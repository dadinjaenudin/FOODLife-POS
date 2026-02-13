# -*- coding: utf-8 -*-
from django.conf import settings
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db.models import Sum, Avg, Q, Count
from django.db.models.deletion import ProtectedError
from datetime import datetime, timedelta
from decimal import Decimal
import logging

from apps.core.models import POSTerminal, Store, Category, Product, User, ProductPhoto, Brand, Company, StoreBrand, MediaGroup, PaymentMethodProfile, DataEntryPrompt, EFTTerminal
from apps.core.models_session import StoreSession
from apps.core.minio_client import get_minio_endpoint_for_request
from apps.pos.models import Bill, Payment, QRISAuditLog, QRISTransaction
from apps.tables.models import Table, TableArea
from apps.promotions.models import Promotion
from .decorators import manager_required, supervisor_required

logger = logging.getLogger(__name__)

def check_store_config(request, template_name):
    """Helper to check if store is configured, return error response if not"""
    store_config = Store.get_current()
    if not store_config:
        return None, render(request, template_name, {
            'error': 'Store configuration not found. Please run setup wizard first.',
            'store_config': None,
        })
    return store_config, None


@manager_required
def dashboard(request):
    """
    Management Dashboard - Real-time metrics & overview
    """
    store_config = Store.get_current()

    if not store_config:
        return render(request, 'management/dashboard.html', {
            'error': 'Store configuration not found. Please run setup wizard first.',
            'store_config': None,
        })

    from apps.core.models import StoreBrand
    if not StoreBrand.objects.filter(store=store_config, is_active=True).exists():
        return render(request, 'management/dashboard.html', {
            'error': 'No active brands found for this store. Please run setup wizard to associate brands.',
            'store_config': store_config,
        })

    today = timezone.now().date()
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True)
    brand_ids = store_brands.values_list('brand_id', flat=True)

    today_bills = Bill.objects.filter(
        created_at__date=today,
        brand_id__in=brand_ids
    )
    closed_bills = today_bills.filter(status='paid')

    today_revenue = closed_bills.aggregate(
        total=Sum('total')
    )['total'] or Decimal('0')

    bills_count = {
        'open': today_bills.filter(status='open').count(),
        'closed': closed_bills.count(),
        'held': today_bills.filter(status='hold').count(),
    }

    avg_bill_value = closed_bills.aggregate(
        avg=Avg('total')
    )['avg'] or Decimal('0')

    payments = Payment.objects.filter(
        bill__created_at__date=today,
        bill__brand_id__in=brand_ids
    ).values('method').annotate(
        total=Sum('amount')
    )

    payment_breakdown = {p['method']: p['total'] for p in payments}

    terminals = POSTerminal.objects.filter(store=store_config, is_active=True)
    five_min_ago = timezone.now() - timedelta(minutes=5)
    online_terminals = terminals.filter(last_heartbeat__gte=five_min_ago).count()
    offline_terminals = terminals.filter(
        Q(last_heartbeat__lt=five_min_ago) | Q(last_heartbeat__isnull=True)
    ).count()

    # Printer health widget (kitchen)
    from apps.kitchen.models import StationPrinter, PrinterHealthCheck
    context_brand_id = request.session.get('context_brand_id')
    if context_brand_id:
        brand_ids = [context_brand_id]

    printers = StationPrinter.objects.filter(brand_id__in=brand_ids, is_active=True).order_by('station_code')
    printer_status_list = []
    online_printers = 0
    offline_printers = 0
    unknown_printers = 0

    for printer in printers:
        latest_health = printer.health_checks.order_by('-checked_at').first()
        if latest_health is None:
            status = 'unknown'
            unknown_printers += 1
        elif latest_health.is_online:
            status = 'online'
            online_printers += 1
        else:
            status = 'offline'
            offline_printers += 1

        printer_status_list.append({
            'printer': printer,
            'status': status,
            'last_checked': latest_health.checked_at if latest_health else None,
        })

    active_cashiers = today_bills.filter(status='open').values(
        'created_by__first_name',
        'created_by__last_name',
        'terminal__terminal_code'
    ).distinct()

    context = {
        'store_config': store_config,
        'current_session': None,
        'business_date': today,
        'hours_open': 0,
        'today_revenue': today_revenue,
        'bills_count': bills_count,
        'avg_bill_value': avg_bill_value,
        'online_terminals': online_terminals,
        'offline_terminals': offline_terminals,
        'total_terminals': terminals.count(),
        'printer_status_list': printer_status_list,
        'printer_status_counts': {
            'total': printers.count(),
            'online': online_printers,
            'offline': offline_printers,
            'unknown': unknown_printers,
        },
        'active_shifts': [],
        'active_cashiers_count': active_cashiers.count(),
        'payment_breakdown': payment_breakdown,
        'last_updated': timezone.now(),
    }

    return render(request, 'management/dashboard.html', context)
@manager_required
def user_create(request):
    """Create New User"""
    store_config = Store.get_current()
    if not store_config:
        return render(request, 'management/user_form.html', {
            'error': 'Store configuration not found. Please run setup wizard first.',
        })
    
    # Get all brands for this store
    from apps.core.models import StoreBrand
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')
    available_brands = [sb.brand for sb in store_brands]

    if request.method == 'GET':
        return render(request, 'management/user_form.html', {
            'is_edit': False,
            'available_brands': available_brands,
        })

    try:
        username = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        role = request.POST.get('role', '').strip()
        role_scope = request.POST.get('role_scope', 'store').strip()
        password = request.POST.get('password', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        # Brand access - can be empty for "All Brands" access
        brand_id = request.POST.get('brand', '').strip()
        selected_brand = None
        if brand_id:
            selected_brand = Brand.objects.filter(id=brand_id).first()

        if not username:
            messages.error(request, 'Username is required')
        elif not first_name:
            messages.error(request, 'First name is required')
        elif not role:
            messages.error(request, 'Role is required')
        elif not password:
            messages.error(request, 'Password is required')
        elif len(password) < 6:
            messages.error(request, 'Password must be at least 6 characters')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists')
        else:
            # Set company - get from selected brand or first brand
            company = selected_brand.company if selected_brand else available_brands[0].company
            
            user = User.objects.create(
                username=username,
                first_name=first_name,
                last_name=last_name,
                email=email,
                role=role,
                role_scope=role_scope,
                brand=selected_brand,  # None if "All Brands" selected
                company=company,
                is_active=is_active,
            )
            user.set_password(password)
            user.save()
            
            brand_access_msg = selected_brand.name if selected_brand else "All Brands"
            messages.success(request, f'User {username} created successfully with access to: {brand_access_msg}')
            return redirect('management:users')

    except Exception as e:
        messages.error(request, str(e))

    return render(request, 'management/user_form.html', {
        'is_edit': False,
        'available_brands': available_brands,
        'user_obj': {
            'username': username,
            'first_name': first_name,
            'last_name': last_name,
            'email': email,
            'role': role,
            'role_scope': role_scope,
            'is_active': is_active,
        }
    })


@manager_required
def user_edit(request, user_id):
    """Edit User"""
    store_config = Store.get_current()
    if not store_config:
        return render(request, 'management/user_form.html', {
            'error': 'Store configuration not found. Please run setup wizard first.',
        })
    
    # Get all brands for this store
    from apps.core.models import StoreBrand
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')
    available_brands = [sb.brand for sb in store_brands]

    user_obj = get_object_or_404(User, id=user_id)

    if request.method == 'GET':
        return render(request, 'management/user_form.html', {
            'is_edit': True,
            'user_obj': user_obj,
            'available_brands': available_brands,
        })

    username = request.POST.get('username', '').strip()
    first_name = request.POST.get('first_name', '').strip()
    last_name = request.POST.get('last_name', '').strip()
    email = request.POST.get('email', '').strip()
    role = request.POST.get('role', '').strip()
    role_scope = request.POST.get('role_scope', 'store').strip()
    password = request.POST.get('password', '').strip()
    is_active = request.POST.get('is_active') == 'on'
    
    # Brand access
    brand_id = request.POST.get('brand', '').strip()
    selected_brand = None
    if brand_id:
        selected_brand = Brand.objects.filter(id=brand_id).first()

    if not username:
        messages.error(request, 'Username is required')
    elif not first_name:
        messages.error(request, 'First name is required')
    elif not role:
        messages.error(request, 'Role is required')
    elif User.objects.filter(username=username).exclude(id=user_obj.id).exists():
        messages.error(request, 'Username already exists')
    else:
        user_obj.username = username
        user_obj.first_name = first_name
        user_obj.last_name = last_name
        user_obj.email = email
        user_obj.role = role
        user_obj.role_scope = role_scope
        user_obj.brand = selected_brand
        user_obj.is_active = is_active
        if password:
            if len(password) < 6:
                messages.error(request, 'Password must be at least 6 characters')
                return render(request, 'management/user_form.html', {
                    'is_edit': True,
                    'user_obj': user_obj,
                })
            user_obj.set_password(password)
        user_obj.save()
        messages.success(request, f'User {username} updated successfully')
        return redirect('management:users')

    return render(request, 'management/user_form.html', {
        'is_edit': True,
        'user_obj': user_obj,
    })


@manager_required
def user_delete(request, user_id):
    """Delete User"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    user_obj = get_object_or_404(User, id=user_id)
    if user_obj.id == request.user.id:
        messages.error(request, 'You cannot delete your own account')
        return redirect('management:users')

    username = user_obj.username
    user_obj.delete()
    messages.success(request, f'User {username} deleted')
    return redirect('management:users')


@manager_required
def terminals_list(request):
    """
    Terminal Management Page - List all terminals with status
    """
    store_config = Store.get_current()
    if not store_config:
        return render(request, 'management/terminals.html', {
            'error': 'Store configuration not found. Please run setup wizard first.',
            'store_config': None,
            'terminals': [],
        })

    base_terminals = POSTerminal.objects.filter(store=store_config)
    
    # Apply context brand filter from session
    context_brand_id = request.session.get('context_brand_id')
    if context_brand_id:
        base_terminals = base_terminals.filter(brand_id=context_brand_id)
    
    # Get all terminals
    terminals = base_terminals.order_by('-is_active', '-last_heartbeat')
    
    # Filter by status
    status_filter = request.GET.get('status', 'all')
    five_min_ago = timezone.now() - timedelta(minutes=5)
    
    if status_filter == 'online':
        terminals = terminals.filter(last_heartbeat__gte=five_min_ago, is_active=True)
    elif status_filter == 'offline':
        terminals = terminals.filter(
            Q(last_heartbeat__lt=five_min_ago) | Q(last_heartbeat__isnull=True),
            is_active=True
        )
    elif status_filter == 'inactive':
        terminals = terminals.filter(is_active=False)
    
    # Filter by device type
    device_filter = request.GET.get('device_type', '')
    if device_filter:
        terminals = terminals.filter(device_type=device_filter)
    
    # Search by terminal code
    search = request.GET.get('search', '')
    if search:
        terminals = terminals.filter(terminal_code__icontains=search)
    
    # Get current shifts per terminal (simplified for now)
    terminal_ids = [t.id for t in terminals]
    
    # Add shift info to terminals
    for terminal in terminals:
        terminal.current_shift = None  # Will implement with CashierShift model
        
        # Calculate online status
        if terminal.is_active and terminal.last_heartbeat:
            time_diff = timezone.now() - terminal.last_heartbeat
            if time_diff < timedelta(minutes=5):
                terminal.status = 'online'
            elif time_diff < timedelta(minutes=10):
                terminal.status = 'warning'
            else:
                terminal.status = 'offline'
        else:
            terminal.status = 'inactive' if not terminal.is_active else 'offline'
    
    total_count = base_terminals.count()
    active_count = base_terminals.filter(is_active=True).count()
    inactive_count = base_terminals.filter(is_active=False).count()
    online_count = base_terminals.filter(last_heartbeat__gte=five_min_ago, is_active=True).count()
    offline_count = base_terminals.filter(
        Q(last_heartbeat__lt=five_min_ago) | Q(last_heartbeat__isnull=True),
        is_active=True
    ).count()

    context = {
        'store_config': store_config,
        'terminals': terminals,
        'status_filter': status_filter,
        'device_filter': device_filter,
        'search': search,
        'device_types': POSTerminal.DEVICE_TYPE_CHOICES,
        'total_count': total_count,
        'active_count': active_count,
        'inactive_count': inactive_count,
        'online_count': online_count,
        'offline_count': offline_count,
    }
    
    return render(request, 'management/terminals.html', context)


@manager_required
def terminal_detail(request, terminal_id):
    """
    Terminal Details Page
    """
    terminal = get_object_or_404(POSTerminal, id=terminal_id)
    
    # Get recent activity (simplified for now)
    recent_shifts = []  # Will implement with CashierShift model
    
    recent_bills = Bill.objects.filter(
        terminal=terminal
    ).order_by('-created_at')[:10]
    
    # Total stats
    total_bills = Bill.objects.filter(terminal=terminal).count()
    total_sales = Bill.objects.filter(
        terminal=terminal,
        status='paid'
    ).aggregate(total=Sum('total'))['total'] or Decimal('0')
    
    context = {
        'terminal': terminal,
        'recent_shifts': recent_shifts,
        'recent_bills': recent_bills,
        'total_bills': total_bills,
        'total_sales': total_sales,
    }
    
    return render(request, 'management/terminal_detail_page.html', context)


@supervisor_required
def terminal_deactivate(request, terminal_id):
    """
    Deactivate Terminal (HTMX endpoint)
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    terminal = get_object_or_404(POSTerminal, id=terminal_id)
    reason = request.POST.get('reason', 'No reason provided')
    
    # Deactivate terminal
    terminal.is_active = False
    terminal.save()
    
    # Set status for display
    terminal.status = 'inactive'
    terminal.current_shift = None
    
    # TODO: Create audit log
    # TODO: Force close active shift if exists
    
    # Return updated table row HTML
    from django.template.loader import render_to_string
    html = render_to_string('management/partials/terminal_row.html', {
        'terminal': terminal,
    })
    return render(request, 'management/partials/terminal_row.html', {'terminal': terminal})


@supervisor_required
def terminal_reactivate(request, terminal_id):
    """
    Reactivate Terminal (HTMX endpoint)
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    terminal = get_object_or_404(POSTerminal, id=terminal_id)
    
    terminal.is_active = True
    terminal.save()
    
    # Recalculate status
    if terminal.last_heartbeat:
        time_diff = timezone.now() - terminal.last_heartbeat
        from datetime import timedelta
        if time_diff < timedelta(minutes=5):
            terminal.status = 'online'
        elif time_diff < timedelta(minutes=10):
            terminal.status = 'warning'
        else:
            terminal.status = 'offline'
    else:
        terminal.status = 'offline'
    
    terminal.current_shift = None
    
    # Return updated table row HTML
    return render(request, 'management/partials/terminal_row.html', {'terminal': terminal})


@supervisor_required
def terminal_delete(request, terminal_id):
    """
    Delete Terminal
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    terminal = get_object_or_404(POSTerminal, id=terminal_id)
    terminal_code = terminal.terminal_code
    
    # Try to delete the terminal
    try:
        terminal.delete()
        messages.success(request, f'Terminal {terminal_code} has been deleted successfully')
    except ProtectedError as e:
        # Handle protected foreign key error
        messages.error(request, 
            f'Cannot delete terminal {terminal_code} because it has related records (shifts, bills, etc). '
            'Please deactivate the terminal instead of deleting it.')
    except Exception as e:
        messages.error(request, f'Error deleting terminal {terminal_code}: {str(e)}')
    
    return redirect('management:terminals')


@manager_required
def terminal_create(request):
    """Create Terminal with Full Configuration"""
    store_config = Store.get_current()
    if not store_config:
        return render(request, 'management/terminal_form.html', {
            'error': 'Store configuration not found. Please run setup wizard first.',
        })

    from apps.core.models import StoreBrand, Brand
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')

    if request.method == 'POST':
        # Core fields
        terminal_code = request.POST.get('terminal_code', '').strip()
        terminal_name = request.POST.get('terminal_name', '').strip()
        device_type = request.POST.get('device_type', '').strip()
        brand_id = request.POST.get('brand')
        is_active = request.POST.get('is_active') == '1'

        # Validation
        if not terminal_code:
            messages.error(request, 'Terminal code is required')
        elif not terminal_name:
            messages.error(request, 'Terminal name is required')
        elif device_type not in dict(POSTerminal.DEVICE_TYPE_CHOICES):
            messages.error(request, 'Invalid device type')
        elif POSTerminal.objects.filter(terminal_code=terminal_code).exists():
            messages.error(request, 'Terminal code already exists')
        else:
            # Get brand
            if brand_id:
                brand = get_object_or_404(Brand, id=brand_id)
            else:
                brand = store_config.brand

            # Create terminal with core fields
            terminal = POSTerminal(
                store=store_config,
                brand=brand,
                terminal_code=terminal_code,
                terminal_name=terminal_name,
                device_type=device_type,
                is_active=is_active,
                registered_by=request.user,
            )

            # Network configuration
            ip_address = request.POST.get('ip_address', '').strip()
            if ip_address:
                terminal.ip_address = ip_address

            # Printer configuration
            printer_type = request.POST.get('printer_type', '').strip()
            terminal.printer_type = printer_type if printer_type else 'thermal'
            terminal.receipt_printer_name = request.POST.get('receipt_printer_name', '').strip()
            terminal.receipt_paper_width = int(request.POST.get('receipt_paper_width', 80))
            terminal.kitchen_printer_name = request.POST.get('kitchen_printer_name', '').strip()
            terminal.print_to = request.POST.get('print_to', 'printer')
            terminal.auto_print_receipt = request.POST.get('auto_print_receipt') == '1'
            terminal.auto_print_kitchen_order = request.POST.get('auto_print_kitchen_order') == '1'
            terminal.print_logo_on_receipt = request.POST.get('print_logo_on_receipt') == '1'
            terminal.print_checker_receipt = request.POST.get('print_checker_receipt') == '1'

            # Hardware integration
            terminal.cash_drawer_enabled = request.POST.get('cash_drawer_enabled') == '1'
            terminal.barcode_scanner_enabled = request.POST.get('barcode_scanner_enabled') == '1'
            terminal.customer_pole_display_enabled = request.POST.get('customer_pole_display_enabled') == '1'

            # Display configuration
            terminal.enable_customer_display = request.POST.get('enable_customer_display') == '1'
            terminal.enable_kitchen_display = request.POST.get('enable_kitchen_display') == '1'
            terminal.enable_kitchen_printer = request.POST.get('enable_kitchen_printer') == '1'

            # Payment configuration
            edc_mode = request.POST.get('edc_integration_mode', '').strip()
            terminal.edc_integration_mode = edc_mode if edc_mode else 'none'
            
            # Payment profiles (M2M) + dual-write legacy
            terminal.save()
            profile_ids = request.POST.getlist('payment_profile_ids')
            if profile_ids:
                profiles = PaymentMethodProfile.objects.filter(id__in=profile_ids)
                terminal.payment_profiles.set(profiles)
                legacy_methods = list(profiles.exclude(legacy_method_id='').values_list('legacy_method_id', flat=True))
                terminal.default_payment_methods = legacy_methods
            else:
                terminal.payment_profiles.clear()
                terminal.default_payment_methods = []
            terminal.save(update_fields=['default_payment_methods'])

            messages.success(request, f'Terminal {terminal.terminal_code} created successfully')
            return redirect('management:terminals')

    # Load payment profiles for the brand
    available_profiles = PaymentMethodProfile.objects.filter(
        brand=store_config.brand, is_active=True
    ).order_by('sort_order', 'name')

    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'device_types': POSTerminal.DEVICE_TYPE_CHOICES,
        'printer_types': POSTerminal.PRINTER_TYPE_CHOICES,
        'print_to_choices': POSTerminal.PRINT_TO_CHOICES,
        'edc_modes': POSTerminal.EDC_MODE_CHOICES,
        'payment_profiles': available_profiles,
        'assigned_profile_ids': [],
        'is_edit': False,
    }
    return render(request, 'management/terminal_form.html', context)


@manager_required
def terminal_edit(request, terminal_id):
    """Edit Terminal with Full Configuration"""
    store_config = Store.get_current()
    if not store_config:
        return render(request, 'management/terminal_form.html', {
            'error': 'Store configuration not found. Please run setup wizard first.',
        })

    terminal = get_object_or_404(POSTerminal, id=terminal_id)
    from apps.core.models import StoreBrand, Brand
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')

    if request.method == 'POST':
        # Core fields
        terminal_code = request.POST.get('terminal_code', '').strip()
        terminal_name = request.POST.get('terminal_name', '').strip()
        device_type = request.POST.get('device_type', '').strip()
        brand_id = request.POST.get('brand')
        is_active = request.POST.get('is_active') == '1'

        # Validation
        if not terminal_code:
            messages.error(request, 'Terminal code is required')
        elif not terminal_name:
            messages.error(request, 'Terminal name is required')
        elif device_type not in dict(POSTerminal.DEVICE_TYPE_CHOICES):
            messages.error(request, 'Invalid device type')
        elif POSTerminal.objects.filter(terminal_code=terminal_code).exclude(id=terminal.id).exists():
            messages.error(request, 'Terminal code already exists')
        else:
            # Get brand
            if brand_id:
                brand = get_object_or_404(Brand, id=brand_id)
            else:
                brand = store_config.brand

            # Update core fields
            terminal.terminal_code = terminal_code
            terminal.terminal_name = terminal_name
            terminal.device_type = device_type
            terminal.brand = brand
            terminal.is_active = is_active

            # Network configuration
            ip_address = request.POST.get('ip_address', '').strip()
            if ip_address:
                terminal.ip_address = ip_address

            # Printer configuration
            printer_type = request.POST.get('printer_type', '').strip()
            terminal.printer_type = printer_type if printer_type else 'thermal'
            terminal.receipt_printer_name = request.POST.get('receipt_printer_name', '').strip()
            terminal.receipt_paper_width = int(request.POST.get('receipt_paper_width', 80))
            terminal.kitchen_printer_name = request.POST.get('kitchen_printer_name', '').strip()
            terminal.print_to = request.POST.get('print_to', 'printer')
            terminal.auto_print_receipt = request.POST.get('auto_print_receipt') == '1'
            terminal.auto_print_kitchen_order = request.POST.get('auto_print_kitchen_order') == '1'
            terminal.print_logo_on_receipt = request.POST.get('print_logo_on_receipt') == '1'
            terminal.print_checker_receipt = request.POST.get('print_checker_receipt') == '1'

            # Hardware integration
            terminal.cash_drawer_enabled = request.POST.get('cash_drawer_enabled') == '1'
            terminal.barcode_scanner_enabled = request.POST.get('barcode_scanner_enabled') == '1'
            terminal.customer_pole_display_enabled = request.POST.get('customer_pole_display_enabled') == '1'

            # Display configuration
            terminal.enable_customer_display = request.POST.get('enable_customer_display') == '1'
            terminal.enable_kitchen_display = request.POST.get('enable_kitchen_display') == '1'
            terminal.enable_kitchen_printer = request.POST.get('enable_kitchen_printer') == '1'

            # Payment configuration
            edc_mode = request.POST.get('edc_integration_mode', '').strip()
            terminal.edc_integration_mode = edc_mode if edc_mode else 'none'
            
            # Payment profiles (M2M) + dual-write legacy
            terminal.save()
            profile_ids = request.POST.getlist('payment_profile_ids')
            if profile_ids:
                profiles = PaymentMethodProfile.objects.filter(id__in=profile_ids)
                terminal.payment_profiles.set(profiles)
                legacy_methods = list(profiles.exclude(legacy_method_id='').values_list('legacy_method_id', flat=True))
                terminal.default_payment_methods = legacy_methods
            else:
                terminal.payment_profiles.clear()
                terminal.default_payment_methods = []
            terminal.save(update_fields=['default_payment_methods'])

            messages.success(request, f'Terminal {terminal.terminal_code} updated successfully')
            return redirect('management:terminals')

    # Load payment profiles for the terminal's brand
    available_profiles = PaymentMethodProfile.objects.filter(
        brand=terminal.brand, is_active=True
    ).order_by('sort_order', 'name')
    assigned_profile_ids = set(terminal.payment_profiles.values_list('id', flat=True))

    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'device_types': POSTerminal.DEVICE_TYPE_CHOICES,
        'printer_types': POSTerminal.PRINTER_TYPE_CHOICES,
        'print_to_choices': POSTerminal.PRINT_TO_CHOICES,
        'edc_modes': POSTerminal.EDC_MODE_CHOICES,
        'payment_profiles': available_profiles,
        'assigned_profile_ids': assigned_profile_ids,
        'terminal': terminal,
        'is_edit': True,
    }
    return render(request, 'management/terminal_form.html', context)


@manager_required
def terminal_duplicate(request, terminal_id):
    """Duplicate Terminal - Copy configuration from existing terminal"""
    store_config = Store.get_current()
    if not store_config:
        return render(request, 'management/terminal_form.html', {
            'error': 'Store configuration not found. Please run setup wizard first.',
        })

    # Get the terminal to duplicate
    source_terminal = get_object_or_404(POSTerminal, id=terminal_id)
    
    from apps.core.models import StoreBrand, Brand
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')

    # If this is a POST request, create the new terminal
    if request.method == 'POST':
        # Core fields
        terminal_code = request.POST.get('terminal_code', '').strip()
        terminal_name = request.POST.get('terminal_name', '').strip()
        device_type = request.POST.get('device_type', '').strip()
        brand_id = request.POST.get('brand')
        is_active = request.POST.get('is_active') == '1'

        # Validation
        if not terminal_code:
            messages.error(request, 'Terminal code is required')
        elif not terminal_name:
            messages.error(request, 'Terminal name is required')
        elif device_type not in dict(POSTerminal.DEVICE_TYPE_CHOICES):
            messages.error(request, 'Invalid device type')
        elif POSTerminal.objects.filter(terminal_code=terminal_code).exists():
            messages.error(request, 'Terminal code already exists')
        else:
            # Get brand
            if brand_id:
                brand = get_object_or_404(Brand, id=brand_id)
            else:
                brand = store_config.brand

            # Create terminal with core fields
            terminal = POSTerminal(
                store=store_config,
                brand=brand,
                terminal_code=terminal_code,
                terminal_name=terminal_name,
                device_type=device_type,
                is_active=is_active,
                registered_by=request.user,
            )

            # Network configuration
            ip_address = request.POST.get('ip_address', '').strip()
            if ip_address:
                terminal.ip_address = ip_address

            # Printer configuration
            printer_type = request.POST.get('printer_type', '').strip()
            terminal.printer_type = printer_type if printer_type else 'thermal'
            terminal.receipt_printer_name = request.POST.get('receipt_printer_name', '').strip()
            terminal.receipt_paper_width = int(request.POST.get('receipt_paper_width', 80))
            terminal.kitchen_printer_name = request.POST.get('kitchen_printer_name', '').strip()
            terminal.print_to = request.POST.get('print_to', 'printer')
            terminal.auto_print_receipt = request.POST.get('auto_print_receipt') == '1'
            terminal.auto_print_kitchen_order = request.POST.get('auto_print_kitchen_order') == '1'
            terminal.print_logo_on_receipt = request.POST.get('print_logo_on_receipt') == '1'
            terminal.print_checker_receipt = request.POST.get('print_checker_receipt') == '1'

            # Hardware integration
            terminal.cash_drawer_enabled = request.POST.get('cash_drawer_enabled') == '1'
            terminal.barcode_scanner_enabled = request.POST.get('barcode_scanner_enabled') == '1'
            terminal.customer_pole_display_enabled = request.POST.get('customer_pole_display_enabled') == '1'

            # Display configuration
            terminal.enable_customer_display = request.POST.get('enable_customer_display') == '1'
            terminal.enable_kitchen_display = request.POST.get('enable_kitchen_display') == '1'
            terminal.enable_kitchen_printer = request.POST.get('enable_kitchen_printer') == '1'

            # Payment configuration
            edc_mode = request.POST.get('edc_integration_mode', '').strip()
            terminal.edc_integration_mode = edc_mode if edc_mode else 'none'
            
            # Payment profiles (M2M) + dual-write legacy
            terminal.save()
            profile_ids = request.POST.getlist('payment_profile_ids')
            if profile_ids:
                profiles = PaymentMethodProfile.objects.filter(id__in=profile_ids)
                terminal.payment_profiles.set(profiles)
                legacy_methods = list(profiles.exclude(legacy_method_id='').values_list('legacy_method_id', flat=True))
                terminal.default_payment_methods = legacy_methods
            else:
                terminal.payment_profiles.clear()
                terminal.default_payment_methods = []
            terminal.save(update_fields=['default_payment_methods'])

            messages.success(request, f'Terminal {terminal.terminal_code} created successfully from duplicate')
            return redirect('management:terminals')

    # Generate suggested terminal code (add -copy or increment number)
    import re
    base_code = source_terminal.terminal_code

    # Try to find a unique code
    counter = 2
    suggested_code = f"{base_code}-copy"
    while POSTerminal.objects.filter(terminal_code=suggested_code).exists():
        suggested_code = f"{base_code}-{counter}"
        counter += 1

    # Create a copy of the terminal object for the form (without saving)
    from copy import copy
    terminal_copy = copy(source_terminal)
    terminal_copy.id = None  # Clear ID so it's treated as new
    terminal_copy.terminal_code = suggested_code
    terminal_copy.terminal_name = f"{source_terminal.terminal_name} (Copy)"

    # Load payment profiles and pre-select from source terminal
    available_profiles = PaymentMethodProfile.objects.filter(
        brand=source_terminal.brand, is_active=True
    ).order_by('sort_order', 'name')
    assigned_profile_ids = set(source_terminal.payment_profiles.values_list('id', flat=True))

    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'device_types': POSTerminal.DEVICE_TYPE_CHOICES,
        'printer_types': POSTerminal.PRINTER_TYPE_CHOICES,
        'print_to_choices': POSTerminal.PRINT_TO_CHOICES,
        'edc_modes': POSTerminal.EDC_MODE_CHOICES,
        'payment_profiles': available_profiles,
        'assigned_profile_ids': assigned_profile_ids,
        'terminal': terminal_copy,
        'is_edit': False,
        'is_duplicate': True,
        'source_terminal_code': source_terminal.terminal_code,
    }
    return render(request, 'management/terminal_form.html', context)


@manager_required
def dashboard_refresh(request):
    """
    Dashboard Metrics Refresh (HTMX partial for auto-update)
    """
    # Reuse dashboard logic but return only the metrics partial
    # This will be called every 30 seconds via HTMX polling
    
    store_config = Store.get_current()
    current_session = StoreSession.get_current(store_config) if store_config else None
    
    # Similar logic as dashboard view...
    # (abbreviated for now, will extract to service function)
    
    context = {
        'last_updated': timezone.now(),
        # ... other metrics
    }
    
    return render(request, 'management/partials/dashboard_metrics.html', context)


@manager_required
def settings(request):
    """
    Settings Page - Store Configuration
    """
    print(f"[DEBUG] settings() called - Method: {request.method}, Path: {request.path}")
    print(f"[DEBUG] Headers: {dict(request.headers)}")
    
    store_config = Store.get_current()
    
    if not store_config:
        return render(request, 'management/settings.html', {
            'error': 'Store configuration not found. Please run setup wizard first.',
        })
    
    brand = store_config.brand
    company = brand.company
    
    # Get terminal count
    terminal_count = POSTerminal.objects.filter(store=store_config).count()
    
    # Get database size (SQLite)
    import os
    from django.conf import settings
    db_size = 0
    if settings.DATABASES['default']['ENGINE'] == 'django.db.backends.sqlite3':
        db_path = settings.DATABASES['default']['NAME']
        if os.path.exists(db_path):
            db_size = os.path.getsize(db_path) / (1024 * 1024)  # MB
    
    context = {
        'store_config': store_config,
        'brand': brand,
        'company': company,
        'terminal_count': terminal_count,
        'db_size': db_size,
    }
    
    return render(request, 'management/settings.html', context)


@manager_required
def settings_update(request):
    """
    Update Settings - HTMX form handler with file upload
    """
    print(f"[DEBUG] settings_update() called - Method: {request.method}")
    print(f"[DEBUG] Headers: {dict(request.headers)}")
    print(f"[DEBUG] POST data: {dict(request.POST)}")
    print(f"[DEBUG] FILES: {dict(request.FILES)}")
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    store_config = Store.get_current()
    if not store_config:
        return JsonResponse({'error': 'Store configuration not found'}, status=404)
    
    try:
        # Update Store (only editable fields: address, phone, login_image)
        # store_code and store_name are read-only, set from /setup/
        store_config.address = request.POST.get('address', store_config.address)
        store_config.phone = request.POST.get('phone', store_config.phone)
        
        # Handle login image upload
        if 'login_image' in request.FILES:
            login_image = request.FILES['login_image']
            # Validate file size (5MB)
            if login_image.size > 5 * 1024 * 1024:
                return JsonResponse({'error': 'File size must be less than 5MB'}, status=400)
            
            # Delete old image if exists
            if store_config.login_image:
                store_config.login_image.delete(save=False)
            
            store_config.login_image = login_image
        
        # Handle remove image
        if request.POST.get('remove_login_image') == 'true':
            if store_config.login_image:
                store_config.login_image.delete(save=False)
                store_config.login_image = None
        
        store_config.save()
        
        # Update Brand (tax & service charge)
        brand = store_config.brand
        brand.tax_rate = Decimal(request.POST.get('tax_rate', brand.tax_rate))
        brand.service_charge = Decimal(request.POST.get('service_charge', brand.service_charge))
        brand.receipt_footer = request.POST.get('receipt_footer', brand.receipt_footer)
        brand.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Settings updated successfully'
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'error': str(e)
        }, status=400)


@manager_required
def master_data(request):
    """
    Master Data Overview - Show all master data counts
    Identifies what data needs to sync with HO/Cloud
    """
    print(f"[DEBUG] master_data() called - Method: {request.method}")
    
    store_config = Store.get_current()
    if not store_config:
        return render(request, 'management/master_data.html', {
            'error': 'Store configuration not found. Please run setup wizard first.',
        })
    
    # Get all brands associated with this store
    from apps.core.models import StoreBrand
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True)
    brand_ids = list(store_brands.values_list('brand_id', flat=True))
    
    # Get global context brand filter from session
    context_brand_id = request.session.get('context_brand_id', '')
    
    # Apply global brand filter if set
    if context_brand_id and context_brand_id in [str(bid) for bid in brand_ids]:
        filtered_brand_ids = [context_brand_id]
        current_brand = Brand.objects.get(id=context_brand_id)
    else:
        filtered_brand_ids = brand_ids
        current_brand = store_brands.first().brand if store_brands.exists() else None
    
    company = current_brand.company if current_brand else None
    
    # Count product images (filtered by selected brands)
    edge_images_count = ProductPhoto.objects.filter(product__brand_id__in=filtered_brand_ids).count()
    last_image_sync = ProductPhoto.objects.filter(
        product__brand_id__in=filtered_brand_ids,
        last_sync_at__isnull=False
    ).order_by('-last_sync_at').first()
    
    # Count all master data
    master_data_summary = [
        {
            'name': 'Company (Tenant)',
            'table': 'core_company',
            'count': 1,  # Always 1 per Edge Server
            'sync': 'pull',  # Download from HO
            'description': 'Company/Tenant configuration',
            'icon': '🏢',
            'url': None,  # Read-only
        },
        {
            'name': 'Brand (Brand)',
            'table': 'core_brand',
            'count': Brand.objects.all().count(),
            'sync': 'pull',
            'description': 'Brand/Brand configuration',
            'icon': '🏷️',
            'url': 'management:brands',
        },
        {
            'name': 'Store Configuration',
            'table': 'core_Store',
            'count': 1,  # Singleton
            'sync': 'pull',
            'description': 'This Edge Server configuration',
            'icon': '🏪',
            'url': 'management:settings',
        },
        {
            'name': 'Master Data',
            'table': 'core_product',
            'count': Product.objects.filter(brand_id__in=filtered_brand_ids).count(),
            'sync': 'pull',
            'description': 'Categories, Products, Modifiers, Modifier Options, Product-Modifier Links',
            'icon': '🍽️',
            'url': 'management:products',
        },
        {
            'name': 'Tables',
            'table': 'tables_table',
            'count': Table.objects.filter(area__brand_id__in=filtered_brand_ids).count(),
            'sync': 'pull',
            'description': 'Tables, Table Areas, Table Groups',
            'icon': '🪑',
            'url': 'management:tables',
        },
        {
            'name': 'Promotions',
            'table': 'promotions_promotion',
            'count': Promotion.objects.filter(
                brand_id__in=filtered_brand_ids,
                is_active=True,
            ).filter(
                Q(store=store_config) | Q(store__isnull=True)
            ).count(),
            'sync': 'pull',
            'description': 'Active promotions & discounts (compiled from HO)',
            'icon': '🎁',
            'url': 'management:promotions',
        },
    ]
    
    # Transaction data (push to HO)
    transaction_data = [
        {
            'name': 'Bills (Transactions)',
            'table': 'pos_bill',
            'count': Bill.objects.filter(brand_id__in=filtered_brand_ids).count(),
            'sync': 'push',
            'description': 'All sales transactions',
            'icon': '💵',
            'url': 'management:bills',
        },
        {
            'name': 'Payments',
            'table': 'pos_payment',
            'count': Payment.objects.filter(bill__brand_id__in=filtered_brand_ids).count(),
            'sync': 'push',
            'description': 'Payment records',
            'icon': '💳',
            'url': 'management:payments',
        },
        {
            'name': 'Store Sessions',
            'table': 'core_storesession',
            'count': StoreSession.objects.filter(store=store_config).count(),
            'sync': 'push',
            'description': 'Daily business sessions & EOD',
            'icon': '??',
            'url': 'management:store_sessions',
        },
    ]
    
    context = {
        'store_config': store_config,
        'brand': current_brand,  # current_brand instance for template
        'company': company,
        'master_data_summary': master_data_summary,
        'transaction_data': transaction_data,
        'total_master_tables': len(master_data_summary),
        'total_transaction_tables': len(transaction_data),
        'edge_images_count': edge_images_count,
        'ho_images_count': 0,  # Will be fetched from HO API
        'last_image_sync': last_image_sync.last_sync_at if last_image_sync else None,
    }
    
    return render(request, 'management/master_data.html', context)


@csrf_exempt
@require_POST
@manager_required
def sync_product_images(request):
    """
    Sync product images from HO MinIO to Edge MinIO
    """
    from apps.core.services_photo_sync import ProductPhotoSyncService
    
    try:
        store_config = Store.get_current()
        if not store_config:
            return JsonResponse({
                'success': False,
                'message': 'Store configuration not found'
            }, status=400)
        
        # Get active StoreBrand to retrieve ho_store_id
        from apps.core.models import StoreBrand
        store_brand = StoreBrand.objects.filter(store=store_config, is_active=True).first()
        if not store_brand:
            return JsonResponse({
                'success': False,
                'message': 'No active brand found for this store'
            }, status=400)
        
        brand = store_brand.brand
        company = brand.company
        ho_store_id = str(store_brand.ho_store_id)  # Use HO store ID, not Edge store ID
        
        # Initialize sync service
        sync_service = ProductPhotoSyncService()
        
        # Sync photos
        logger.info(f"Starting photo sync for company={company.id}, brand={brand.id}, ho_store_id={ho_store_id}")
        result = sync_service.sync_photos(
            company_id=str(company.id),
            brand_id=str(brand.id),
            store_id=ho_store_id,  # Send HO store ID to HO API
            limit=50  # Sync 50 photos at a time
        )
        
        if result['success']:
            return JsonResponse({
                'success': True,
                'message': f"Synced {result['synced_count']} images successfully",
                'synced_count': result['synced_count'],
                'skipped_count': result['skipped_count'],
                'failed_count': result['failed_count'],
                'total_size': result['total_size'],
            })
        else:
            return JsonResponse({
                'success': False,
                'message': result.get('error', 'Unknown error'),
                'synced_count': result.get('synced_count', 0),
                'skipped_count': result.get('skipped_count', 0),
                'failed_count': result.get('failed_count', 0),
            }, status=500)
        
    except Exception as e:
        logger.error(f"Error syncing product images: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@csrf_exempt
@require_POST
def configure_bucket_policy(request):
    """
    Configure MinIO bucket policy for public read access (product images).
    Creates bucket if not exists and sets public GetObject policy.
    """
    import json as json_lib
    from minio import Minio
    from minio.error import S3Error
    from django.conf import settings as django_settings
    
    # Check authentication - return JSON for AJAX requests
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'Authentication required. Please login.'}, status=401)
    
    try:
        endpoint = django_settings.EDGE_MINIO_ENDPOINT
        access_key = django_settings.EDGE_MINIO_ACCESS_KEY
        secret_key = django_settings.EDGE_MINIO_SECRET_KEY
        secure = django_settings.EDGE_MINIO_SECURE
        bucket_name = 'product-images'
        
        logger.info(f"Configuring bucket policy - endpoint: {endpoint}")
        client = Minio(endpoint, access_key=access_key, secret_key=secret_key, secure=secure)
        
        # Create bucket if not exists
        bucket_created = False
        if not client.bucket_exists(bucket_name):
            client.make_bucket(bucket_name)
            bucket_created = True
            logger.info(f"Created bucket: {bucket_name}")
        
        # Set public read policy
        policy = {
            "Version": "2012-10-17",
            "Statement": [
                {
                    "Effect": "Allow",
                    "Principal": {"AWS": "*"},
                    "Action": ["s3:GetObject"],
                    "Resource": [f"arn:aws:s3:::{bucket_name}/*"]
                }
            ]
        }
        client.set_bucket_policy(bucket_name, json_lib.dumps(policy))
        logger.info(f"Set public read policy for bucket: {bucket_name}")
        
        # Verify by reading back
        current_policy = client.get_bucket_policy(bucket_name)
        policy_set = 's3:GetObject' in current_policy
        
        return JsonResponse({
            'success': True,
            'message': f"Bucket '{bucket_name}' configured successfully!",
            'bucket_created': bucket_created,
            'policy_set': policy_set,
            'endpoint': endpoint,
        })
        
    except S3Error as e:
        logger.error(f"MinIO S3 error configuring bucket: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'MinIO error: {str(e)}'
        }, status=500)
    except Exception as e:
        logger.error(f"Error configuring bucket policy: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Connection error: {str(e)}. Make sure MinIO is running.'
        }, status=500)


@csrf_exempt
@require_POST
@manager_required
def sync_from_ho(request):
    """
    Sync selected master data tables from HO Server using HOAPIClient
    """
    import json
    from apps.core.ho_api import HOAPIClient, HOAPIException
    from decimal import Decimal
    
    try:
        # Parse request body
        body = json.loads(request.body)
        selected_tables = body.get('tables', [])
        
        if not selected_tables:
            return JsonResponse({
                'success': False,
                'error': 'No tables selected'
            }, status=400)
        
        # Get store config
        store_config = Store.get_current()
        if not store_config:
            return JsonResponse({
                'success': False,
                'error': 'Store not configured'
            }, status=400)
        
        from apps.core.models import StoreBrand
        store_brand = StoreBrand.objects.filter(store=store_config, is_active=True).first()
        if not store_brand:
            return JsonResponse({
                'success': False,
                'error': 'No active brand found for this store'
            }, status=400)
        
        brand = store_brand.brand
        company = brand.company
        company_id = str(company.id)
        
        # Validate HO Store ID is configured
        if not store_brand.ho_store_id:
            return JsonResponse({
                'success': False,
                'error': 'HO Store ID not configured. Please run Setup Store wizard first or update StoreBrand.ho_store_id in database.'
            }, status=400)
        
        ho_store_id = str(store_brand.ho_store_id)
        
        # Initialize HO API Client
        client = HOAPIClient()
        
        synced_tables = []
        sync_results = {}
        
        # Sync each selected table
        for table_name in selected_tables:
            try:
                if table_name == 'core_brand':
                    # Sync Brands
                    print(f"[SYNC] Requesting brands from HO with payload: company_id={company_id}, store_id={ho_store_id}")
                    brands = client.get_brands(company_id=company_id, store_id=ho_store_id)
                    saved_count = 0
                    updated_count = 0
                    
                    # Log all brands and their brand_type
                    print(f"[SYNC] Total brands from HO: {len(brands)}")
                    for idx, b in enumerate(brands, 1):
                        print(f"[SYNC] Brand {idx}: code='{b.get('code')}' name='{b.get('name')}' brand_type='{b.get('brand_type')}'")
                    
                    # Get all brand IDs from HO response
                    ho_brand_ids = [brand_data['id'] for brand_data in brands]
                    
                    # Handle brands not in HO response - use soft delete if they have transactions
                    brands_not_in_ho = Brand.objects.filter(
                        company_id=company_id
                    ).exclude(
                        id__in=ho_brand_ids
                    )
                    
                    soft_deleted_count = 0
                    hard_deleted_count = 0
                    
                    for brand in brands_not_in_ho:
                        # Check if brand has any transactions
                        from apps.pos.models import Bill, BillItem
                        has_transactions = (
                            Bill.objects.filter(brand=brand).exists() or
                            BillItem.objects.filter(brand=brand).exists()
                        )
                        
                        if has_transactions:
                            # Soft delete - set is_active=False to preserve data integrity
                            brand.is_active = False
                            brand.save(update_fields=['is_active'])
                            soft_deleted_count += 1
                            print(f"[SYNC] ⚠️ Brand '{brand.name}' has transactions - soft deleted (is_active=False)")
                        else:
                            # Hard delete - safe to remove as no transactions exist
                            brand_name = brand.name
                            brand.delete()
                            hard_deleted_count += 1
                            print(f"[SYNC] 🗑️ Brand '{brand_name}' has no transactions - hard deleted")
                    
                    deleted_count = soft_deleted_count + hard_deleted_count
                    if deleted_count > 0:
                        print(f"[SYNC] Brands not in HO: {soft_deleted_count} soft-deleted, {hard_deleted_count} hard-deleted")
                    
                    for brand_data in brands:
                        try:
                            # Get company_id from API response
                            brand_company_id = brand_data.get('company_id')
                            if not brand_company_id:
                                print(f"[SYNC] Brand {brand_data.get('name')} skipped - missing company_id")
                                continue
                            
                            brand_company = Company.objects.filter(id=brand_company_id).first()
                            if not brand_company:
                                print(f"[SYNC] Brand {brand_data.get('name')} skipped - company not found: {brand_company_id}")
                                continue
                            
                            defaults = {
                                'company': brand_company,
                                'code': brand_data.get('code', ''),
                                'name': brand_data.get('name', 'Unnamed Brand'),
                                'address': brand_data.get('address', ''),
                                'phone': brand_data.get('phone', ''),
                                'tax_id': brand_data.get('tax_id', ''),
                                'tax_rate': Decimal(str(brand_data.get('tax_rate', 11.00))),
                                'service_charge': Decimal(str(brand_data.get('service_charge', 5.00))),
                                'receipt_footer': brand_data.get('receipt_footer', 'Terima Kasih Atas Kunjungan Anda'),
                                'brand_type': brand_data.get('brand_type', 'restaurant'),  # Get brand_type from HO API
                                'is_active': brand_data.get('is_active', True),
                                'point_expiry_months_override': brand_data.get('point_expiry_months_override'),
                            }
                            
                            brand_obj, created = Brand.objects.update_or_create(
                                id=brand_data['id'],
                                defaults=defaults
                            )
                            
                            if created:
                                saved_count += 1
                            else:
                                updated_count += 1
                                
                        except Exception as brand_error:
                            print(f"Error saving brand {brand_data.get('name', 'unknown')}: {brand_error}")
                            import traceback
                            traceback.print_exc()
                            continue
                    
                    synced_tables.append(table_name)
                    sync_results[table_name] = {
                        'success': True,
                        'records_count': saved_count + updated_count,
                        'details': f'{saved_count} created, {updated_count} updated'
                    }
                
                elif table_name == 'core_storebrand':
                    # Sync Store-Brand relationships
                    print(f"[SYNC] Requesting store-brands from HO with payload: company_id={company_id}, store_id={ho_store_id}")
                    store_brands_data = client.get_store_brands(company_id=company_id, store_id=ho_store_id)
                    saved_count = 0
                    updated_count = 0
                    
                    print(f"[SYNC] Total store-brands from HO: {len(store_brands_data)}")
                    
                    # Get all store-brand IDs from HO response
                    ho_storebrand_ids = [sb['id'] for sb in store_brands_data if 'id' in sb]
                    
                    # Handle store-brands not in HO response - use soft delete if brand has transactions
                    storebrands_not_in_ho = StoreBrand.objects.filter(
                        store=store_config
                    ).exclude(
                        id__in=ho_storebrand_ids
                    ).select_related('brand')
                    
                    soft_deleted_count = 0
                    hard_deleted_count = 0
                    
                    for store_brand in storebrands_not_in_ho:
                        # Check if the brand has any transactions
                        from apps.pos.models import Bill, BillItem
                        has_transactions = (
                            Bill.objects.filter(brand=store_brand.brand).exists() or
                            BillItem.objects.filter(brand=store_brand.brand).exists()
                        )
                        
                        if has_transactions:
                            # Soft delete - set is_active=False to preserve data integrity
                            store_brand.is_active = False
                            store_brand.save(update_fields=['is_active'])
                            soft_deleted_count += 1
                            print(f"[SYNC] ⚠️ StoreBrand '{store_brand.brand.name}' has transactions - soft deleted (is_active=False)")
                        else:
                            # Hard delete - safe to remove as no transactions exist
                            brand_name = store_brand.brand.name
                            store_brand.delete()
                            hard_deleted_count += 1
                            print(f"[SYNC] 🗑️ StoreBrand '{brand_name}' has no transactions - hard deleted")
                    
                    deleted_count = soft_deleted_count + hard_deleted_count
                    if deleted_count > 0:
                        print(f"[SYNC] StoreBrands not in HO: {soft_deleted_count} soft-deleted, {hard_deleted_count} hard-deleted")
                    
                    for sb_data in store_brands_data:
                        try:
                            brand_id = sb_data.get('brand_id')
                            if not brand_id:
                                print(f"[SYNC] Store-Brand skipped - missing brand_id")
                                continue
                            
                            brand_obj = Brand.objects.filter(id=brand_id).first()
                            if not brand_obj:
                                print(f"[SYNC] Store-Brand skipped - brand not found: {brand_id}")
                                continue
                            
                            # store_id from HO API response is the HO Store ID
                            ho_store_id_from_api = sb_data.get('store_id')
                            
                            defaults = {
                                'store': store_config,
                                'brand': brand_obj,
                                'ho_store_id': ho_store_id_from_api,  # This is HO Store ID
                                'is_active': sb_data.get('is_active', True),
                            }
                            
                            print(f"[SYNC] Syncing StoreBrand: brand={brand_obj.name}, ho_store_id={ho_store_id_from_api}")
                            
                            storebrand_obj, created = StoreBrand.objects.update_or_create(
                                id=sb_data['id'],
                                defaults=defaults
                            )
                            
                            if created:
                                saved_count += 1
                                print(f"[SYNC] ✓ StoreBrand created: {brand_obj.name}")
                            else:
                                updated_count += 1
                                print(f"[SYNC] ✓ StoreBrand updated: {brand_obj.name}")
                                
                        except Exception as sb_error:
                            print(f"Error saving store-brand: {sb_error}")
                            import traceback
                            traceback.print_exc()
                            continue
                    
                    synced_tables.append(table_name)
                    sync_results[table_name] = {
                        'success': True,
                        'records_count': saved_count + updated_count,
                        'details': f'{saved_count} created, {updated_count} updated, {deleted_count} deleted'
                    }
                
                elif table_name == 'core_category':
                    # Sync Categories
                    categories = client.get_categories(company_id=company_id, store_id=ho_store_id)
                    saved_count = 0
                    skipped_no_brand = 0
                    
                    # Log sample to check structure
                    if categories and len(categories) > 0:
                        print(f"[SYNC] Sample category data: {categories[0]}")
                    
                    # First pass: Create/update all categories without parent
                    for cat_data in categories:
                        try:
                            # Get brand_id from API response
                            category_brand_id = cat_data.get('brand_id')
                            if not category_brand_id:
                                print(f"[SYNC] Category {cat_data.get('name')} skipped - missing brand_id")
                                skipped_no_brand += 1
                                continue
                            
                            category_brand = Brand.objects.filter(id=category_brand_id).first()
                            if not category_brand:
                                print(f"[SYNC] Category {cat_data.get('name')} skipped - brand not found: {category_brand_id}")
                                skipped_no_brand += 1
                                continue
                            
                            defaults = {
                                'brand': category_brand,
                                'name': cat_data.get('name', 'Unnamed Category'),
                                'is_active': cat_data.get('is_active', True),
                                'sort_order': cat_data.get('sort_order', 0),
                                'icon': cat_data.get('icon', ''),
                            }
                            
                            Category.objects.update_or_create(
                                id=cat_data['id'],
                                defaults=defaults
                            )
                            saved_count += 1
                        except Exception as cat_error:
                            print(f"Error saving category {cat_data.get('name', 'unknown')}: {cat_error}")
                            continue
                    
                    # Second pass: Update parent relationships
                    for cat_data in categories:
                        try:
                            if cat_data.get('parent_id'):
                                category = Category.objects.get(id=cat_data['id'])
                                parent = Category.objects.filter(id=cat_data['parent_id']).first()
                                if parent:
                                    category.parent = parent
                                    category.save(update_fields=['parent'])
                        except Exception as parent_error:
                            print(f"Error setting parent for category {cat_data.get('name', 'unknown')}: {parent_error}")
                            continue
                    
                    synced_tables.append(table_name)
                    sync_results[table_name] = {
                        'success': True,
                        'records_count': saved_count,
                        'details': f'{saved_count} saved, {skipped_no_brand} skipped (no brand)'
                    }
                
                elif table_name == 'tables_table':
                    # Sync Tables with dependencies in correct order:
                    # 1. Table Areas → 2. Tables → 3. Table Groups
                    
                    from apps.tables.models import TableArea, Table, TableGroup
                    
                    # Step 1: Sync Table Areas
                    table_areas = client.get_table_areas(company_id=company_id, store_id=ho_store_id)
                    area_count = 0
                    area_skipped_no_brand = 0
                    
                    if table_areas and len(table_areas) > 0:
                        print(f"[SYNC] Sample table area data: {table_areas[0]}")
                    
                    for area_data in table_areas:
                        try:
                            # Get brand_id from API response
                            area_brand_id = area_data.get('brand_id')
                            if not area_brand_id:
                                print(f"[SYNC] Table area {area_data.get('name')} skipped - missing brand_id")
                                area_skipped_no_brand += 1
                                continue
                            
                            area_brand = Brand.objects.filter(id=area_brand_id).first()
                            if not area_brand:
                                print(f"[SYNC] Table area {area_data.get('name')} skipped - brand not found: {area_brand_id}")
                                area_skipped_no_brand += 1
                                continue
                            
                            TableArea.objects.update_or_create(
                                id=area_data['id'],
                                defaults={
                                    'brand': area_brand,
                                    'company': company,
                                    'store': store_config,
                                    'name': area_data['name'],
                                    'description': area_data.get('description', ''),
                                    'sort_order': area_data.get('sort_order', 0),
                                    'is_active': area_data.get('is_active', True),
                                    'floor_width': area_data.get('floor_width'),
                                    'floor_height': area_data.get('floor_height'),
                                }
                            )
                            area_count += 1
                        except Exception as area_error:
                            print(f"Error saving table area {area_data.get('name', 'unknown')}: {area_error}")
                            continue
                    
                    # Step 2: Sync Tables
                    tables = client.get_tables(company_id=company_id, store_id=ho_store_id)
                    table_count = 0
                    
                    if tables and len(tables) > 0:
                        print(f"[SYNC] Sample table data: {tables[0]}")
                    
                    for table_data in tables:
                        try:
                            area = TableArea.objects.filter(id=table_data.get('area_id')).first()
                            if area:
                                Table.objects.update_or_create(
                                    id=table_data['id'],
                                    defaults={
                                        'area': area,
                                        'number': table_data['number'],
                                        'capacity': table_data.get('capacity', 4),
                                        'status': table_data.get('status', 'available'),
                                        'qr_code': table_data.get('qr_code', ''),
                                        'is_active': table_data.get('is_active', True),
                                        'pos_x': table_data.get('pos_x'),
                                        'pos_y': table_data.get('pos_y'),
                                        'shape': table_data.get('shape', 'rect'),
                                    }
                                )
                                table_count += 1
                            else:
                                print(f"Table {table_data.get('number')} skipped - area not found: {table_data.get('area_id')}")
                        except Exception as table_error:
                            print(f"Error saving table {table_data.get('number', 'unknown')}: {table_error}")
                            continue
                    
                    # Step 3: Sync Table Groups (optional, may be empty)
                    try:
                        table_groups = client.get_table_groups(company_id=company_id, store_id=ho_store_id)
                        group_count = 0
                        for group_data in table_groups:
                            try:
                                main_table = Table.objects.filter(id=group_data.get('main_table_id')).first()
                                if main_table:
                                    # Get brand from main_table's area
                                    group_brand = main_table.area.brand
                                    
                                    TableGroup.objects.update_or_create(
                                        id=group_data['id'],
                                        defaults={
                                            'main_table': main_table,
                                            'brand': group_brand,
                                            'created_by_id': group_data.get('created_by_id'),
                                        }
                                    )
                                    group_count += 1
                            except Exception as group_error:
                                print(f"Error saving table group: {group_error}")
                                continue
                    except Exception as e:
                        print(f"Table groups sync skipped: {e}")
                        group_count = 0
                    
                    synced_tables.append(table_name)
                    sync_results[table_name] = {
                        'success': True,
                        'records_count': table_count,
                        'details': f'{area_count} areas ({area_skipped_no_brand} skipped), {table_count} tables, {group_count} groups'
                    }
                
                elif table_name == 'promotions_promotion':
                    # Sync Promotions from HO (denormalized schema)
                    from apps.promotions.models import Promotion, PromotionSyncLog
                    from datetime import datetime
                    import json
                    
                    sync_start = timezone.now()
                    promotions = client.get_promotions(company_id=company_id, store_id=ho_store_id)
                    promo_count = 0
                    updated_count = 0
                    promo_skipped_no_brand = 0
                    
                    if promotions and len(promotions) > 0:
                        print(f"[SYNC] Sample promotion data: {promotions[0]}")
                    
                    for promo_data in promotions:
                        try:
                            # Get brand_id from API response
                            promo_brand_id = promo_data.get('brand_id')
                            if not promo_brand_id:
                                print(f"[SYNC] Promotion {promo_data.get('code')} skipped - missing brand_id")
                                promo_skipped_no_brand += 1
                                continue
                            
                            promo_brand = Brand.objects.filter(id=promo_brand_id).first()
                            if not promo_brand:
                                print(f"[SYNC] Promotion {promo_data.get('code')} skipped - brand not found: {promo_brand_id}")
                                promo_skipped_no_brand += 1
                                continue
                            
                            # Extract validity fields
                            validity = promo_data.get('validity', {})
                            start_date_str = validity.get('start_date')
                            end_date_str = validity.get('end_date')
                            
                            # Parse dates
                            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
                            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None
                            
                            if not start_date or not end_date:
                                print(f"Promotion {promo_data.get('code')} skipped - missing dates")
                                continue
                            
                            # Parse time fields
                            time_start = validity.get('time_start')
                            time_end = validity.get('time_end')
                            if time_start:
                                time_start = datetime.strptime(time_start, '%H:%M:%S').time()
                            if time_end:
                                time_end = datetime.strptime(time_end, '%H:%M:%S').time()
                            
                            # Parse compiled_at
                            compiled_at_str = promo_data.get('compiled_at')
                            if compiled_at_str:
                                compiled_at = datetime.fromisoformat(compiled_at_str.replace('Z', '+00:00'))
                            else:
                                compiled_at = timezone.now()
                            
                            # Create or update promotion
                            promo, created = Promotion.objects.update_or_create(
                                id=promo_data['id'],
                                defaults={
                                    'company': company,
                                    'brand': promo_brand,
                                    'store': store_config,
                                    'code': promo_data['code'],
                                    'name': promo_data['name'],
                                    'description': promo_data.get('description', ''),
                                    'terms_conditions': promo_data.get('terms_conditions', ''),
                                    'promo_type': promo_data.get('promo_type', 'percent_discount'),
                                    'apply_to': promo_data.get('apply_to', 'all'),
                                    'execution_stage': promo_data.get('execution_stage', 'item_level'),
                                    'execution_priority': promo_data.get('execution_priority', 500),
                                    'is_active': promo_data.get('is_active', True),
                                    'is_auto_apply': promo_data.get('is_auto_apply', False),
                                    'require_voucher': promo_data.get('require_voucher', False),
                                    'member_only': promo_data.get('member_only', False),
                                    'is_stackable': promo_data.get('is_stackable', False),
                                    'start_date': start_date,
                                    'end_date': end_date,
                                    'time_start': time_start,
                                    'time_end': time_end,
                                    'valid_days': json.dumps(validity.get('days_of_week', [])),
                                    'exclude_holidays': validity.get('exclude_holidays', False),
                                    'rules_json': json.dumps(promo_data.get('rules', {})),
                                    'scope_json': json.dumps(promo_data.get('scope', {})),
                                    'targeting_json': json.dumps(promo_data.get('targeting', {})),
                                    'max_uses': promo_data.get('limits', {}).get('max_uses'),
                                    'max_uses_per_customer': promo_data.get('limits', {}).get('max_uses_per_customer'),
                                    'max_uses_per_day': promo_data.get('limits', {}).get('max_uses_per_day'),
                                    'current_uses': promo_data.get('limits', {}).get('current_uses', 0),
                                    'compiled_at': compiled_at,
                                }
                            )
                            
                            if created:
                                promo_count += 1
                            else:
                                updated_count += 1
                                
                        except Exception as promo_error:
                            print(f"Error saving promotion {promo_data.get('code', 'unknown')}: {promo_error}")
                            continue
                    
                    # Log sync operation
                    sync_end = timezone.now()
                    duration = int((sync_end - sync_start).total_seconds())
                    
                    PromotionSyncLog.objects.create(
                        sync_type='manual',
                        sync_status='success',
                        promotions_received=len(promotions),
                        promotions_added=promo_count,
                        promotions_updated=updated_count,
                        promotions_deleted=0,
                        company=company,
                        store=store_config,
                        started_at=sync_start,
                        completed_at=sync_end,
                        duration_seconds=duration,
                        edge_version='1.0'
                    )
                    
                    synced_tables.append(table_name)
                    sync_results[table_name] = {
                        'success': True,
                        'records_count': promo_count + updated_count,
                        'details': f'{promo_count} added, {updated_count} updated, {promo_skipped_no_brand} skipped (no brand)'
                    }
                
                elif table_name == 'core_product':
                    # Sync Products with dependencies in correct order:
                    # 1. Modifiers → 2. Modifier Options → 3. Products → 4. Product-Modifier relationships
                    
                    from apps.core.models import Modifier, ModifierOption, ProductModifier
                    
                    # API call order (as requested):
                    # 0) categories, 1) product-modifiers, 2) modifier-options, 3) modifiers, 4) products
                    categories = client.get_categories(company_id=company_id, store_id=ho_store_id)
                    product_modifiers = client.get_product_modifiers(company_id=company_id, store_id=ho_store_id)
                    modifier_options = client.get_modifier_options(company_id=company_id, store_id=ho_store_id)
                    modifiers = client.get_modifiers(company_id=company_id, store_id=ho_store_id)
                    products = client.get_products(company_id=company_id, store_id=ho_store_id)
                    product_received = len(products)
                    
                    # Log first product to see structure
                    if products and len(products) > 0:
                        print(f"[SYNC] Sample product data from HO API: {products[0]}")
                    else:
                        print(f"[SYNC] No products received from HO API")

                    # Sync categories first (needed for products)
                    category_count = 0
                    category_skipped_no_brand = 0
                    for cat_data in categories:
                        try:
                            # Get brand_id from API response
                            cat_brand_id = cat_data.get('brand_id')
                            if not cat_brand_id:
                                print(f"[SYNC] Category {cat_data.get('name')} skipped - missing brand_id")
                                category_skipped_no_brand += 1
                                continue
                            
                            cat_brand = Brand.objects.filter(id=cat_brand_id).first()
                            if not cat_brand:
                                print(f"[SYNC] Category {cat_data.get('name')} skipped - brand not found: {cat_brand_id}")
                                category_skipped_no_brand += 1
                                continue
                            
                            defaults = {
                                'brand': cat_brand,
                                'name': cat_data.get('name', 'Unnamed Category'),
                                'is_active': cat_data.get('is_active', True),
                                'sort_order': cat_data.get('sort_order', 0),
                                'icon': cat_data.get('icon', ''),
                            }

                            Category.objects.update_or_create(
                                id=cat_data['id'],
                                defaults=defaults
                            )
                            category_count += 1
                        except Exception as cat_error:
                            print(f"Error saving category {cat_data.get('name', 'unknown')}: {cat_error}")
                            continue

                    for cat_data in categories:
                        try:
                            if cat_data.get('parent_id'):
                                category = Category.objects.get(id=cat_data['id'])
                                parent = Category.objects.filter(id=cat_data['parent_id']).first()
                                if parent:
                                    category.parent = parent
                                    category.save(update_fields=['parent'])
                        except Exception as parent_error:
                            print(f"Error setting parent for category {cat_data.get('name', 'unknown')}: {parent_error}")
                            continue

                    # Process in safe FK order:
                    # 1) modifiers -> 2) modifier options -> 3) products -> 4) product-modifier links
                    modifier_count = 0
                    modifier_skipped_no_brand = 0
                    
                    if modifiers and len(modifiers) > 0:
                        print(f"[SYNC] Sample modifier data: {modifiers[0]}")
                    
                    for mod_data in modifiers:
                        try:
                            # Get brand_id from API response
                            mod_brand_id = mod_data.get('brand_id')
                            if not mod_brand_id:
                                print(f"[SYNC] Modifier {mod_data.get('name')} skipped - missing brand_id")
                                modifier_skipped_no_brand += 1
                                continue
                            
                            mod_brand = Brand.objects.filter(id=mod_brand_id).first()
                            if not mod_brand:
                                print(f"[SYNC] Modifier {mod_data.get('name')} skipped - brand not found: {mod_brand_id}")
                                modifier_skipped_no_brand += 1
                                continue
                            
                            Modifier.objects.update_or_create(
                                id=mod_data['id'],
                                defaults={
                                    'brand': mod_brand,
                                    'name': mod_data['name'],
                                    'is_required': mod_data.get('is_required', False),
                                    'max_selections': mod_data.get('max_selections', 1),
                                    'is_active': mod_data.get('is_active', True),
                                }
                            )
                            modifier_count += 1
                        except Exception as mod_error:
                            print(f"Error saving modifier {mod_data.get('name', 'unknown')}: {mod_error}")
                            continue

                    option_count = 0
                    for opt_data in modifier_options:
                        try:
                            modifier = Modifier.objects.filter(id=opt_data.get('modifier_id')).first()
                            if modifier:
                                ModifierOption.objects.update_or_create(
                                    id=opt_data['id'],
                                    defaults={
                                        'modifier': modifier,
                                        'name': opt_data['name'],
                                        'price_adjustment': Decimal(str(opt_data.get('price_adjustment', 0))),
                                        'is_default': opt_data.get('is_default', False),
                                        'sort_order': opt_data.get('sort_order', 0),
                                        'is_active': opt_data.get('is_active', True),
                                    }
                                )
                                option_count += 1
                        except Exception as opt_error:
                            print(f"Error saving modifier option {opt_data.get('name', 'unknown')}: {opt_error}")
                            continue

                    product_count = 0
                    product_duplicate_sku_updates = 0
                    product_skipped_missing_category = 0
                    product_skipped_missing_brand = 0
                    product_errors = 0
                    
                    if products and len(products) > 0:
                        print(f"[SYNC] Sample product data (first product):")
                        print(f"  - ID: {products[0].get('id')}")
                        print(f"  - Name: {products[0].get('name')}")
                        print(f"  - Brand ID: {products[0].get('brand_id')}")
                        print(f"  - Category ID: {products[0].get('category_id')}")
                        print(f"  - Printer Target: {products[0].get('printer_target', 'NOT FOUND IN API')}")
                        print(f"  - Full keys: {list(products[0].keys())}")
                    
                    for prod_data in products:
                        try:
                            category = Category.objects.filter(id=prod_data.get('category_id')).first()
                            
                            # Get brand_id from product data (products belong to specific brands)
                            product_brand_id = prod_data.get('brand_id')
                            if not product_brand_id:
                                print(f"[SYNC] Product {prod_data.get('name')} skipped - missing brand_id in API response")
                                product_skipped_missing_brand += 1
                                continue
                            
                            product_brand = Brand.objects.filter(id=product_brand_id).first()
                            if not product_brand:
                                print(f"[SYNC] Product {prod_data.get('name')} skipped - brand not found: {product_brand_id}")
                                product_skipped_missing_brand += 1
                                continue
                            
                            if category:
                                try:
                                    # Get company for product
                                    product_company = None
                                    if prod_data.get('company_id'):
                                        product_company = Company.objects.filter(id=prod_data['company_id']).first()
                                    
                                    Product.objects.update_or_create(
                                        id=prod_data['id'],
                                        defaults={
                                            'brand': product_brand,
                                            'category': category,
                                            'company': product_company,
                                            'name': prod_data['name'],
                                            'description': prod_data.get('description', ''),
                                            'price': Decimal(str(prod_data['price'])),
                                            'cost': Decimal(str(prod_data.get('cost', 0))),
                                            'sku': prod_data.get('sku', ''),
                                            'image': prod_data.get('image', ''),
                                            'printer_target': prod_data.get('printer_target', ''),
                                            'track_stock': prod_data.get('track_stock', False),
                                            'stock_quantity': Decimal(str(prod_data.get('stock_quantity', 0))),
                                            'is_active': prod_data.get('is_active', True),
                                            'sort_order': prod_data.get('sort_order', 0),
                                        }
                                    )
                                    product_count += 1
                                except IntegrityError:
                                    sku = prod_data.get('sku', '')
                                    existing = None
                                    if sku:
                                        existing = Product.objects.filter(brand=product_brand, sku=sku).first()
                                    if existing:
                                        # Update company if provided
                                        if prod_data.get('company_id'):
                                            existing.company = Company.objects.filter(id=prod_data['company_id']).first()
                                        
                                        existing.category = category
                                        existing.name = prod_data['name']
                                        existing.description = prod_data.get('description', '')
                                        existing.price = Decimal(str(prod_data['price']))
                                        existing.cost = Decimal(str(prod_data.get('cost', 0)))
                                        existing.image = prod_data.get('image', '')
                                        existing.printer_target = prod_data.get('printer_target', '')
                                        existing.track_stock = prod_data.get('track_stock', False)
                                        existing.stock_quantity = Decimal(str(prod_data.get('stock_quantity', 0)))
                                        existing.is_active = prod_data.get('is_active', True)
                                        existing.sort_order = prod_data.get('sort_order', 0)
                                        existing.save(update_fields=[
                                            'company', 'category', 'name', 'description', 'price', 'cost',
                                            'image', 'printer_target', 'track_stock', 'stock_quantity',
                                            'is_active', 'sort_order'
                                        ])
                                        product_count += 1
                                        product_duplicate_sku_updates += 1
                                        print(f"Product {prod_data.get('name', 'unknown')} updated by SKU due to duplicate constraint (sku={sku})")
                                    else:
                                        raise
                            else:
                                print(f"Product {prod_data.get('name')} skipped - category not found: {prod_data.get('category_id')}")
                                product_skipped_missing_category += 1
                        except Exception as prod_error:
                            print(f"Error saving product {prod_data.get('name', 'unknown')}: {prod_error}")
                            product_errors += 1
                            continue

                    link_count = 0
                    for pm_data in product_modifiers:
                        try:
                            product = Product.objects.filter(id=pm_data.get('product_id')).first()
                            modifier = Modifier.objects.filter(id=pm_data.get('modifier_id')).first()
                            if product and modifier:
                                product_modifier, created = ProductModifier.objects.get_or_create(
                                    product=product,
                                    modifier=modifier,
                                    defaults={
                                        'sort_order': pm_data.get('sort_order', 0)
                                    }
                                )
                                if created:
                                    link_count += 1
                        except Exception as pm_error:
                            print(f"Error linking product-modifier: {pm_error}")
                            continue
                    
                    # Count totals in Edge DB for all brands in this store
                    from apps.core.models import StoreBrand
                    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True)
                    brand_ids = list(store_brands.values_list('brand_id', flat=True))
                    
                    brand_category_total = Category.objects.filter(brand_id__in=brand_ids).count()
                    brand_product_total = Product.objects.filter(brand_id__in=brand_ids).count()
                    brand_modifier_total = Modifier.objects.filter(brand_id__in=brand_ids).count()
                    modifier_ids = list(Modifier.objects.filter(brand_id__in=brand_ids).values_list('id', flat=True))
                    brand_option_total = ModifierOption.objects.filter(modifier_id__in=modifier_ids).count()
                    brand_link_total = ProductModifier.objects.filter(modifier__brand_id__in=brand_ids).count()
                    
                    # Track HO counts
                    ho_category_count = len(categories)
                    ho_modifier_count = len(modifiers)
                    ho_option_count = len(modifier_options)
                    ho_product_count = len(products)
                    ho_link_count = len(product_modifiers)
                    
                    # Build comparison checklist
                    checklist = {
                        'categories': {
                            'ho': ho_category_count,
                            'edge': brand_category_total,
                            'match': ho_category_count == brand_category_total
                        },
                        'modifiers': {
                            'ho': ho_modifier_count,
                            'edge': brand_modifier_total,
                            'match': ho_modifier_count == brand_modifier_total
                        },
                        'options': {
                            'ho': ho_option_count,
                            'edge': brand_option_total,
                            'match': ho_option_count == brand_option_total
                        },
                        'products': {
                            'ho': ho_product_count,
                            'edge': brand_product_total,
                            'match': ho_product_count == brand_product_total
                        },
                        'links': {
                            'ho': ho_link_count,
                            'edge': brand_link_total,
                            'match': ho_link_count == brand_link_total
                        }
                    }
                    
                    synced_tables.append(table_name)
                    sync_results[table_name] = {
                        'success': True,
                        'records_count': brand_product_total,
                        'checklist': checklist,
                        'details': (
                            f'{category_count} categories ({category_skipped_no_brand} skipped), '
                            f'{modifier_count} modifiers ({modifier_skipped_no_brand} skipped), '
                            f'{option_count} options, '
                            f'{product_count} products processed (received {product_received}, '
                            f'updated_by_sku {product_duplicate_sku_updates}, '
                            f'skipped_no_category {product_skipped_missing_category}, '
                            f'skipped_no_brand {product_skipped_missing_brand}, '
                            f'errors {product_errors}, total_in_db {brand_product_total}), '
                            f'{link_count} links'
                        )
                    }
                
                else:
                    # Table not yet implemented
                    sync_results[table_name] = {
                        'success': False,
                        'error': f'Sync for table "{table_name}" not yet implemented. Coming soon!'
                    }
            
            except HOAPIException as e:
                sync_results[table_name] = {
                    'success': False,
                    'error': f'HO API Error: {str(e)}'
                }
            except Exception as e:
                import traceback
                traceback.print_exc()
                sync_results[table_name] = {
                    'success': False,
                    'error': str(e)
                }
        
        # Prepare response with checklist validation
        if synced_tables:
            # Build checklist items with HO vs Edge comparison
            all_match = True
            checklist_items = []
            
            label_map = {
                'categories': 'categories',
                'modifiers': 'modifiers',
                'options': 'modifier-options',
                'products': 'products',
                'links': 'product-modifiers',
            }

            for table_name, result in sync_results.items():
                if result.get('success'):
                    # For tables with detailed checklist
                    if 'checklist' in result:
                        checklist = result['checklist']
                        for item_name, comparison in checklist.items():
                            is_match = comparison.get('match', False)
                            all_match = all_match and is_match
                            icon = '✓' if is_match else '✗'
                            checklist_items.append({
                                'item': label_map.get(item_name, item_name),
                                'ho': comparison['ho'],
                                'edge': comparison['edge'],
                                'match': is_match,
                                'icon': icon
                            })
                    # For simple tables (Brand, StoreBrand, etc) - use records_count
                    elif 'records_count' in result:
                        checklist_items.append({
                            'item': table_name.replace('core_', '').replace('_', '-').title(),
                            'ho': result['records_count'],
                            'edge': result['records_count'],
                            'match': True,
                            'icon': '✓'
                        })
            
            return JsonResponse({
                'success': True,
                'message': f'Successfully synced {len(synced_tables)} table(s)',
                'synced_tables': synced_tables,
                'results': sync_results,
                'checklist_items': checklist_items,
                'all_match': all_match
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'No tables were synced successfully',
                'results': sync_results
            }, status=400)
    
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON in request body'
        }, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@manager_required
def brands_list(request):
    """Brands List - Display all brands in the system"""
    store_config = Store.get_current()
    
    # Get all brands (only from HO after sync)
    brands = Brand.objects.all().order_by('code')
    
    # Get store brands (brands assigned to this store)
    from apps.core.models import StoreBrand
    store_brands = StoreBrand.objects.filter(
        store=store_config
    ).select_related('brand', 'brand__company').order_by('brand__name')
    
    context = {
        'brands': brands,
        'store_brands': store_brands,
        'store_config': store_config,
        'total_count': brands.count(),
        'store_brands_count': store_brands.count(),
    }
    
    return render(request, 'management/brands_list.html', context)


@manager_required
def categories(request):
    """Categories Management"""
    store_config = Store.get_current()
    
    # Get all brands associated with this store
    from apps.core.models import StoreBrand
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True)
    brand_ids = list(store_brands.values_list('brand_id', flat=True))
    
    # Get global context brand filter from session
    context_brand_id = request.session.get('context_brand_id', '')
    
    # Apply global brand filter if set
    if context_brand_id and context_brand_id in [str(bid) for bid in brand_ids]:
        brand_ids = [context_brand_id]
    
    categories_list = Category.objects.filter(
        brand_id__in=brand_ids
    ).select_related('parent', 'brand').order_by('brand__name', 'sort_order', 'name')
    
    # Calculate counts for template
    active_count = categories_list.filter(is_active=True).count()
    parent_count = categories_list.filter(parent__isnull=True).count()
    
    context = {
        'categories': categories_list,
        'total_count': categories_list.count(),
        'active_count': active_count,
        'parent_count': parent_count,
    }
    
    return render(request, 'management/categories.html', context)


@manager_required
def products(request):
    """Products Management"""
    from django.conf import settings
    from django.db.models import Prefetch
    
    store_config = Store.get_current()
    
    # Get all brands associated with this store
    from apps.core.models import StoreBrand
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True)
    brand_ids = list(store_brands.values_list('brand_id', flat=True))
    
    # Get global context brand filter from session
    context_brand_id = request.session.get('context_brand_id', '')
    
    # Apply global brand filter if set
    if context_brand_id and context_brand_id in [str(bid) for bid in brand_ids]:
        brand_ids = [context_brand_id]
    
    # Get all products from filtered brands
    products_list = Product.objects.filter(
        brand_id__in=brand_ids
    ).select_related('category', 'brand').prefetch_related(
        'product_modifiers__modifier',
        Prefetch(
            'photos',  # Use 'photos' related_name from ProductPhoto model
            queryset=ProductPhoto.objects.filter(is_primary=True).order_by('sort_order'),
            to_attr='primary_photos'
        )
    ).order_by('brand__name', 'category__name', 'name')
    
    # Filter by brand (local filter on page)
    brand_filter = request.GET.get('brand', '')
    if brand_filter:
        products_list = products_list.filter(brand_id=brand_filter)
    
    # Get categories for filter dropdown (filtered by selected brand if any)
    if brand_filter:
        categories_list = Category.objects.filter(brand_id=brand_filter).order_by('name')
    elif context_brand_id:
        categories_list = Category.objects.filter(brand_id=context_brand_id).order_by('name')
    else:
        categories_list = Category.objects.filter(brand_id__in=brand_ids).order_by('brand__name', 'name')
    
    # Filter by category
    category_filter = request.GET.get('category', '')
    if category_filter:
        products_list = products_list.filter(category_id=category_filter)
    
    # Search
    search = request.GET.get('search', '')
    if search:
        products_list = products_list.filter(
            Q(name__icontains=search) | Q(sku__icontains=search)
        )
    
    # Filter by active status
    status_filter = request.GET.get('status', 'active')
    if status_filter == 'active':
        products_list = products_list.filter(is_active=True)
    elif status_filter == 'inactive':
        products_list = products_list.filter(is_active=False)
    
    # Get brands for filter dropdown
    brands_list = Brand.objects.filter(id__in=brand_ids).order_by('name')
    
    # Calculate counts
    total_count = products_list.count()
    active_count = Product.objects.filter(brand_id__in=brand_ids, is_active=True).count()
    inactive_count = Product.objects.filter(brand_id__in=brand_ids, is_active=False).count()
    
    # Pagination
    paginator = Paginator(products_list, 20)  # 20 products per page
    page = request.GET.get('page', 1)
    
    try:
        products_page = paginator.page(page)
    except PageNotAnInteger:
        products_page = paginator.page(1)
    except EmptyPage:
        products_page = paginator.page(paginator.num_pages)
    
    # Get MinIO settings (browser-accessible URL)
    minio_endpoint = get_minio_endpoint_for_request(request)
    minio_bucket = 'product-images'
    
    context = {
        'products': products_page,
        'categories': categories_list,
        'brands': brands_list,
        'total_count': total_count,
        'active_count': active_count,
        'inactive_count': inactive_count,
        'category_filter': category_filter,
        'brand_filter': brand_filter,
        'search': search,
        'status_filter': status_filter,
        'paginator': paginator,
        'minio_endpoint': minio_endpoint,
        'minio_bucket': minio_bucket,
    }
    
    return render(request, 'management/products.html', context)


@require_POST
@manager_required
def products_set_stock_default(request):
    """Set default stock quantity for all products in current brand"""
    store_config = Store.get_current()
    
    # Get all brands for this store
    store_brands = StoreBrand.objects.filter(store=store_config).select_related('brand')
    brand_ids = [sb.brand_id for sb in store_brands]
    
    # Apply global brand filter from session
    context_brand_id = request.session.get('context_brand_id')
    if context_brand_id:
        brand_ids = [context_brand_id]

    updated_count = Product.objects.filter(brand_id__in=brand_ids).update(
        track_stock=True,
        stock_quantity=1000,
    )

    if context_brand_id:
        messages.success(request, f'Stock default set to 1000 for {updated_count} products in selected brand')
    else:
        messages.success(request, f'Stock default set to 1000 for {updated_count} products in all brands')
    return redirect('management:products')


@manager_required
def tables_list(request):
    """Tables Management"""
    store_config, error_response = check_store_config(request, 'management/tables_management.html')
    if error_response:
        return error_response
    
    # Get all brands for this store
    store_brands = StoreBrand.objects.filter(store=store_config).select_related('brand')
    brand_ids = [sb.brand_id for sb in store_brands]
    
    # Apply global brand filter from session
    context_brand_id = request.session.get('context_brand_id')
    if context_brand_id:
        brand_ids = [context_brand_id]
    
    tables = Table.objects.filter(
        area__brand_id__in=brand_ids
    ).select_related('area').order_by('area__name', 'number')
    
    # Filter by area
    area_filter = request.GET.get('area', '')
    if area_filter:
        tables = tables.filter(area_id=area_filter)
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        tables = tables.filter(status=status_filter)
    
    # Get areas for filter
    areas = TableArea.objects.filter(brand_id__in=brand_ids).order_by('name')
    
    # Calculate status counts
    available_count = tables.filter(status='available').count()
    occupied_count = tables.filter(status='occupied').count()
    reserved_count = tables.filter(status='reserved').count()
    
    context = {
        'tables': tables,
        'areas': areas,
        'total_count': tables.count(),
        'available_count': available_count,
        'occupied_count': occupied_count,
        'reserved_count': reserved_count,
        'area_filter': area_filter,
        'status_filter': status_filter,
    }
    
    return render(request, 'management/tables_management.html', context)


@manager_required
def users_list(request):
    """Users Management"""
    store_config, error_response = check_store_config(request, 'management/users.html')
    if error_response:
        return error_response
    
    Brand = store_config.brand
    
    users = User.objects.filter(
        Q(brand=Brand) | Q(company=Brand.company) | Q(is_superuser=True)
    ).order_by('role', 'username')
    
    # Filter by role
    role_filter = request.GET.get('role', '')
    if role_filter:
        users = users.filter(role=role_filter)
    
    # Search
    search = request.GET.get('search', '')
    if search:
        users = users.filter(
            Q(username__icontains=search) | 
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search)
        )
    
    # Calculate role counts
    admin_count = users.filter(role='admin').count()
    manager_count = users.filter(role='manager').count()
    cashier_count = users.filter(role='cashier').count()
    waiter_count = users.filter(role='waiter').count()
    kitchen_count = users.filter(role='kitchen').count()
    
    context = {
        'users': users,
        'total_count': users.count(),
        'admin_count': admin_count,
        'manager_count': manager_count,
        'cashier_count': cashier_count,
        'waiter_count': waiter_count,
        'kitchen_count': kitchen_count,
        'role_filter': role_filter,
        'search': search,
    }
    
    return render(request, 'management/users.html', context)




@csrf_exempt
@require_POST
@manager_required
def user_set_password(request, user_id):
    """Set/Reset User Password (HTMX)"""
    
    user = get_object_or_404(User, id=user_id)
    new_password = request.POST.get('password', '').strip()
    
    if not new_password:
        return JsonResponse({'error': 'Password cannot be empty'}, status=400)
    
    if len(new_password) < 8:
        return JsonResponse({'error': 'Password must be at least 8 characters'}, status=400)
    
    # Set new password
    user.set_password(new_password)
    user.save()
    
    return JsonResponse({
        'success': True,
        'message': f'Password updated for {user.username}'
    })


@csrf_exempt
@require_POST
@manager_required
def user_set_pin(request, user_id):
    """Set User PIN for POS Login (HTMX)"""
    
    user = get_object_or_404(User, id=user_id)
    new_pin = request.POST.get('pin', '').strip()
    
    if not new_pin:
        return JsonResponse({'error': 'PIN cannot be empty'}, status=400)
    
    if not new_pin.isdigit():
        return JsonResponse({'error': 'PIN must be numeric'}, status=400)
    
    if len(new_pin) != 6:
        return JsonResponse({'error': 'PIN must be exactly 6 digits'}, status=400)
    
    # Set PIN (stored as hashed value for security)
    user.pin = new_pin  # Model will hash it automatically
    user.save()
    
    return JsonResponse({
        'success': True,
        'message': f'PIN set for {user.username}'
    })


@manager_required
def table_areas_list(request):
    """Table Areas Management"""
    store_config = Store.get_current()
    Brand = store_config.brand
    
    areas = TableArea.objects.filter(
        brand=Brand
    ).order_by('sort_order', 'name')
    
    # Count tables per area
    from django.db.models import Count
    areas = areas.annotate(table_count=Count('tables'))
    
    # Search
    search = request.GET.get('search', '')
    if search:
        areas = areas.filter(name__icontains=search)
    
    context = {
        'areas': areas,
        'total_count': areas.count(),
        'search': search,
    }
    
    return render(request, 'management/table_areas.html', context)


@manager_required
def promotions_list(request):
    """Promotions Management - Shows Promotions from HO"""
    store_config = Store.get_current()
    Brand = store_config.brand
    
    # Use new Promotion model (denormalized schema)
    promotions = Promotion.objects.filter(
        brand=Brand
    ).order_by('-start_date', 'name')
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        promotions = promotions.filter(is_active=True)
    elif status_filter == 'inactive':
        promotions = promotions.filter(is_active=False)
    
    # Filter by type
    type_filter = request.GET.get('type', '')
    if type_filter:
        promotions = promotions.filter(promo_type=type_filter)
    
    # Search
    search = request.GET.get('search', '')
    if search:
        promotions = promotions.filter(
            Q(name__icontains=search) | Q(code__icontains=search)
        )
    
    # Calculate counts
    active_count = Promotion.objects.filter(brand=Brand, is_active=True).count()
    inactive_count = Promotion.objects.filter(brand=Brand, is_active=False).count()
    
    # Check current validity
    now = timezone.now()
    valid_count = Promotion.objects.filter(
        brand=Brand,
        is_active=True,
        start_date__lte=now,
        end_date__gte=now
    ).count()
    
    context = {
        'promotions': promotions,
        'total_count': promotions.count(),
        'active_count': active_count,
        'inactive_count': inactive_count,
        'valid_count': valid_count,
        'status_filter': status_filter,
        'type_filter': type_filter,
        'search': search,
    }
    
    return render(request, 'management/promotions.html', context)


@manager_required
def promotion_detail(request, promotion_id):
    """Promotion Detail - Show all promotion information"""
    store_config = Store.get_current()
    Brand = store_config.brand
    
    promotion = get_object_or_404(
        Promotion.objects.select_related('brand', 'company', 'store'),
        id=promotion_id,
        brand=Brand
    )
    
    # Parse JSON fields
    rules = promotion.get_rules()
    scope = promotion.get_scope()
    targeting = promotion.get_targeting()
    valid_days = promotion.get_valid_days()
    
    # Resolve category/product UUIDs to names
    if scope:
        # Resolve categories
        if 'categories' in scope and scope['categories']:
            category_ids = scope['categories']
            categories = Category.objects.filter(id__in=category_ids).values_list('name', flat=True)
            scope['category_names'] = list(categories)
        
        if 'exclude_categories' in scope and scope['exclude_categories']:
            category_ids = scope['exclude_categories']
            categories = Category.objects.filter(id__in=category_ids).values_list('name', flat=True)
            scope['exclude_category_names'] = list(categories)
        
        # Resolve products
        if 'products' in scope and scope['products']:
            product_ids = scope['products']
            products = Product.objects.filter(id__in=product_ids).values_list('name', flat=True)
            scope['product_names'] = list(products)
        
        if 'exclude_products' in scope and scope['exclude_products']:
            product_ids = scope['exclude_products']
            products = Product.objects.filter(id__in=product_ids).values_list('name', flat=True)
            scope['exclude_product_names'] = list(products)
    
    # Get usage statistics
    from apps.promotions.models import PromotionUsage
    usage_stats = PromotionUsage.objects.filter(promotion=promotion).aggregate(
        total_usage=Count('id'),
        total_discount=Sum('discount_amount'),
        total_original=Sum('original_amount')
    )
    
    # Recent usage
    recent_usage = PromotionUsage.objects.filter(
        promotion=promotion
    ).order_by('-used_at')[:10]
    
    context = {
        'promotion': promotion,
        'rules': rules,
        'scope': scope,
        'targeting': targeting,
        'valid_days': valid_days,
        'usage_stats': usage_stats,
        'recent_usage': recent_usage,
    }
    
    return render(request, 'management/promotion_detail.html', context)


@manager_required
def vouchers_list(request):
    """Vouchers Management"""
    store_config = Store.get_current()
    Brand = store_config.brand
    
    vouchers = Voucher.objects.filter(
        promotion__brand=Brand
    ).select_related('promotion').order_by('-created_at')
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        vouchers = vouchers.filter(is_active=True, is_used=False)
    elif status_filter == 'used':
        vouchers = vouchers.filter(is_used=True)
    elif status_filter == 'inactive':
        vouchers = vouchers.filter(is_active=False)
    
    # Search by code or customer
    search = request.GET.get('search', '')
    if search:
        vouchers = vouchers.filter(
            Q(code__icontains=search) | 
            Q(customer_name__icontains=search) |
            Q(customer_phone__icontains=search)
        )
    
    # Calculate counts
    active_count = Voucher.objects.filter(
        promotion__brand=Brand,
        is_active=True,
        is_used=False
    ).count()
    used_count = Voucher.objects.filter(
        promotion__brand=Brand,
        is_used=True
    ).count()
    expired_count = Voucher.objects.filter(
        promotion__brand=Brand,
        expiry_date__lt=timezone.now(),
        is_used=False
    ).count()
    
    context = {
        'vouchers': vouchers,
        'total_count': vouchers.count(),
        'active_count': active_count,
        'used_count': used_count,
        'expired_count': expired_count,
        'status_filter': status_filter,
        'search': search,
    }
    
    return render(request, 'management/vouchers.html', context)


@manager_required
def bills_list(request):
    """Bills/Transactions List"""
    store_config = Store.get_current()
    Brand = store_config.brand
    
    bills = Bill.objects.filter(
        brand=Brand
    ).select_related('table', 'created_by', 'terminal').order_by('-created_at')
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        bills = bills.filter(status=status_filter)
    
    # Filter by date range
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        bills = bills.filter(created_at__date__gte=date_from)
    if date_to:
        bills = bills.filter(created_at__date__lte=date_to)
    
    # Search by bill number
    search = request.GET.get('search', '')
    if search:
        bills = bills.filter(bill_number__icontains=search)
    
    # Calculate counts and totals (before pagination)
    total_bills = Bill.objects.filter(brand=Brand).count()
    open_count = Bill.objects.filter(brand=Brand, status='open').count()
    paid_count = Bill.objects.filter(brand=Brand, status='paid').count()
    void_count = Bill.objects.filter(brand=Brand, status='void').count()
    
    total_revenue = Bill.objects.filter(
        brand=Brand, 
        status='paid'
    ).aggregate(total=Sum('total'))['total'] or Decimal('0')
    
    # Pagination
    paginator = Paginator(bills, 50)  # 50 bills per page
    page = request.GET.get('page', 1)
    
    try:
        bills_page = paginator.page(page)
    except PageNotAnInteger:
        bills_page = paginator.page(1)
    except EmptyPage:
        bills_page = paginator.page(paginator.num_pages)
    
    context = {
        'bills': bills_page,
        'total_count': total_bills,
        'open_count': open_count,
        'paid_count': paid_count,
        'void_count': void_count,
        'total_revenue': total_revenue,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search': search,
    }
    
    return render(request, 'management/bills.html', context)


@manager_required
def payments_list(request):
    """Payments List"""
    store_config = Store.get_current()
    Brand = store_config.brand
    
    payments = Payment.objects.filter(
        bill__brand=Brand
    ).select_related('bill', 'bill__created_by', 'payment_profile').order_by('-created_at')
    
    # Filter by payment method
    method_filter = request.GET.get('method', '')
    if method_filter:
        payments = payments.filter(method=method_filter)
    
    # Filter by date range
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        payments = payments.filter(created_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(created_at__date__lte=date_to)
    
    # Search by bill number
    search = request.GET.get('search', '')
    if search:
        payments = payments.filter(bill__bill_number__icontains=search)
    
    # Calculate totals by method (before pagination)
    total_payments = Payment.objects.filter(bill__brand=Brand).count()
    cash_total = Payment.objects.filter(
        bill__brand=Brand, method='cash'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    card_total = Payment.objects.filter(
        bill__brand=Brand, method='card'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    qris_total = Payment.objects.filter(
        bill__brand=Brand, method='qris'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    total_amount = Payment.objects.filter(
        bill__brand=Brand
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Pagination
    paginator = Paginator(payments, 50)  # 50 payments per page
    page = request.GET.get('page', 1)
    
    try:
        payments_page = paginator.page(page)
    except PageNotAnInteger:
        payments_page = paginator.page(1)
    except EmptyPage:
        payments_page = paginator.page(paginator.num_pages)
    
    context = {
        'payments': payments_page,
        'total_count': total_payments,
        'cash_total': cash_total,
        'card_total': card_total,
        'qris_total': qris_total,
        'total_amount': total_amount,
        'method_filter': method_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search': search,
    }
    
    return render(request, 'management/payments.html', context)


@manager_required
def store_sessions_list(request):
    """Store Sessions List"""
    store_config = Store.get_current()
    
    sessions = StoreSession.objects.filter(
        store=store_config
    ).order_by('-opened_at')
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter == 'open':
        sessions = sessions.filter(closed_at__isnull=True)
    elif status_filter == 'closed':
        sessions = sessions.filter(closed_at__isnull=False)
    
    # Calculate counts (before pagination)
    total_sessions = StoreSession.objects.filter(store=store_config).count()
    open_sessions = StoreSession.objects.filter(
        store=store_config, closed_at__isnull=True
    ).count()
    closed_sessions = StoreSession.objects.filter(
        store=store_config, closed_at__isnull=False
    ).count()
    
    # Pagination
    paginator = Paginator(sessions, 30)  # 30 sessions per page
    page = request.GET.get('page', 1)
    
    try:
        sessions_page = paginator.page(page)
    except PageNotAnInteger:
        sessions_page = paginator.page(1)
    except EmptyPage:
        sessions_page = paginator.page(paginator.num_pages)
    
    context = {
        'sessions': sessions_page,
        'total_count': total_sessions,
        'open_count': open_sessions,
        'closed_count': closed_sessions,
        'status_filter': status_filter,
    }
    
    return render(request, 'management/store_sessions.html', context)


# ===========================
# REPORTS & ANALYTICS
# ===========================

@manager_required
def reports_dashboard(request):
    """Reports overview dashboard"""
    store_config = Store.get_current()
    Brand = store_config.brand
    
    # Quick stats for last 7 days
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    
    last_week_revenue = Bill.objects.filter(
        brand=Brand,
        status='paid',
        created_at__date__gte=week_ago
    ).aggregate(total=Sum('total'))['total'] or Decimal('0')
    
    last_week_bills = Bill.objects.filter(
        brand=Brand,
        status='paid',
        created_at__date__gte=week_ago
    ).count()
    
    context = {
        'last_week_revenue': last_week_revenue,
        'last_week_bills': last_week_bills,
        'avg_transaction': last_week_revenue / last_week_bills if last_week_bills > 0 else 0,
    }
    
    return render(request, 'management/reports/dashboard.html', context)


@manager_required
def sales_report(request):
    """Sales report with period selection"""
    store_config = Store.get_current()
    Brand = store_config.brand
    
    # Get date range from request
    period = request.GET.get('period', 'today')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    today = timezone.now().date()
    
    # Calculate date range based on period
    if period == 'today':
        date_from = today
        date_to = today
    elif period == 'yesterday':
        date_from = today - timedelta(days=1)
        date_to = today - timedelta(days=1)
    elif period == 'week':
        date_from = today - timedelta(days=today.weekday())
        date_to = today
    elif period == 'last_week':
        date_from = today - timedelta(days=today.weekday() + 7)
        date_to = today - timedelta(days=today.weekday() + 1)
    elif period == 'month':
        date_from = today.replace(day=1)
        date_to = today
    elif period == 'last_month':
        last_month = today.replace(day=1) - timedelta(days=1)
        date_from = last_month.replace(day=1)
        date_to = last_month
    elif period == 'custom' and start_date and end_date:
        date_from = datetime.strptime(start_date, '%Y-%m-%d').date()
        date_to = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        date_from = today
        date_to = today
    
    # Get bills in date range
    bills = Bill.objects.filter(
        brand=Brand,
        status='paid',
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    )
    
    # Calculate metrics
    total_revenue = bills.aggregate(total=Sum('total'))['total'] or Decimal('0')
    total_bills = bills.count()
    avg_bill = total_revenue / total_bills if total_bills > 0 else Decimal('0')
    
    # Subtotals breakdown
    subtotal_sum = bills.aggregate(sum=Sum('subtotal'))['sum'] or Decimal('0')
    discount_sum = bills.aggregate(sum=Sum('discount_amount'))['sum'] or Decimal('0')
    tax_sum = bills.aggregate(sum=Sum('tax_amount'))['sum'] or Decimal('0')
    service_sum = bills.aggregate(sum=Sum('service_charge'))['sum'] or Decimal('0')
    
    # Daily breakdown
    from django.db.models.functions import TruncDate
    daily_sales = bills.annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        revenue=Sum('total'),
        count=Count('id')
    ).order_by('date')
    
    # Comparison with previous period
    period_days = (date_to - date_from).days + 1
    prev_start = date_from - timedelta(days=period_days)
    prev_end = date_from - timedelta(days=1)
    
    prev_bills = Bill.objects.filter(
        brand=Brand,
        status='paid',
        created_at__date__gte=prev_start,
        created_at__date__lte=prev_end
    )
    
    prev_revenue = prev_bills.aggregate(total=Sum('total'))['total'] or Decimal('0')
    revenue_growth = ((total_revenue - prev_revenue) / prev_revenue * 100) if prev_revenue > 0 else Decimal('0')
    
    context = {
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'total_revenue': total_revenue,
        'total_bills': total_bills,
        'avg_bill': avg_bill,
        'subtotal_sum': subtotal_sum,
        'discount_sum': discount_sum,
        'tax_sum': tax_sum,
        'service_sum': service_sum,
        'daily_sales': list(daily_sales),
        'prev_revenue': prev_revenue,
        'revenue_growth': revenue_growth,
    }
    
    return render(request, 'management/reports/sales.html', context)


@manager_required
def products_report(request):
    """Top products and categories analysis"""
    store_config = Store.get_current()
    Brand = store_config.brand
    
    # Get date range
    period = request.GET.get('period', 'week')
    today = timezone.now().date()
    
    if period == 'today':
        date_from = today
    elif period == 'week':
        date_from = today - timedelta(days=7)
    elif period == 'month':
        date_from = today - timedelta(days=30)
    else:
        date_from = today - timedelta(days=7)
    
    # Top products by quantity
    from apps.pos.models import BillItem
    top_products_qty = BillItem.objects.filter(
        bill__brand=Brand,
        bill__status='paid',
        bill__created_at__date__gte=date_from,
        is_void=False
    ).values(
        'product__name',
        'product__category__name'
    ).annotate(
        total_qty=Sum('quantity'),
        total_revenue=Sum('total')
    ).order_by('-total_qty')[:20]
    
    # Top products by revenue
    top_products_revenue = BillItem.objects.filter(
        bill__brand=Brand,
        bill__status='paid',
        bill__created_at__date__gte=date_from,
        is_void=False
    ).values(
        'product__name',
        'product__category__name'
    ).annotate(
        total_qty=Sum('quantity'),
        total_revenue=Sum('total')
    ).order_by('-total_revenue')[:20]
    
    # Category performance
    category_performance = BillItem.objects.filter(
        bill__brand=Brand,
        bill__status='paid',
        bill__created_at__date__gte=date_from,
        is_void=False
    ).values(
        'product__category__name'
    ).annotate(
        total_qty=Sum('quantity'),
        total_revenue=Sum('total'),
        product_count=Count('product', distinct=True)
    ).order_by('-total_revenue')
    
    context = {
        'period': period,
        'date_from': date_from,
        'top_products_qty': list(top_products_qty),
        'top_products_revenue': list(top_products_revenue),
        'category_performance': list(category_performance),
    }
    
    return render(request, 'management/reports/products.html', context)


@manager_required
def cashier_report(request):
    """Cashier performance report"""
    store_config = Store.get_current()
    Brand = store_config.brand
    
    # Get date range
    period = request.GET.get('period', 'today')
    today = timezone.now().date()
    
    if period == 'today':
        date_from = today
        date_to = today
    elif period == 'week':
        date_from = today - timedelta(days=7)
        date_to = today
    elif period == 'month':
        date_from = today - timedelta(days=30)
        date_to = today
    else:
        date_from = today
        date_to = today
    
    # Cashier performance
    cashier_stats = Bill.objects.filter(
        brand=Brand,
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    ).values(
        'created_by__username',
        'created_by__first_name',
        'created_by__last_name'
    ).annotate(
        total_bills=Count('id'),
        paid_bills=Count('id', filter=Q(status='paid')),
        cancelled_bills=Count('id', filter=Q(status='cancelled')),
        total_revenue=Sum('total', filter=Q(status='paid')),
        avg_bill=Avg('total', filter=Q(status='paid'))
    ).order_by('-total_revenue')
    
    # Void items by cashier
    from apps.pos.models import BillItem
    void_stats = BillItem.objects.filter(
        bill__brand=Brand,
        bill__created_at__date__gte=date_from,
        bill__created_at__date__lte=date_to,
        is_void=True
    ).values(
        'bill__created_by__username'
    ).annotate(
        void_count=Count('id'),
        void_amount=Sum('total')
    )
    
    void_dict = {item['bill__created_by__username']: item for item in void_stats}
    
    # Merge void stats into cashier stats
    cashier_list = []
    for cashier in cashier_stats:
        username = cashier['created_by__username']
        cashier['void_count'] = void_dict.get(username, {}).get('void_count', 0)
        cashier['void_amount'] = void_dict.get(username, {}).get('void_amount', Decimal('0'))
        cashier_list.append(cashier)
    
    context = {
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'cashier_stats': cashier_list,
    }
    
    return render(request, 'management/reports/cashier.html', context)


@manager_required
def payment_report(request):
    """Payment method breakdown"""
    store_config = Store.get_current()
    Brand = store_config.brand
    
    # Get date range
    period = request.GET.get('period', 'today')
    today = timezone.now().date()
    
    if period == 'today':
        date_from = today
        date_to = today
    elif period == 'week':
        date_from = today - timedelta(days=7)
        date_to = today
    elif period == 'month':
        date_from = today - timedelta(days=30)
        date_to = today
    else:
        date_from = today
        date_to = today
    
    # Payment method breakdown
    payment_breakdown = Payment.objects.filter(
        bill__brand=Brand,
        bill__status='paid',
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    ).values('method').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('-total')
    
    # Daily payment trends
    from django.db.models.functions import TruncDate
    daily_payments = Payment.objects.filter(
        bill__brand=Brand,
        bill__status='paid',
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    ).annotate(
        date=TruncDate('created_at')
    ).values('date', 'method').annotate(
        total=Sum('amount')
    ).order_by('date', 'method')
    
    # Calculate totals
    total_amount = sum(item['total'] for item in payment_breakdown)
    
    context = {
        'period': period,
        'date_from': date_from,
        'date_to': date_to,
        'payment_breakdown': list(payment_breakdown),
        'daily_payments': list(daily_payments),
        'total_amount': total_amount,
    }
    
    return render(request, 'management/reports/payment.html', context)


@manager_required
def void_discount_report(request):
    """Void and discount analysis"""
    store_config = Store.get_current()
    Brand = store_config.brand
    
    # Get date range
    period = request.GET.get('period', 'week')
    today = timezone.now().date()
    
    if period == 'today':
        date_from = today
    elif period == 'week':
        date_from = today - timedelta(days=7)
    elif period == 'month':
        date_from = today - timedelta(days=30)
    else:
        date_from = today - timedelta(days=7)
    
    # Void items analysis
    from apps.pos.models import BillItem
    void_items = BillItem.objects.filter(
        bill__brand=Brand,
        bill__created_at__date__gte=date_from,
        is_void=True
    ).select_related('product', 'bill__created_by')
    
    void_summary = void_items.aggregate(
        count=Count('id'),
        total_amount=Sum('total')
    )
    
    # Void by product
    void_by_product = void_items.values(
        'product__name'
    ).annotate(
        count=Count('id'),
        amount=Sum('total')
    ).order_by('-count')[:10]
    
    # Void by cashier
    void_by_cashier = void_items.values(
        'bill__created_by__username'
    ).annotate(
        count=Count('id'),
        amount=Sum('total')
    ).order_by('-count')
    
    # Discount analysis
    bills_with_discount = Bill.objects.filter(
        brand=Brand,
        created_at__date__gte=date_from,
        discount_amount__gt=0
    )
    
    discount_summary = bills_with_discount.aggregate(
        count=Count('id'),
        total_discount=Sum('discount_amount')
    )
    
    # Discount by type (if you have discount reason field)
    discount_breakdown = bills_with_discount.values(
        'discount_percent'
    ).annotate(
        count=Count('id'),
        total=Sum('discount_amount')
    ).order_by('-total')
    
    context = {
        'period': period,
        'date_from': date_from,
        'void_summary': void_summary,
        'void_by_product': list(void_by_product),
        'void_by_cashier': list(void_by_cashier),
        'discount_summary': discount_summary,
        'discount_breakdown': list(discount_breakdown),
    }
    
    return render(request, 'management/reports/void_discount.html', context)


@manager_required
def peak_hours_report(request):
    """Peak hours and time-based analysis"""
    store_config = Store.get_current()
    Brand = store_config.brand
    
    # Get date range
    period = request.GET.get('period', 'week')
    today = timezone.now().date()
    
    if period == 'today':
        date_from = today
    elif period == 'week':
        date_from = today - timedelta(days=7)
    elif period == 'month':
        date_from = today - timedelta(days=30)
    else:
        date_from = today - timedelta(days=7)
    
    # Hourly sales distribution
    from django.db.models.functions import ExtractHour
    hourly_sales = Bill.objects.filter(
        brand=Brand,
        status='paid',
        created_at__date__gte=date_from
    ).annotate(
        hour=ExtractHour('created_at')
    ).values('hour').annotate(
        count=Count('id'),
        revenue=Sum('total')
    ).order_by('hour')
    
    # Day of week analysis
    from django.db.models.functions import ExtractWeekDay
    daily_sales = Bill.objects.filter(
        brand=Brand,
        status='paid',
        created_at__date__gte=date_from
    ).annotate(
        weekday=ExtractWeekDay('created_at')
    ).values('weekday').annotate(
        count=Count('id'),
        revenue=Sum('total')
    ).order_by('weekday')
    
    # Map weekday numbers to names
    weekday_names = {
        1: 'Sunday', 2: 'Monday', 3: 'Tuesday', 4: 'Wednesday',
        5: 'Thursday', 6: 'Friday', 7: 'Saturday'
    }
    
    daily_sales_list = []
    for item in daily_sales:
        item['weekday_name'] = weekday_names.get(item['weekday'], 'Unknown')
        daily_sales_list.append(item)
    
    # Find peak hour
    peak_hour = max(hourly_sales, key=lambda x: x['revenue']) if hourly_sales else None
    
    context = {
        'period': period,
        'date_from': date_from,
        'hourly_sales': list(hourly_sales),
        'daily_sales': daily_sales_list,
        'peak_hour': peak_hour,
    }
    
    return render(request, 'management/reports/peak_hours.html', context)


@manager_required
def export_sales_excel(request):
    """Export sales report to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    store_config = Store.get_current()
    Brand = store_config.brand
    
    # Get date range from request
    period = request.GET.get('period', 'today')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    today = timezone.now().date()
    
    # Calculate date range
    if period == 'today':
        date_from = today
        date_to = today
    elif period == 'week':
        date_from = today - timedelta(days=7)
        date_to = today
    elif period == 'month':
        date_from = today - timedelta(days=30)
        date_to = today
    elif period == 'custom' and start_date and end_date:
        date_from = datetime.strptime(start_date, '%Y-%m-%d').date()
        date_to = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        date_from = today
        date_to = today
    
    # Get bills
    bills = Bill.objects.filter(
        brand=Brand,
        status='paid',
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    ).order_by('created_at')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Title
    ws['A1'] = f"Sales Report: {Brand.name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f"Period: {date_from} to {date_to}"
    
    # Headers
    headers = ['Date', 'Bill Number', 'Table', 'Cashier', 'Subtotal', 'Discount', 'Tax', 'Service', 'Total', 'Payment Method']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row = 5
    for bill in bills:
        payment_methods = ', '.join([p.method for p in bill.payments.all()])
        
        ws.cell(row=row, column=1).value = bill.created_at.strftime('%Y-%m-%d %H:%M')
        ws.cell(row=row, column=2).value = bill.bill_number
        ws.cell(row=row, column=3).value = str(bill.table) if bill.table else '-'
        ws.cell(row=row, column=4).value = bill.created_by.username
        ws.cell(row=row, column=5).value = float(bill.subtotal)
        ws.cell(row=row, column=6).value = float(bill.discount_amount)
        ws.cell(row=row, column=7).value = float(bill.tax_amount)
        ws.cell(row=row, column=8).value = float(bill.service_charge)
        ws.cell(row=row, column=9).value = float(bill.total)
        ws.cell(row=row, column=10).value = payment_methods
        row += 1
    
    # Totals
    total_row = row + 1
    ws.cell(row=total_row, column=4).value = "TOTAL:"
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    
    total_revenue = bills.aggregate(total=Sum('total'))['total'] or Decimal('0')
    ws.cell(row=total_row, column=9).value = float(total_revenue)
    ws.cell(row=total_row, column=9).font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 20
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=sales_report_{date_from}_{date_to}.xlsx'
    
    wb.save(response)
    return response


@manager_required
def product_photos(request, product_id):
    """Product photo gallery management"""
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        # Handle photo upload
        image = request.FILES.get('image')
        caption = request.POST.get('caption', '')
        order = int(request.POST.get('order', 0))
        
        if image:
            # Validate file size (5MB max)
            if image.size > 5 * 1024 * 1024:
                return render(request, 'management/product_photos.html', {
                    'product': product,
                    'photos': product.photos.all(),
                    'error': 'File size must be less than 5MB'
                })
            
            # Create photo
            ProductPhoto.objects.create(
                product=product,
                image=image,
                caption=caption,
                order=order
            )
    
    photos = product.photos.all()
    
    return render(request, 'management/product_photos.html', {
        'product': product,
        'photos': photos,
    })


@manager_required
def product_photo_toggle(request, product_id, photo_id):
    """Toggle photo active status"""
    photo = get_object_or_404(ProductPhoto, id=photo_id, product_id=product_id)
    photo.is_active = not photo.is_active
    photo.save()
    
    from django.shortcuts import redirect
    return redirect('management:product_photos', product_id=product_id)


@manager_required
def product_photo_delete(request, product_id, photo_id):
    """Delete product photo"""
    photo = get_object_or_404(ProductPhoto, id=photo_id, product_id=product_id)
    photo.delete()
    
    from django.shortcuts import redirect
    return redirect('management:product_photos', product_id=product_id)


@manager_required
def product_detail(request, product_id):
    """View Product Detail"""
    store_config = Store.get_current()
    
    # Get all brands for this store
    from apps.core.models import StoreBrand
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True)
    brand_ids = list(store_brands.values_list('brand_id', flat=True))
    
    product = get_object_or_404(
        Product.objects.select_related('category', 'brand').prefetch_related(
            'product_modifiers__modifier__options', 'photos'
        ),
        id=product_id,
        brand_id__in=brand_ids
    )
    
    # MinIO settings for product images
    minio_endpoint = get_minio_endpoint_for_request(request)
    minio_bucket = 'product-images'
    
    context = {
        'product': product,
        'minio_endpoint': minio_endpoint,
        'minio_bucket': minio_bucket,
    }
    
    return render(request, 'management/product_detail.html', context)


@manager_required
@manager_required
def product_edit(request, product_id):
    """Edit Product"""
    store_config = Store.get_current()
    
    # Get all brands for this store
    from apps.core.models import StoreBrand
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True)
    brand_ids = list(store_brands.values_list('brand_id', flat=True))
    
    product = get_object_or_404(
        Product.objects.select_related('category', 'brand'),
        id=product_id,
        brand_id__in=brand_ids
    )
    
    if request.method == 'POST':
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Product edit POST request for {product_id}")
        logger.info(f"POST data: {request.POST}")
        
        try:
            # Update product
            name = request.POST.get('name', product.name)
            logger.info(f"Name: {name}")
            
            product.name = name
            product.sku = request.POST.get('sku', product.sku)
            product.description = request.POST.get('description', '')
            product.price = Decimal(request.POST.get('price', product.price))
            product.cost = Decimal(request.POST.get('cost', 0))
            product.stock_quantity = int(request.POST.get('stock_quantity', product.stock_quantity))
            product.low_stock_alert = int(request.POST.get('low_stock_alert', product.low_stock_alert))
            product.printer_target = request.POST.get('printer_target', product.printer_target)
            product.is_active = request.POST.get('is_active') == 'on'
            product.track_stock = request.POST.get('track_stock') == 'on'
            
            logger.info(f"Updated fields successfully")
            
            # Handle category
            category_id = request.POST.get('category')
            logger.info(f"Category ID: {category_id}")
            if category_id:
                try:
                    category = Category.objects.get(id=category_id, brand=Brand)
                    product.category = category
                    logger.info(f"Category updated to {category.name}")
                except Category.DoesNotExist:
                    logger.warning(f"Category {category_id} not found")
                    pass
            
            # Handle image upload
            if 'image' in request.FILES:
                # Delete old image if exists
                if product.image:
                    product.image.delete(save=False)
                product.image = request.FILES['image']
                logger.info(f"Image uploaded")
            
            product.save()
            logger.info(f"Product saved successfully")
            
            from django.contrib import messages
            from django.shortcuts import redirect
            messages.success(request, f'Product "{product.name}" updated successfully!')
            return redirect('management:products')
        except Exception as e:
            import traceback
            logger.error(f"Error updating product: {str(e)}")
            logger.error(traceback.format_exc())
            from django.http import HttpResponse
            return HttpResponse(f'Error: {str(e)}', status=400)
    
    # GET request - show form
    categories = Category.objects.filter(brand=Brand).order_by('name')
    
    context = {
        'product': product,
        'categories': categories,
    }
    
    return render(request, 'management/product_edit.html', context)


@manager_required
def import_excel_page(request):
    """Import Excel Page"""
    return render(request, 'management/import_excel.html')


@manager_required
def download_excel_template(request):
    """Download Excel Template for Product Import - Multi-Sheet (Products + Condiment Groups)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # ============================================
    # SHEET 1: PRODUCTS
    # ============================================
    ws = wb.active
    ws.title = "Products"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers with new structure
    headers = [
        'Category',
        'Menu Category', 
        'Nama Product',
        'PLU Product',
        'Printer Kitchen',     # Printer target (kitchen/bar/dessert/none)
        'Condiment Groups',    # 📝 Comma-separated group names (e.g., "Coffee Taste,Size,Ice Level")
        'Price Product',
        'Image Product'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Example data - NEW FORMAT: Comma-separated group references
    examples = [
        # Hot Americano - references 3 groups
        ['Beverage', 'Hot Coffee', 'Hot Americano', '03835821', 'bar', 'Coffee Taste,Size,Ice Level', 27000, 'avril/images/menus/americano.jpg'],
        
        # Hot Cappuccino - references 3 groups (Coffee Taste is REUSED!)
        ['Beverage', 'Hot Coffee', 'Hot Cappuccino', '03835822', 'bar', 'Coffee Taste,Size,Sugar Level', 33000, 'avril/images/menus/cappuccino.jpg'],
        
        # Ice Latte - references 3 groups (Coffee Taste and Size are REUSED!)
        ['Beverage', 'Ice Coffee', 'Ice Latte', '03835823', 'bar', 'Coffee Taste,Size,Ice Level', 35000, 'avril/images/menus/latte.jpg'],
        
        # Food with Spicy Level
        ['Food', 'Nasi Goreng', 'Nasi Goreng Special', '12345', 'kitchen', 'Spicy Level,Extra Topping', 35000, 'avril/images/menus/nasi_goreng.jpg'],
        
        # Food with multiple groups
        ['Food', 'Mie Goreng', 'Mie Goreng Special', '12346', 'kitchen', 'Spicy Level,Extra Topping', 30000, 'avril/images/menus/mie_goreng.jpg'],
        
        # Product without condiments (empty column is OK)
        ['Beverage', 'Bottled Drinks', 'Mineral Water', '99999', 'none', '', 5000, 'avril/images/menus/water.jpg'],
    ]
    
    # Write example data
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            if col_idx == 4:  # PLU Product - monospace
                cell.font = Font(name='Courier New')
    
    # Set column widths
    ws.column_dimensions['A'].width = 12  # Category
    ws.column_dimensions['B'].width = 15  # Menu Category
    ws.column_dimensions['C'].width = 22  # Nama Product
    ws.column_dimensions['D'].width = 12  # PLU Product
    ws.column_dimensions['E'].width = 18  # Printer Kitchen
    ws.column_dimensions['F'].width = 35  # Condiment Groups (comma-separated)
    ws.column_dimensions['G'].width = 13  # Price Product
    ws.column_dimensions['H'].width = 45  # Image Product
    
    # ============================================
    # SHEET 2: CONDIMENT GROUPS
    # ============================================
    ws_condiments = wb.create_sheet("Condiment Groups")
    
    # Header styling (green theme for condiments)
    condiment_header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    condiment_header_font = Font(color="FFFFFF", bold=True, size=11)
    
    condiment_headers = ['Group Name', 'Option Name', 'Fee', 'Is Required', 'Max Selections']
    
    # Write condiment headers
    for col, header in enumerate(condiment_headers, 1):
        cell = ws_condiments.cell(row=1, column=col)
        cell.value = header
        cell.fill = condiment_header_fill
        cell.font = condiment_header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Example condiment data
    condiment_examples = [
        ['Coffee Taste', 'Bold', 0, 'No', 1],
        ['Coffee Taste', 'Fruity', 0, 'No', 1],
        ['Spicy Level', 'No Chili', 0, 'No', 1],
        ['Spicy Level', 'Medium', 1000, 'No', 1],
        ['Spicy Level', 'Hot', 2000, 'No', 1],
        ['Size', 'Regular', 0, 'Yes', 1],
        ['Size', 'Large', 5000, 'Yes', 1],
        ['Sugar Level', 'No Sugar', 0, 'No', 1],
        ['Sugar Level', 'Normal', 0, 'No', 1],
        ['Ice Level', 'Less Ice', 0, 'No', 1],
        ['Ice Level', 'Normal', 0, 'No', 1],
        ['Milk Upgrade', 'Regular Milk', 0, 'No', 1],
        ['Milk Upgrade', 'Oat Milk', 8000, 'No', 1],
        ['Extra Topping', 'Telur', 5000, 'No', 3],
        ['Extra Topping', 'Keju', 8000, 'No', 3],
    ]
    
    # Write condiment example data
    for row_idx, row_data in enumerate(condiment_examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_condiments.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            if col_idx in [3, 4, 5]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set condiment column widths
    ws_condiments.column_dimensions['A'].width = 20  # Group Name
    ws_condiments.column_dimensions['B'].width = 25  # Option Name
    ws_condiments.column_dimensions['C'].width = 12  # Fee
    ws_condiments.column_dimensions['D'].width = 15  # Is Required
    ws_condiments.column_dimensions['E'].width = 18  # Max Selections
    
    # ============================================
    # SHEET 3: INSTRUCTIONS
    # ============================================
    ws2 = wb.create_sheet("Instructions")
    ws2['A1'] = "PRODUCT IMPORT TEMPLATE v4.0 - SINGLE-FILE MULTI-SHEET SYSTEM"
    ws2['A1'].font = Font(size=14, bold=True, color="10b981")
    
    instructions = [
        "",
        "SINGLE-FILE MULTI-SHEET IMPORT SYSTEM:",
        "",
        "This template has 2 sheets in ONE file:",
        "  - Sheet 1: Products (your menu items)",
        "  - Sheet 2: Condiment Groups (modifier groups & options)",
        "",
        "The system will automatically:",
        "  1. Process 'Condiment Groups' sheet FIRST (creates all groups & options)",
        "  2. Process 'Products' sheet SECOND (links products to groups)",
        "",
        "HOW TO USE:",
        "  Step 1: Fill 'Condiment Groups' sheet with your modifier groups",
        "  Step 2: Fill 'Products' sheet and reference groups by name (comma-separated)",
        "  Step 3: Upload ONE file - system handles both sheets automatically!",
        "",
        "PRODUCTS SHEET - Column Descriptions:",
        "",
        "1. Category: Main product category (e.g., Beverage, Food)",
        "2. Menu Category: Sub-category for grouping (e.g., Hot Coffee, Ice Coffee)",
        "3. Nama Product: Product name",
        "4. PLU Product: Product PLU/SKU code (unique identifier)",
        "5. Printer Kitchen: Printer target (kitchen/bar/dessert/none) - default: kitchen",
        "6. Condiment Groups: COMMA-SEPARATED group names (e.g., Coffee Taste,Size,Ice Level)",
        "7. Price Product: Base product price",
        "8. Image Product: Path to product image (e.g., avril/images/menus/product.jpg)",
        "",
        "CONDIMENT GROUPS SHEET - Column Descriptions:",
        "",
        "1. Group Name: Modifier group name (Coffee Taste, Spicy Level, Size)",
        "2. Option Name: Option within group (Bold, Medium, Large)",
        "3. Fee: Additional charge (0 = free, positive number = add cost)",
        "4. Is Required: Yes/No - must customer choose this group?",
        "5. Max Selections: How many options can be selected (usually 1, use 3+ for multiple toppings)",
        "",
        "EXAMPLE WORKFLOW:",
        "",
        "Condiment Groups sheet:",
        "  Coffee Taste | Bold   | 0    | No  | 1",
        "  Coffee Taste | Fruity | 0    | No  | 1",
        "  Size         | Regular| 0    | Yes | 1",
        "  Size         | Large  | 5000 | Yes | 1",
        "",
        "Products sheet:",
        "  Beverage | Hot Coffee | Americano  | 001 | bar | Coffee Taste,Size | 27000 | ...",
        "  Beverage | Hot Coffee | Cappuccino | 002 | bar | Coffee Taste,Size | 33000 | ...",
        "",
        "CONDIMENT GROUPS COLUMN (in Products sheet):",
        "",
        "  - Comma-separated list of group names (NO spaces after comma)",
        "  - Reference groups defined in 'Condiment Groups' sheet",
        "  - Leave empty if product has no modifiers",
        "  - Examples:",
        "      Coffee Taste,Size (Product uses 2 groups)",
        "      Spicy Level,Extra Topping (Food with 2 groups)",
        "      (empty) (Product has no modifiers)",
        "",
        "PRINTER KITCHEN OPTIONS:",
        "",
        "  kitchen: Food items (Nasi Goreng, Mie Goreng, Steak)",
        "  bar: Beverages (Coffee, Juice, Cocktails)",
        "  dessert: Desserts (Cake, Ice Cream, Pudding)",
        "  none: Items that don't need kitchen printing",
        "",
        "IMPORTANT NOTES:",
        "",
        "  - ONE FILE: All data in one Excel file with multiple sheets",
        "  - AUTOMATIC: System processes both sheets in correct order",
        "  - One row = One product (much cleaner!)",
        "  - Groups are REUSABLE - define once, use everywhere",
        "  - Update group once affects ALL products using it",
        "  - Images must be placed in media/products/ directory before import",
        "",
        "BENEFITS:",
        "",
        "  - Simpler workflow: Upload ONE file instead of TWO",
        "  - No import order confusion: System handles it automatically",
        "  - Cleaner Excel: 1 row per product (vs 10+ rows in old format)",
        "  - No duplication: Define 'Coffee Taste' once, use for all coffees",
        "  - Easy updates: Change group options affects all products",
        "  - Better organization: All data in one place",
        "",
        "COMMON MISTAKES:",
        "",
        "  - Don't fill Condiment Groups sheet: Products won't find them",
        "  - Typo in group name: System won't find the group",
        "  - Spaces after commas: Use 'Coffee Taste,Size' not 'Coffee Taste, Size'",
        "  - Wrong sheet names: Must be exactly 'Products' and 'Condiment Groups'",
        "",
        "BACKWARD COMPATIBILITY:",
        "",
        "  - The old two-file system still works!",
        "  - You can still import groups and products separately if needed",
        "  - But single-file is MUCH easier!",
    ]
    
    for idx, instruction in enumerate(instructions, 1):
        ws2[f'A{idx}'] = instruction
        if instruction and (instruction.isupper() or instruction.endswith(':')):
            ws2[f'A{idx}'].font = Font(bold=True, size=11, color="10b981")
        else:
            ws2[f'A{idx}'].font = Font(size=10)
    
    ws2.column_dimensions['A'].width = 100
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Product_Import_Template_v4_MultiSheet.xlsx'
    
    wb.save(response)
    return response


@manager_required
def download_condiment_groups_template(request):
    """Download Excel Template for Condiment Groups Import"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Condiment Groups"
    
    # Header styling
    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['Group Name', 'Option Name', 'Fee', 'Is Required', 'Max Selections']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Example data
    examples = [
        ['Coffee Taste', 'Bold', 0, 'No', 1],
        ['Coffee Taste', 'Fruity', 0, 'No', 1],
        ['Spicy Level', 'No Chili', 0, 'No', 1],
        ['Spicy Level', 'Medium', 1000, 'No', 1],
        ['Spicy Level', 'Hot', 2000, 'No', 1],
        ['Size', 'Regular', 0, 'Yes', 1],
        ['Size', 'Large', 5000, 'Yes', 1],
        ['Sugar Level', 'No Sugar', 0, 'No', 1],
        ['Sugar Level', 'Normal', 0, 'No', 1],
        ['Ice Level', 'Less Ice', 0, 'No', 1],
        ['Ice Level', 'Normal', 0, 'No', 1],
        ['Milk Upgrade', 'Regular Milk', 0, 'No', 1],
        ['Milk Upgrade', 'Oat Milk', 8000, 'No', 1],
        ['Extra Topping', 'Telur', 5000, 'No', 3],
        ['Extra Topping', 'Keju', 8000, 'No', 3],
    ]
    
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            if col_idx in [3, 4, 5]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    
    # Instructions sheet
    ws2 = wb.create_sheet(title="Instructions")
    instructions = [
        "",
        "Condiment Groups Import Instructions",
        "",
        "1. Group Name: Modifier group name (Coffee Taste, Spicy Level)",
        "2. Option Name: Option within group (Bold, Medium, Large)",
        "3. Fee: Additional charge (0 = free)",
        "4. Is Required: Yes/No (must customer choose?)",
        "5. Max Selections: How many options can be selected (usually 1)",
        "",
        "Groups are REUSABLE across products!",
        "Import groups FIRST, then products will reference them.",
    ]
    
    for idx, instruction in enumerate(instructions, 1):
        ws2[f'A{idx}'] = instruction
        ws2[f'A{idx}'].font = Font(size=10)
    
    ws2.column_dimensions['A'].width = 100
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Condiment_Groups_Template.xlsx'
    
    wb.save(response)
    return response


@manager_required
def import_condiment_groups(request):
    """Condiment Groups Import Page"""
    from apps.core.models import Modifier
    
    store_config = Store.get_current()
    Brand = store_config.brand
    groups_count = Modifier.objects.filter(brand=Brand).count()
    
    context = {'groups_count': groups_count}
    return render(request, 'management/import_condiment_groups.html', context)


@manager_required
def import_condiment_groups_process(request):
    """Process Condiment Groups Excel Import"""
    from django.contrib import messages
    from django.shortcuts import redirect
    import openpyxl
    from apps.core.models import Modifier, ModifierOption, ProductModifier
    
    if request.method != 'POST':
        return redirect('management:import_condiment_groups')
    
    if 'excel_file' not in request.FILES:
        messages.error(request, 'Please select an Excel file')
        return redirect('management:import_condiment_groups')
    
    excel_file = request.FILES['excel_file']
    update_existing = request.POST.get('update_existing') == 'on'
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        
        store_config = Store.get_current()
        Brand = store_config.brand
        
        stats = {'groups': 0, 'options': 0, 'updated': 0, 'errors': []}
        
        headers = [cell.value for cell in sheet[1]]
        col_indices = {}
        for idx, header in enumerate(headers):
            if header:
                col_indices[header.strip()] = idx
        
        required_cols = ['Group Name', 'Option Name', 'Fee']
        missing_cols = [col for col in required_cols if col not in col_indices]
        if missing_cols:
            messages.error(request, f'Missing required columns: {", ".join(missing_cols)}')
            return redirect('management:import_condiment_groups')
        
        processed_groups = {}
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            
            try:
                group_name = row[col_indices['Group Name']]
                option_name = row[col_indices['Option Name']]
                fee = row[col_indices['Fee']]
                is_required = row[col_indices.get('Is Required', -1)] if 'Is Required' in col_indices else None
                max_selections = row[col_indices.get('Max Selections', -1)] if 'Max Selections' in col_indices else None
                
                if not group_name or not option_name:
                    continue
                
                group_name = str(group_name).strip()
                option_name = str(option_name).strip()
                
                if group_name not in processed_groups:
                    is_req = False
                    if is_required:
                        is_req_str = str(is_required).strip().lower()
                        is_req = is_req_str in ['yes', 'y', 'true', '1']
                    
                    max_sel = 1
                    if max_selections:
                        try:
                            max_sel = int(max_selections)
                        except:
                            max_sel = 1
                    
                    modifier, created = Modifier.objects.get_or_create(
                        name=group_name,
                        brand=Brand,
                        defaults={'is_required': is_req, 'max_selections': max_sel}
                    )
                    
                    if created:
                        stats['groups'] += 1
                    elif update_existing:
                        modifier.is_required = is_req
                        modifier.max_selections = max_sel
                        modifier.save()
                        stats['updated'] += 1
                    
                    processed_groups[group_name] = modifier
                else:
                    modifier = processed_groups[group_name]
                
                option, opt_created = ModifierOption.objects.get_or_create(
                    modifier=modifier,
                    name=option_name,
                    defaults={'price_adjustment': Decimal(str(fee)) if fee else Decimal('0')}
                )
                
                if opt_created:
                    stats['options'] += 1
                elif update_existing:
                    new_fee = Decimal(str(fee)) if fee else Decimal('0')
                    if option.price_adjustment != new_fee:
                        option.price_adjustment = new_fee
                        option.save()
                        stats['updated'] += 1
                
            except Exception as e:
                stats['errors'].append(f'Row {row_idx}: {str(e)}')
                continue
        
        msg_parts = []
        if stats['groups'] > 0:
            msg_parts.append(f"{stats['groups']} groups created")
        if stats['options'] > 0:
            msg_parts.append(f"{stats['options']} options added")
        if stats['updated'] > 0:
            msg_parts.append(f"{stats['updated']} items updated")
        
        if msg_parts:
            messages.success(request, f'Import successful: {", ".join(msg_parts)}')
        
        if stats['errors']:
            for error in stats['errors'][:5]:
                messages.warning(request, error)
        
        return redirect('management:import_condiment_groups')
        
    except Exception as e:
        messages.error(request, f'Import failed: {str(e)}')
        return redirect('management:import_condiment_groups')


@csrf_exempt
@require_POST
@manager_required
def import_excel_reset(request):
    """Reset all products, categories, and modifiers"""
    from apps.core.models import Modifier, ModifierOption
    
    try:
        store_config = Store.get_current()
        Brand = store_config.brand
        
        # Count before deletion
        modifier_count = Modifier.objects.filter(brand=Brand).count()
        product_count = Product.objects.filter(category__brand=Brand).count()
        category_count = Category.objects.filter(brand=Brand).count()
        
        # Delete in correct order to avoid FK constraints
        # 1. Clear product-modifier relationships first
        ProductModifier.objects.filter(modifier__brand=Brand).delete()
        
        # 2. Delete modifier options
        modifier_ids = list(Modifier.objects.filter(brand=Brand).values_list('id', flat=True))
        ModifierOption.objects.filter(modifier_id__in=modifier_ids).delete()
        
        # 3. Delete modifiers
        Modifier.objects.filter(brand=Brand).delete()
        
        # 4. Delete products
        Product.objects.filter(category__brand=Brand).delete()
        
        # 5. Delete categories
        Category.objects.filter(brand=Brand).delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully deleted {category_count} categories, {product_count} products, and {modifier_count} modifiers'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@manager_required
def import_excel_process(request):
    """Process Excel Import - Multi-Sheet Support (Condiment Groups + Products)"""
    from django.contrib import messages
    from django.shortcuts import redirect
    import openpyxl
    import os
    from pathlib import Path
    
    if request.method != 'POST':
        return redirect('management:import_excel')
    
    if 'excel_file' not in request.FILES:
        messages.error(request, 'Please select an Excel file')
        return redirect('management:import_excel')
    
    excel_file = request.FILES['excel_file']
    skip_duplicates = request.POST.get('skip_duplicates') == 'on'
    update_existing = request.POST.get('update_existing') == 'on'
    create_modifiers = request.POST.get('create_modifiers') == 'on'
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_file)
        
        # Detect file type: Multi-sheet or Single-sheet
        sheet_names = wb.sheetnames
        has_condiment_sheet = 'Condiment Groups' in sheet_names
        has_products_sheet = 'Products' in sheet_names
        
        # Determine which sheet to use for products
        if has_products_sheet:
            sheet = wb['Products']
        else:
            sheet = wb.active  # Fallback to active sheet (backward compatibility)
        
        store_config = Store.get_current()
        Brand = store_config.brand
        
        stats = {
            'categories': 0,
            'products': 0,
            'modifiers': 0,
            'modifier_groups': 0,
            'modifier_options': 0,
            'skipped': 0,
            'updated': 0,
            'errors': []
        }
        
        # ============================================
        # STEP 1: PROCESS CONDIMENT GROUPS SHEET (if exists)
        # ============================================
        from apps.core.models import Modifier, ModifierOption, ProductModifier
        
        if has_condiment_sheet:
            messages.info(request, '🔄 Processing Condiment Groups sheet...')
            condiment_sheet = wb['Condiment Groups']
            
            # Get condiment headers
            condiment_headers = [cell.value for cell in condiment_sheet[1]]
            condiment_col_indices = {}
            for idx, header in enumerate(condiment_headers):
                if header:
                    condiment_col_indices[header.strip()] = idx
            
            # Required columns for condiments
            required_condiment_cols = ['Group Name', 'Option Name', 'Fee']
            missing_condiment_cols = [col for col in required_condiment_cols if col not in condiment_col_indices]
            
            if not missing_condiment_cols:
                processed_groups = {}
                
                for row_idx, row in enumerate(condiment_sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not any(row):
                        continue
                    
                    try:
                        group_name = row[condiment_col_indices['Group Name']]
                        option_name = row[condiment_col_indices['Option Name']]
                        fee = row[condiment_col_indices['Fee']]
                        is_required = row[condiment_col_indices.get('Is Required', -1)] if 'Is Required' in condiment_col_indices else None
                        max_selections = row[condiment_col_indices.get('Max Selections', -1)] if 'Max Selections' in condiment_col_indices else None
                        
                        if not group_name or not option_name:
                            continue
                        
                        group_name = str(group_name).strip()
                        option_name = str(option_name).strip()
                        
                        # Create or get modifier group
                        if group_name not in processed_groups:
                            is_req = False
                            if is_required:
                                is_req_str = str(is_required).strip().lower()
                                is_req = is_req_str in ['yes', 'y', 'true', '1']
                            
                            max_sel = 1
                            if max_selections:
                                try:
                                    max_sel = int(max_selections)
                                except:
                                    max_sel = 1
                            
                            modifier, mod_created = Modifier.objects.get_or_create(
                                name=group_name,
                                brand=Brand,
                                defaults={'is_required': is_req, 'max_selections': max_sel}
                            )
                            
                            if mod_created:
                                stats['modifier_groups'] += 1
                            elif update_existing:
                                modifier.is_required = is_req
                                modifier.max_selections = max_sel
                                modifier.save()
                                stats['updated'] += 1
                            
                            processed_groups[group_name] = modifier
                        else:
                            modifier = processed_groups[group_name]
                        
                        # Create or update modifier option
                        option, opt_created = ModifierOption.objects.get_or_create(
                            modifier=modifier,
                            name=option_name,
                            defaults={'price_adjustment': Decimal(str(fee)) if fee else Decimal('0')}
                        )
                        
                        if opt_created:
                            stats['modifier_options'] += 1
                        elif update_existing:
                            new_fee = Decimal(str(fee)) if fee else Decimal('0')
                            if option.price_adjustment != new_fee:
                                option.price_adjustment = new_fee
                                option.save()
                                stats['updated'] += 1
                    
                    except Exception as e:
                        stats['errors'].append(f'Condiment Groups Row {row_idx}: {str(e)}')
                        continue
                
                messages.success(request, f'✅ Condiment Groups: {stats["modifier_groups"]} groups, {stats["modifier_options"]} options processed')
            else:
                messages.warning(request, f'⚠️ Condiment Groups sheet missing columns: {missing_condiment_cols}. Skipping condiments.')
        
        # ============================================
        # STEP 2: PROCESS PRODUCTS SHEET
        # ============================================
        messages.info(request, '🔄 Processing Products sheet...')
        
        # Get header row
        headers = [cell.value for cell in sheet[1]]
        
        # Find column indices
        col_indices = {}
        for idx, header in enumerate(headers):
            if header:
                col_indices[header.strip()] = idx
        
        # Required columns (Printer Kitchen is optional, defaults to 'kitchen')
        required_cols = ['Category', 'Menu Category', 'Nama Product', 'PLU Product', 'Price Product', 'Image Product']
        missing_cols = [col for col in required_cols if col not in col_indices]
        if missing_cols:
            messages.error(request, f'Missing required columns: {", ".join(missing_cols)}')
            return redirect('management:import_excel')
        
        # Track modifiers by (product_id, group_name) to avoid duplicates
        processed_modifiers = {}
        
        # Process product rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            
            try:
                # Extract data
                category_name = row[col_indices['Category']]
                menu_category_name = row[col_indices['Menu Category']]
                product_name = row[col_indices['Nama Product']]
                plu_code = str(row[col_indices['PLU Product']]) if row[col_indices['PLU Product']] else ''
                price = row[col_indices['Price Product']]
                image_path = row[col_indices['Image Product']]
                
                # Optional printer kitchen column
                printer_target = row[col_indices.get('Printer Kitchen', -1)] if 'Printer Kitchen' in col_indices else None
                if printer_target:
                    printer_target = str(printer_target).strip().lower()
                    if printer_target not in ['kitchen', 'bar', 'dessert', 'none']:
                        printer_target = 'kitchen'
                else:
                    printer_target = 'kitchen'
                
                # NEW: Condiment Groups (comma-separated references)
                condiment_groups_str = row[col_indices.get('Condiment Groups', -1)] if 'Condiment Groups' in col_indices else None
                
                # BACKWARD COMPATIBILITY: Old format with inline condiments
                condiment_group = row[col_indices.get('Condiment Group', -1)] if 'Condiment Group' in col_indices else None
                condiment_name = row[col_indices.get('Condiment', -1)] if 'Condiment' in col_indices else None
                condiment_plu = row[col_indices.get('PLU Condiment', -1)] if 'PLU Condiment' in col_indices else None
                condiment_fee = row[col_indices.get('Fee Condiment', -1)] if 'Fee Condiment' in col_indices else 0
                
                if not condiment_group and condiment_name:
                    condiment_group = condiment_name.strip()
                
                if not product_name or not price:
                    continue
                
                # Create/Get Category
                category, created = Category.objects.get_or_create(
                    name=category_name,
                    brand=Brand,
                    defaults={'is_active': True, 'sort_order': 0}
                )
                if created:
                    stats['categories'] += 1
                
                # Check for existing product based on: Category + Menu Category (description) + Product Name + PLU
                existing_product = Product.objects.filter(
                    name=product_name,
                    category=category,
                    description=menu_category_name or '',
                    sku=plu_code
                ).first()
                
                if existing_product:
                    if skip_duplicates and not update_existing:
                        stats['skipped'] += 1
                        # Still create modifiers for skipped products
                        product = existing_product
                    elif update_existing:
                        # Update existing product
                        existing_product.name = product_name
                        existing_product.category = category
                        existing_product.sku = plu_code
                        existing_product.price = Decimal(str(price))
                        existing_product.description = menu_category_name or ''
                        existing_product.printer_target = printer_target
                        
                        # Handle image path
                        if image_path:
                            # Extract filename from path
                            filename = os.path.basename(image_path)
                            existing_product.image = f'products/{filename}'
                        
                        existing_product.save()
                        stats['updated'] += 1
                        product = existing_product
                    else:
                        # Allow duplicate - create new product
                        product = Product.objects.create(
                            name=product_name,
                            category=category,
                            sku=plu_code,
                            price=Decimal(str(price)),
                            description=menu_category_name or '',
                            stock_quantity=100,
                            is_active=True,
                            printer_target=printer_target
                        )
                        
                        # Handle image path
                        if image_path:
                            filename = os.path.basename(image_path)
                            product.image = f'products/{filename}'
                            product.save()
                        
                        stats['products'] += 1
                else:
                    # Create new product
                    product = Product.objects.create(
                        name=product_name,
                        category=category,
                        sku=plu_code,
                        price=Decimal(str(price)),
                        description=menu_category_name or '',
                        stock_quantity=100,
                        is_active=True,
                        printer_target=printer_target
                    )
                    
                    # Handle image path
                    if image_path:
                        # Extract filename from path
                        filename = os.path.basename(image_path)
                        product.image = f'products/{filename}'
                        product.save()
                    
                    stats['products'] += 1
                
                # NEW: Link product to existing condiment groups (comma-separated)
                if condiment_groups_str and str(condiment_groups_str).strip():
                    group_names = [g.strip() for g in str(condiment_groups_str).split(',') if g.strip()]
                    for group_name in group_names:
                        try:
                            # Find existing modifier by name
                            modifier = Modifier.objects.get(name=group_name, brand=Brand)
                            # Link product to modifier
                            product_modifier, created = ProductModifier.objects.get_or_create(
                                product=product,
                                modifier=modifier,
                                defaults={'sort_order': 0}
                            )
                            if created:
                                stats['modifiers'] += 1
                        except Modifier.DoesNotExist:
                            # Group not found - add to errors
                            error_msg = f'Row {row_idx}: Condiment group "{group_name}" not found. Import groups first!'
                            stats['errors'].append(error_msg)
                            continue
                
                # BACKWARD COMPATIBILITY: Create modifier if condiment data exists (old format)
                if create_modifiers and condiment_name and condiment_name.strip() and condiment_group and condiment_group.strip():
                    modifier_group_name = condiment_group.strip()
                    modifier_key = modifier_group_name
                    
                    if modifier_key not in processed_modifiers:
                        modifier, mod_created = Modifier.objects.get_or_create(
                            name=modifier_group_name,
                            brand=Brand,
                            defaults={
                                'is_required': False,
                                'max_selections': 1
                            }
                        )
                        processed_modifiers[modifier_key] = modifier
                    else:
                        modifier = processed_modifiers[modifier_key]
                    
                    # Add product to modifier (many-to-many relationship)
                    ProductModifier.objects.get_or_create(
                        product=product,
                        modifier=modifier,
                        defaults={'sort_order': 0}
                    )
                    
                    # Create modifier option (Note: ModifierOption doesn't have SKU field)
                    option, opt_created = ModifierOption.objects.get_or_create(
                        modifier=modifier,
                        name=condiment_name.strip(),
                        defaults={
                            'price_adjustment': Decimal(str(condiment_fee)) if condiment_fee else Decimal('0')
                        }
                    )
                    
                    # Update price if option already exists but has different fee
                    if not opt_created and condiment_fee is not None:
                        new_fee = Decimal(str(condiment_fee))
                        if option.price_adjustment != new_fee:
                            option.price_adjustment = new_fee
                            option.save()
                    
                    if opt_created:
                        stats['modifiers'] += 1
            
            except Exception as e:
                stats['errors'].append(f'Row {row_idx}: {str(e)}')
                continue
        
        # Success message
        message_parts = []
        if stats['modifier_groups'] > 0 or stats['modifier_options'] > 0:
            message_parts.append(f"📝 {stats['modifier_groups']} groups + {stats['modifier_options']} options")
        if stats['categories'] > 0:
            message_parts.append(f"{stats['categories']} categories")
        if stats['products'] > 0:
            message_parts.append(f"{stats['products']} products imported")
        if stats['updated'] > 0:
            message_parts.append(f"{stats['updated']} items updated")
        if stats['modifiers'] > 0:
            message_parts.append(f"{stats['modifiers']} product-modifier links")
        if stats['skipped'] > 0:
            message_parts.append(f"{stats['skipped']} products skipped")
        
        if message_parts:
            messages.success(request, '✅ Import completed! ' + ' | '.join(message_parts))
        
        if has_condiment_sheet:
            messages.info(request, '💡 Single-file multi-sheet import successful! Both sheets processed automatically.')
        
        if stats['errors']:
            for error in stats['errors'][:10]:  # Show first 10 errors
                messages.warning(request, error)
            if len(stats['errors']) > 10:
                messages.warning(request, f'... and {len(stats["errors"]) - 10} more errors')
        
        # Add instruction for image files
        messages.info(request, f'?? Remember to copy your image files to: media/products/ directory')
        
        return redirect('management:master_data')
    
    except Exception as e:
        messages.error(request, f'Error processing Excel file: {str(e)}')
        return redirect('management:import_excel')


# ==================== SESSION MANAGEMENT ====================

@login_required
@manager_required
def session_management(request):
    """Session management page for opening/closing store sessions"""
    store_config = Store.get_current()
    current_session = StoreSession.get_current(store_config)
    
    context = {
        'current_session': current_session,
        'today': timezone.now().date(),
    }
    
    return render(request, 'management/session_management.html', context)


@login_required
@manager_required
def session_close(request):
    """Close the current store session"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        store_config = Store.get_current()
        current_session = StoreSession.get_current(store_config)
        
        if not current_session:
            return JsonResponse({'success': False, 'error': 'No active session found'}, status=400)
        
        # Check for open cashier shifts
        from apps.core.models_session import CashierShift
        open_shifts = CashierShift.objects.filter(
            store_session=current_session,
            status='open'
        )

        if open_shifts.exists():
            shift_users = ', '.join([shift.cashier.get_full_name() or shift.cashier.username
                                     for shift in open_shifts])
            return JsonResponse({
                'success': False,
                'error': f'Cannot close session. Open shifts found for: {shift_users}'
            }, status=400)

        # Close the session
        closing_notes = request.POST.get('closing_notes', '').strip()
        current_session.closed_at = timezone.now()
        current_session.closed_by = request.user
        current_session.eod_notes = closing_notes
        current_session.status = 'closed'
        current_session.is_current = False
        current_session.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Session closed successfully at {current_session.closed_at.strftime("%H:%M")}'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ========================================
# CUSTOMER DISPLAY SLIDESHOW MANAGEMENT
# ========================================

@manager_required
def customer_display_slides(request):
    """
    Customer Display Slideshow Management
    View, upload, edit, delete slides
    """
    store_config, error_response = check_store_config(request, 'management/customer_display_slides.html')
    if error_response:
        return error_response
    
    # Get company and brand from store
    from apps.core.models import CustomerDisplaySlide
    
    # Get context brand if set
    context_brand_id = request.session.get('context_brand_id')
    
    # Build filter
    slides = CustomerDisplaySlide.objects.filter(
        company=store_config.company
    ).select_related('brand', 'store', 'created_by')
    
    # Filter by brand if context set
    if context_brand_id:
        slides = slides.filter(
            Q(brand_id=context_brand_id) | 
            Q(brand__isnull=True, store__isnull=True)
        )
    
    # Order by order field and created date
    slides = slides.order_by('order', '-created_at')
    
    # Get all brands for dropdown
    from apps.core.models import StoreBrand
    store_brands = StoreBrand.objects.filter(
        store=store_config,
        is_active=True
    ).select_related('brand')
    
    context = {
        'store_config': store_config,
        'slides': slides,
        'store_brands': store_brands,
        'context_brand_id': context_brand_id,
    }
    
    return render(request, 'management/customer_display_slides.html', context)


@manager_required
@require_POST
def customer_display_slide_upload(request):
    """
    Upload new slide to MinIO
    """
    try:
        store_config = Store.get_current()
        if not store_config:
            return JsonResponse({
                'success': False,
                'error': 'Store configuration not found'
            }, status=400)
        
        # Get form data
        image = request.FILES.get('image')
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        brand_id = request.POST.get('brand_id')
        store_id = request.POST.get('store_id')
        order = int(request.POST.get('order', 0))
        duration = int(request.POST.get('duration', 5))
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Validate
        if not image:
            return JsonResponse({
                'success': False,
                'error': 'Image file is required'
            }, status=400)
        
        if not title:
            return JsonResponse({
                'success': False,
                'error': 'Title is required'
            }, status=400)
        
        # Validate file size (10MB max)
        if image.size > 10 * 1024 * 1024:
            return JsonResponse({
                'success': False,
                'error': 'File size must be less than 10MB'
            }, status=400)
        
        # Validate file type
        allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif', 'image/webp']
        if image.content_type not in allowed_types:
            return JsonResponse({
                'success': False,
                'error': f'Invalid file type. Allowed: JPG, PNG, GIF, WebP'
            }, status=400)
        
        # Get brand and store objects
        brand = None
        if brand_id:
            brand = Brand.objects.filter(id=brand_id).first()
        
        store = None
        if store_id:
            store = Store.objects.filter(id=store_id).first()
        
        # Upload to MinIO
        from apps.core.minio_client import upload_to_minio
        import uuid
        
        # Generate unique filename
        file_ext = image.name.split('.')[-1]
        unique_id = str(uuid.uuid4())[:8]
        filename = f"slide_{unique_id}.{file_ext}"
        
        # Upload path
        bucket_name = 'customer-display'
        brand_code = brand.code if brand else 'all'
        object_path = f"{store_config.company.code}/{brand_code}/{filename}"
        
        # Upload to MinIO
        image_url = upload_to_minio(
            bucket_name=bucket_name,
            object_name=object_path,
            file_data=image.read(),
            content_type=image.content_type
        )
        
        # Create database record
        from apps.core.models import CustomerDisplaySlide
        
        slide = CustomerDisplaySlide.objects.create(
            company=store_config.company,
            brand=brand,
            store=store,
            title=title,
            description=description,
            image_url=image_url,
            image_path=object_path,
            order=order,
            duration_seconds=duration,
            is_active=True,
            start_date=start_date if start_date else None,
            end_date=end_date if end_date else None,
            created_by=request.user,
            updated_by=request.user
        )
        
        messages.success(request, f'Slide "{title}" uploaded successfully!')
        
        return JsonResponse({
            'success': True,
            'slide': {
                'id': slide.id,
                'title': slide.title,
                'image_url': slide.image_url,
                'order': slide.order
            }
        })
        
    except Exception as e:
        logger.error(f"Error uploading slide: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@manager_required
@require_POST
def customer_display_slide_update(request, slide_id):
    """
    Update slide metadata (not image)
    """
    try:
        from apps.core.models import CustomerDisplaySlide, Brand
        
        slide = get_object_or_404(CustomerDisplaySlide, id=slide_id)
        
        # Get form data
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        brand_id = request.POST.get('brand_id', '').strip()
        order = int(request.POST.get('order', 0))
        duration = int(request.POST.get('duration', 5))
        is_active = request.POST.get('is_active') == 'true'
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Update fields
        if title:
            slide.title = title
        slide.description = description
        
        # Update brand
        if brand_id:
            slide.brand = Brand.objects.get(id=brand_id)
        else:
            slide.brand = None
            
        slide.order = order
        slide.duration_seconds = duration
        slide.is_active = is_active
        slide.start_date = start_date if start_date else None
        slide.end_date = end_date if end_date else None
        slide.updated_by = request.user
        slide.save()
        
        messages.success(request, f'Slide "{slide.title}" updated successfully!')
        
        return JsonResponse({
            'success': True,
            'message': 'Slide updated successfully'
        })
        
    except Exception as e:
        logger.error(f"Error updating slide: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@manager_required
@require_POST
def customer_display_slide_delete(request, slide_id):
    """
    Delete slide from database and MinIO
    """
    try:
        from apps.core.models import CustomerDisplaySlide
        
        slide = get_object_or_404(CustomerDisplaySlide, id=slide_id)
        slide_title = slide.title
        
        # Delete from MinIO
        try:
            from apps.core.minio_client import delete_from_minio
            delete_from_minio(slide.image_path)
        except Exception as e:
            logger.warning(f"Failed to delete from MinIO: {e}")
            # Continue with database deletion even if MinIO fails
        
        # Delete from database
        slide.delete()
        
        messages.success(request, f'Slide "{slide_title}" deleted successfully!')
        
        return JsonResponse({
            'success': True,
            'message': 'Slide deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error deleting slide: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@manager_required
@require_POST
def customer_display_slide_toggle(request, slide_id):
    """
    Toggle slide active status
    """
    try:
        from apps.core.models import CustomerDisplaySlide
        
        slide = get_object_or_404(CustomerDisplaySlide, id=slide_id)
        slide.is_active = not slide.is_active
        slide.updated_by = request.user
        slide.save()
        
        status = 'activated' if slide.is_active else 'deactivated'
        messages.success(request, f'Slide "{slide.title}" {status}!')
        
        return JsonResponse({
            'success': True,
            'is_active': slide.is_active,
            'message': f'Slide {status}'
        })
        
    except Exception as e:
        logger.error(f"Error toggling slide: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ==========================================
# CUSTOMER DISPLAY CONFIG MANAGEMENT
# ==========================================

@manager_required
def display_config_list(request):
    """List all customer display configurations"""
    from apps.core.models import CustomerDisplayConfig
    
    store_config, error_response = check_store_config(request, 'management/display_config_list.html')
    if error_response:
        return error_response
    
    # Get configs for current store
    configs = CustomerDisplayConfig.objects.filter(
        company=store_config.company
    ).select_related('brand', 'store').order_by('-created_at')
    
    # Apply brand context filter
    context_brand_id = request.session.get('context_brand_id')
    if context_brand_id:
        configs = configs.filter(
            Q(brand_id=context_brand_id) | Q(brand__isnull=True)
        )
    
    context = {
        'store_config': store_config,
        'configs': configs,
        'context_brand_id': context_brand_id,
    }
    
    return render(request, 'management/display_config_list.html', context)


@manager_required
def display_config_create(request):
    """Create new customer display configuration"""
    from apps.core.models import CustomerDisplayConfig
    
    store_config, error_response = check_store_config(request, 'management/display_config_form.html')
    if error_response:
        return error_response
    
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')
    
    if request.method == 'POST':
        brand_name = request.POST.get('brand_name', '').strip()
        brand_logo = request.FILES.get('brand_logo')
        brand_logo_url = request.POST.get('brand_logo_url', '').strip()
        brand_tagline = request.POST.get('brand_tagline', '').strip()
        running_text = request.POST.get('running_text', '').strip()
        running_text_speed = request.POST.get('running_text_speed', '50')
        theme_primary_color = request.POST.get('theme_primary_color', '#4F46E5').strip()
        theme_secondary_color = request.POST.get('theme_secondary_color', '#10B981').strip()
        theme_text_color = request.POST.get('theme_text_color', '#1F2937').strip()
        theme_billing_bg = request.POST.get('theme_billing_bg', 'gradient').strip()
        theme_billing_text = request.POST.get('theme_billing_text', '#FFFFFF').strip()
        brand_id = request.POST.get('brand')
        store_id = request.POST.get('store')
        is_active = request.POST.get('is_active') == '1'
        
        # Validation
        if not brand_name:
            messages.error(request, 'Brand name is required')
        elif not running_text:
            messages.error(request, 'Running text is required')
        else:
            try:
                # Check unique constraint
                existing = CustomerDisplayConfig.objects.filter(
                    company=store_config.company,
                    brand_id=brand_id if brand_id else None,
                    store_id=store_id if store_id else None
                ).exists()
                
                if existing:
                    messages.error(request, 'Configuration already exists for this company/brand/store combination')
                else:
                    config = CustomerDisplayConfig(
                        company=store_config.company,
                        brand_id=brand_id if brand_id else None,
                        store_id=store_id if store_id else None,
                        brand_name=brand_name,
                        brand_logo=brand_logo,
                        brand_logo_url=brand_logo_url,
                        brand_tagline=brand_tagline,
                        running_text=running_text,
                        running_text_speed=int(running_text_speed),
                        theme_primary_color=theme_primary_color,
                        theme_secondary_color=theme_secondary_color,
                        theme_text_color=theme_text_color,
                        theme_billing_bg=theme_billing_bg,
                        theme_billing_text=theme_billing_text,
                        is_active=is_active,
                        created_by=request.user
                    )
                    config.save()
                    messages.success(request, f'Display config "{brand_name}" created successfully')
                    return redirect('management:display_config_list')
            except Exception as e:
                messages.error(request, f'Error creating config: {str(e)}')
    
    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'is_edit': False,
    }
    
    return render(request, 'management/display_config_form.html', context)


@manager_required
def display_config_edit(request, config_id):
    """Edit customer display configuration"""
    from apps.core.models import CustomerDisplayConfig
    
    store_config, error_response = check_store_config(request, 'management/display_config_form.html')
    if error_response:
        return error_response
    
    config = get_object_or_404(CustomerDisplayConfig, id=config_id, company=store_config.company)
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')
    
    if request.method == 'POST':
        brand_name = request.POST.get('brand_name', '').strip()
        brand_logo = request.FILES.get('brand_logo')
        brand_logo_url = request.POST.get('brand_logo_url', '').strip()
        brand_tagline = request.POST.get('brand_tagline', '').strip()
        running_text = request.POST.get('running_text', '').strip()
        running_text_speed = request.POST.get('running_text_speed', '50')
        theme_primary_color = request.POST.get('theme_primary_color', '#4F46E5').strip()
        theme_secondary_color = request.POST.get('theme_secondary_color', '#10B981').strip()
        theme_text_color = request.POST.get('theme_text_color', '#1F2937').strip()
        theme_billing_bg = request.POST.get('theme_billing_bg', 'gradient').strip()
        theme_billing_text = request.POST.get('theme_billing_text', '#FFFFFF').strip()
        brand_id = request.POST.get('brand')
        store_id = request.POST.get('store')
        is_active = request.POST.get('is_active') == '1'
        
        # Validation
        if not brand_name:
            messages.error(request, 'Brand name is required')
        elif not running_text:
            messages.error(request, 'Running text is required')
        else:
            try:
                # Check unique constraint (excluding current record)
                existing = CustomerDisplayConfig.objects.filter(
                    company=store_config.company,
                    brand_id=brand_id if brand_id else None,
                    store_id=store_id if store_id else None
                ).exclude(id=config_id).exists()
                
                if existing:
                    messages.error(request, 'Configuration already exists for this company/brand/store combination')
                else:
                    config.brand_id = brand_id if brand_id else None
                    config.store_id = store_id if store_id else None
                    config.brand_name = brand_name
                    if brand_logo:
                        config.brand_logo = brand_logo
                    config.brand_logo_url = brand_logo_url
                    config.brand_tagline = brand_tagline
                    config.running_text = running_text
                    config.running_text_speed = int(running_text_speed)
                    config.theme_primary_color = theme_primary_color
                    config.theme_secondary_color = theme_secondary_color
                    config.theme_text_color = theme_text_color
                    config.theme_billing_bg = theme_billing_bg
                    config.theme_billing_text = theme_billing_text
                    config.is_active = is_active
                    config.updated_by = request.user
                    config.save()
                    messages.success(request, f'Display config "{brand_name}" updated successfully')
                    return redirect('management:display_config_list')
            except Exception as e:
                messages.error(request, f'Error updating config: {str(e)}')
    
    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'config': config,
        'is_edit': True,
    }
    
    return render(request, 'management/display_config_form.html', context)


@manager_required
@require_POST
def display_config_delete(request, config_id):
    """Delete customer display configuration"""
    from apps.core.models import CustomerDisplayConfig
    
    store_config = Store.get_current()
    if not store_config:
        messages.error(request, 'Store configuration not found')
        return redirect('management:display_config_list')
    
    try:
        config = get_object_or_404(CustomerDisplayConfig, id=config_id, company=store_config.company)
        config_name = config.brand_name
        config.delete()
        messages.success(request, f'Display config "{config_name}" deleted successfully')
    except Exception as e:
        messages.error(request, f'Error deleting config: {str(e)}')
    
    return redirect('management:display_config_list')


@manager_required
@require_POST
def display_config_toggle(request, config_id):
    """Toggle customer display configuration active status"""
    from apps.core.models import CustomerDisplayConfig
    
    store_config = Store.get_current()
    if not store_config:
        return JsonResponse({'success': False, 'error': 'Store not configured'}, status=400)
    
    try:
        config = get_object_or_404(CustomerDisplayConfig, id=config_id, company=store_config.company)
        config.is_active = not config.is_active
        config.updated_by = request.user
        config.save()
        
        status = 'activated' if config.is_active else 'deactivated'
        messages.success(request, f'Config "{config.brand_name}" {status}!')
        
        return JsonResponse({
            'success': True,
            'is_active': config.is_active,
            'message': f'Config {status}'
        })
        
    except Exception as e:
        logger.error(f"Error toggling config: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ===================== Receipt Template Views =====================

@manager_required
def receipt_template_list(request):
    """List all receipt templates"""
    from apps.core.models import ReceiptTemplate
    
    store_config, error_response = check_store_config(request, 'management/receipt_template_list.html')
    if error_response:
        return error_response
    
    # Get templates for current store
    templates = ReceiptTemplate.objects.filter(
        company=store_config.company
    ).select_related('brand', 'store').order_by('-created_at')
    
    # Apply brand context filter
    context_brand_id = request.session.get('context_brand_id')
    if context_brand_id:
        templates = templates.filter(
            Q(brand_id=context_brand_id) | Q(brand__isnull=True)
        )
    
    context = {
        'store_config': store_config,
        'templates': templates,
        'context_brand_id': context_brand_id,
    }
    
    return render(request, 'management/receipt_template_list.html', context)


@manager_required
def receipt_template_create(request):
    """Create new receipt template"""
    from apps.core.models import ReceiptTemplate
    
    store_config, error_response = check_store_config(request, 'management/receipt_template_form.html')
    if error_response:
        return error_response
    
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')
    
    if request.method == 'POST':
        template_name = request.POST.get('template_name', '').strip()
        paper_width = request.POST.get('paper_width', '58')
        brand_id = request.POST.get('brand')
        store_id = request.POST.get('store')
        is_active = request.POST.get('is_active') == '1'
        
        # Header fields
        show_logo = request.POST.get('show_logo') == '1'
        header_line_1 = request.POST.get('header_line_1', '').strip()
        header_line_2 = request.POST.get('header_line_2', '').strip()
        header_line_3 = request.POST.get('header_line_3', '').strip()
        header_line_4 = request.POST.get('header_line_4', '').strip()
        
        # Content display options
        show_receipt_number = request.POST.get('show_receipt_number') == '1'
        show_date_time = request.POST.get('show_date_time') == '1'
        show_cashier_name = request.POST.get('show_cashier_name') == '1'
        show_customer_name = request.POST.get('show_customer_name') == '1'
        show_table_number = request.POST.get('show_table_number') == '1'
        show_item_code = request.POST.get('show_item_code') == '1'
        show_item_category = request.POST.get('show_item_category') == '1'
        show_modifiers = request.POST.get('show_modifiers') == '1'
        
        # Formatting
        price_alignment = request.POST.get('price_alignment', 'right')
        show_currency_symbol = request.POST.get('show_currency_symbol') == '1'
        
        # Summary section
        show_subtotal = request.POST.get('show_subtotal') == '1'
        show_tax = request.POST.get('show_tax') == '1'
        show_service_charge = request.POST.get('show_service_charge') == '1'
        show_discount = request.POST.get('show_discount') == '1'
        show_payment_method = request.POST.get('show_payment_method') == '1'
        show_paid_amount = request.POST.get('show_paid_amount') == '1'
        show_change = request.POST.get('show_change') == '1'
        
        # Footer
        footer_line_1 = request.POST.get('footer_line_1', '').strip()
        footer_line_2 = request.POST.get('footer_line_2', '').strip()
        footer_line_3 = request.POST.get('footer_line_3', '').strip()
        show_qr_payment = request.POST.get('show_qr_payment') == '1'
        
        # Print settings
        auto_print = request.POST.get('auto_print') == '1'
        auto_cut = request.POST.get('auto_cut') == '1'
        feed_lines = request.POST.get('feed_lines', '3')
        
        # Validation
        if not template_name:
            messages.error(request, 'Template name is required')
        else:
            try:
                template = ReceiptTemplate(
                    company=store_config.company,
                    brand_id=brand_id if brand_id else None,
                    store_id=store_id if store_id else None,
                    template_name=template_name,
                    is_active=is_active,
                    paper_width=int(paper_width),
                    show_logo=show_logo,
                    header_line_1=header_line_1,
                    header_line_2=header_line_2,
                    header_line_3=header_line_3,
                    header_line_4=header_line_4,
                    show_receipt_number=show_receipt_number,
                    show_date_time=show_date_time,
                    show_cashier_name=show_cashier_name,
                    show_customer_name=show_customer_name,
                    show_table_number=show_table_number,
                    show_item_code=show_item_code,
                    show_item_category=show_item_category,
                    show_modifiers=show_modifiers,
                    price_alignment=price_alignment,
                    show_currency_symbol=show_currency_symbol,
                    show_subtotal=show_subtotal,
                    show_tax=show_tax,
                    show_service_charge=show_service_charge,
                    show_discount=show_discount,
                    show_payment_method=show_payment_method,
                    show_paid_amount=show_paid_amount,
                    show_change=show_change,
                    footer_line_1=footer_line_1,
                    footer_line_2=footer_line_2,
                    footer_line_3=footer_line_3,
                    show_qr_payment=show_qr_payment,
                    auto_print=auto_print,
                    auto_cut=auto_cut,
                    feed_lines=int(feed_lines),
                    created_by=request.user
                )
                
                # Handle logo upload
                if 'logo' in request.FILES:
                    template.logo = request.FILES['logo']
                
                template.save()
                messages.success(request, f'Receipt template "{template_name}" created successfully')
                return redirect('management:receipt_template_list')
            except Exception as e:
                messages.error(request, f'Error creating template: {str(e)}')
    
    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'is_edit': False,
    }
    
    return render(request, 'management/receipt_template_form.html', context)


@manager_required
def receipt_template_edit(request, template_id):
    """Edit receipt template"""
    from apps.core.models import ReceiptTemplate
    
    store_config, error_response = check_store_config(request, 'management/receipt_template_form.html')
    if error_response:
        return error_response
    
    template = get_object_or_404(ReceiptTemplate, id=template_id, company=store_config.company)
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')
    
    if request.method == 'POST':
        template_name = request.POST.get('template_name', '').strip()
        paper_width = request.POST.get('paper_width', '58')
        brand_id = request.POST.get('brand')
        store_id = request.POST.get('store')
        is_active = request.POST.get('is_active') == '1'
        
        # Header fields
        show_logo = request.POST.get('show_logo') == '1'
        header_line_1 = request.POST.get('header_line_1', '').strip()
        header_line_2 = request.POST.get('header_line_2', '').strip()
        header_line_3 = request.POST.get('header_line_3', '').strip()
        header_line_4 = request.POST.get('header_line_4', '').strip()
        
        # Content display options
        show_receipt_number = request.POST.get('show_receipt_number') == '1'
        show_date_time = request.POST.get('show_date_time') == '1'
        show_cashier_name = request.POST.get('show_cashier_name') == '1'
        show_customer_name = request.POST.get('show_customer_name') == '1'
        show_table_number = request.POST.get('show_table_number') == '1'
        show_item_code = request.POST.get('show_item_code') == '1'
        show_item_category = request.POST.get('show_item_category') == '1'
        show_modifiers = request.POST.get('show_modifiers') == '1'
        
        # Formatting
        price_alignment = request.POST.get('price_alignment', 'right')
        show_currency_symbol = request.POST.get('show_currency_symbol') == '1'
        
        # Summary section
        show_subtotal = request.POST.get('show_subtotal') == '1'
        show_tax = request.POST.get('show_tax') == '1'
        show_service_charge = request.POST.get('show_service_charge') == '1'
        show_discount = request.POST.get('show_discount') == '1'
        show_payment_method = request.POST.get('show_payment_method') == '1'
        show_paid_amount = request.POST.get('show_paid_amount') == '1'
        show_change = request.POST.get('show_change') == '1'
        
        # Footer
        footer_line_1 = request.POST.get('footer_line_1', '').strip()
        footer_line_2 = request.POST.get('footer_line_2', '').strip()
        footer_line_3 = request.POST.get('footer_line_3', '').strip()
        show_qr_payment = request.POST.get('show_qr_payment') == '1'
        
        # Print settings
        auto_print = request.POST.get('auto_print') == '1'
        auto_cut = request.POST.get('auto_cut') == '1'
        feed_lines = request.POST.get('feed_lines', '3')
        
        # Validation
        if not template_name:
            messages.error(request, 'Template name is required')
        else:
            try:
                template.brand_id = brand_id if brand_id else None
                template.store_id = store_id if store_id else None
                template.template_name = template_name
                template.is_active = is_active
                template.paper_width = int(paper_width)
                template.show_logo = show_logo
                template.header_line_1 = header_line_1
                template.header_line_2 = header_line_2
                template.header_line_3 = header_line_3
                template.header_line_4 = header_line_4
                template.show_receipt_number = show_receipt_number
                template.show_date_time = show_date_time
                template.show_cashier_name = show_cashier_name
                template.show_customer_name = show_customer_name
                template.show_table_number = show_table_number
                template.show_item_code = show_item_code
                template.show_item_category = show_item_category
                template.show_modifiers = show_modifiers
                template.price_alignment = price_alignment
                template.show_currency_symbol = show_currency_symbol
                template.show_subtotal = show_subtotal
                template.show_tax = show_tax
                template.show_service_charge = show_service_charge
                template.show_discount = show_discount
                template.show_payment_method = show_payment_method
                template.show_paid_amount = show_paid_amount
                template.show_change = show_change
                template.footer_line_1 = footer_line_1
                template.footer_line_2 = footer_line_2
                template.footer_line_3 = footer_line_3
                template.show_qr_payment = show_qr_payment
                template.auto_print = auto_print
                template.auto_cut = auto_cut
                template.feed_lines = int(feed_lines)
                template.updated_by = request.user
                
                # Handle logo upload
                if 'logo' in request.FILES:
                    template.logo = request.FILES['logo']
                
                template.save()
                messages.success(request, f'Receipt template "{template_name}" updated successfully')
                return redirect('management:receipt_template_list')
            except Exception as e:
                messages.error(request, f'Error updating template: {str(e)}')
    
    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'template': template,
        'is_edit': True,
    }
    
    return render(request, 'management/receipt_template_form.html', context)


@manager_required
def receipt_template_duplicate(request, template_id):
    """Duplicate receipt template"""
    from apps.core.models import ReceiptTemplate
    
    store_config, error_response = check_store_config(request, 'management/receipt_template_form.html')
    if error_response:
        return error_response
    
    # Get original template
    original = get_object_or_404(ReceiptTemplate, id=template_id, company=store_config.company)
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')
    
    if request.method == 'POST':
        template_name = request.POST.get('template_name', '').strip()
        paper_width = request.POST.get('paper_width', '58')
        brand_id = request.POST.get('brand')
        store_id = request.POST.get('store')
        is_active = request.POST.get('is_active') == '1'
        
        # Header fields
        show_logo = request.POST.get('show_logo') == '1'
        header_line_1 = request.POST.get('header_line_1', '').strip()
        header_line_2 = request.POST.get('header_line_2', '').strip()
        header_line_3 = request.POST.get('header_line_3', '').strip()
        header_line_4 = request.POST.get('header_line_4', '').strip()
        
        # Content display options
        show_receipt_number = request.POST.get('show_receipt_number') == '1'
        show_date_time = request.POST.get('show_date_time') == '1'
        show_cashier_name = request.POST.get('show_cashier_name') == '1'
        show_customer_name = request.POST.get('show_customer_name') == '1'
        show_table_number = request.POST.get('show_table_number') == '1'
        show_item_code = request.POST.get('show_item_code') == '1'
        show_item_category = request.POST.get('show_item_category') == '1'
        show_modifiers = request.POST.get('show_modifiers') == '1'
        
        # Formatting
        price_alignment = request.POST.get('price_alignment', 'right')
        show_currency_symbol = request.POST.get('show_currency_symbol') == '1'
        
        # Summary section
        show_subtotal = request.POST.get('show_subtotal') == '1'
        show_tax = request.POST.get('show_tax') == '1'
        show_service_charge = request.POST.get('show_service_charge') == '1'
        show_discount = request.POST.get('show_discount') == '1'
        show_payment_method = request.POST.get('show_payment_method') == '1'
        show_paid_amount = request.POST.get('show_paid_amount') == '1'
        show_change = request.POST.get('show_change') == '1'
        
        # Footer
        footer_line_1 = request.POST.get('footer_line_1', '').strip()
        footer_line_2 = request.POST.get('footer_line_2', '').strip()
        footer_line_3 = request.POST.get('footer_line_3', '').strip()
        show_qr_payment = request.POST.get('show_qr_payment') == '1'
        
        # Print settings
        auto_print = request.POST.get('auto_print') == '1'
        auto_cut = request.POST.get('auto_cut') == '1'
        feed_lines = request.POST.get('feed_lines', '3')
        
        # Validation
        if not template_name:
            messages.error(request, 'Template name is required')
        else:
            try:
                template = ReceiptTemplate(
                    company=store_config.company,
                    brand_id=brand_id if brand_id else None,
                    store_id=store_id if store_id else None,
                    template_name=template_name,
                    is_active=is_active,
                    paper_width=int(paper_width),
                    show_logo=show_logo,
                    header_line_1=header_line_1,
                    header_line_2=header_line_2,
                    header_line_3=header_line_3,
                    header_line_4=header_line_4,
                    show_receipt_number=show_receipt_number,
                    show_date_time=show_date_time,
                    show_cashier_name=show_cashier_name,
                    show_customer_name=show_customer_name,
                    show_table_number=show_table_number,
                    show_item_code=show_item_code,
                    show_item_category=show_item_category,
                    show_modifiers=show_modifiers,
                    price_alignment=price_alignment,
                    show_currency_symbol=show_currency_symbol,
                    show_subtotal=show_subtotal,
                    show_tax=show_tax,
                    show_service_charge=show_service_charge,
                    show_discount=show_discount,
                    show_payment_method=show_payment_method,
                    show_paid_amount=show_paid_amount,
                    show_change=show_change,
                    footer_line_1=footer_line_1,
                    footer_line_2=footer_line_2,
                    footer_line_3=footer_line_3,
                    show_qr_payment=show_qr_payment,
                    auto_print=auto_print,
                    auto_cut=auto_cut,
                    feed_lines=int(feed_lines),
                    created_by=request.user
                )
                template.save()
                messages.success(request, f'Receipt template "{template_name}" duplicated successfully')
                return redirect('management:receipt_template_list')
            except Exception as e:
                messages.error(request, f'Error duplicating template: {str(e)}')
    
    # Pre-fill form with original template data but change name
    template = original
    template.template_name = f"{original.template_name} (Copy)"
    template.is_active = False  # Set duplicate as inactive by default
    
    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'template': template,
        'is_edit': False,
        'is_duplicate': True,
    }
    
    return render(request, 'management/receipt_template_form.html', context)


@manager_required
@require_POST
def receipt_template_delete(request, template_id):
    """Delete receipt template"""
    from apps.core.models import ReceiptTemplate
    
    store_config = Store.get_current()
    if not store_config:
        messages.error(request, 'Store configuration not found')
        return redirect('management:receipt_template_list')
    
    try:
        template = get_object_or_404(ReceiptTemplate, id=template_id, company=store_config.company)
        template_name = template.template_name
        template.delete()
        messages.success(request, f'Receipt template "{template_name}" deleted successfully')
    except Exception as e:
        messages.error(request, f'Error deleting template: {str(e)}')
    
    return redirect('management:receipt_template_list')


@manager_required
@require_POST
def receipt_template_toggle(request, template_id):
    """Toggle receipt template active status"""
    from apps.core.models import ReceiptTemplate
    
    store_config = Store.get_current()
    if not store_config:
        return JsonResponse({'success': False, 'error': 'Store not configured'}, status=400)
    
    try:
        template = get_object_or_404(ReceiptTemplate, id=template_id, company=store_config.company)
        template.is_active = not template.is_active
        template.updated_by = request.user
        template.save()
        
        status = 'activated' if template.is_active else 'deactivated'
        messages.success(request, f'Template "{template.template_name}" {status}!')
        
        return JsonResponse({
            'success': True,
            'is_active': template.is_active,
            'message': f'Template {status}'
        })
        
    except Exception as e:
        logger.error(f"Error toggling template: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@manager_required
@require_POST
def receipt_template_create_sample(request):
    """Create sample receipt template with default values"""
    from apps.core.models import ReceiptTemplate
    
    store_config = Store.get_current()
    if not store_config:
        messages.error(request, 'Store not configured')
        return redirect('management:receipt_template_list')
    
    try:
        # Check if sample already exists
        existing_sample = ReceiptTemplate.objects.filter(
            company=store_config.company,
            template_name__icontains='Sample Template'
        ).first()
        
        if existing_sample:
            messages.warning(request, f'Sample template already exists: "{existing_sample.template_name}"')
            return redirect('management:receipt_template_list')
        
        # Create sample template with comprehensive default values
        sample_template = ReceiptTemplate.objects.create(
            company=store_config.company,
            brand=None,  # Company-wide template
            store=None,
            
            # Template Info
            template_name='Sample Template - 58mm',
            is_active=False,  # Inactive by default
            
            # Paper Settings
            paper_width=58,
            
            # Header
            show_logo=True,
            header_line_1=store_config.company.name,
            header_line_2='Jl. Contoh No. 123, Jakarta',
            header_line_3='Telp: (021) 1234-5678',
            header_line_4='NPWP: 01.234.567.8-901.000',
            
            # Content Display Options
            show_receipt_number=True,
            show_date_time=True,
            show_cashier_name=True,
            show_customer_name=True,
            show_table_number=True,
            show_item_code=False,
            show_item_category=False,
            show_modifiers=True,
            
            # Formatting
            price_alignment='right',
            show_currency_symbol=True,
            
            # Summary Section
            show_subtotal=True,
            show_tax=True,
            show_service_charge=True,
            show_discount=True,
            show_payment_method=True,
            show_paid_amount=True,
            show_change=True,
            
            # Footer
            footer_line_1='Terima kasih atas kunjungan Anda!',
            footer_line_2='Barang yang sudah dibeli tidak dapat ditukar',
            footer_line_3='www.yogyagroup.com',
            show_qr_payment=False,
            
            # Print Settings
            auto_print=True,
            auto_cut=True,
            feed_lines=3,
            
            # Metadata
            created_by=request.user,
            updated_by=request.user
        )
        
        messages.success(
            request, 
            f'✅ Sample template created: "{sample_template.template_name}". '
            f'You can now edit it to customize for your needs.'
        )
        
        # Redirect to edit page
        return redirect('management:receipt_template_edit', template_id=sample_template.id)
        
    except Exception as e:
        logger.error(f"Error creating sample template: {str(e)}")
        messages.error(request, f'Failed to create sample template: {str(e)}')
        return redirect('management:receipt_template_list')


# =============================================================================
# STORE PRODUCT STOCK MANAGEMENT
# =============================================================================

@manager_required
def stock_management(request):
    """
    Store Product Stock Management - Daily stock tracking.
    Only products added here are stock-tracked.
    Products NOT in this table are always available.
    """
    from apps.pos.models import StoreProductStock

    store_config = Store.get_current()
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True)
    brand_ids = list(store_brands.values_list('brand_id', flat=True))

    # Get global context brand filter from session
    context_brand_id = request.session.get('context_brand_id', '')
    if context_brand_id and context_brand_id in [str(bid) for bid in brand_ids]:
        brand_ids = [context_brand_id]

    # Get stock records
    stocks = StoreProductStock.objects.filter(
        brand_id__in=brand_ids, is_active=True
    ).order_by('product_name')

    # Search filter
    search = request.GET.get('search', '')
    if search:
        stocks = stocks.filter(
            Q(product_name__icontains=search) | Q(product_sku__icontains=search)
        )

    # Status filter
    status_filter = request.GET.get('status', 'all')
    if status_filter == 'out_of_stock':
        # Filter in Python since remaining_stock is a property
        stock_list = [s for s in stocks if s.is_out_of_stock]
    elif status_filter == 'low_stock':
        stock_list = [s for s in stocks if s.is_low_stock]
    else:
        stock_list = list(stocks)

    # Get available products (not yet tracked) for "Add Product" dropdown
    tracked_skus = StoreProductStock.objects.filter(
        brand_id__in=brand_ids, is_active=True
    ).values_list('product_sku', flat=True)

    available_products = Product.objects.filter(
        brand_id__in=brand_ids, is_active=True
    ).exclude(sku__in=tracked_skus).order_by('name')

    brands_list = Brand.objects.filter(id__in=brand_ids).order_by('name')

    context = {
        'stocks': stock_list,
        'available_products': available_products,
        'brands': brands_list,
        'search': search,
        'status_filter': status_filter,
        'total_tracked': StoreProductStock.objects.filter(brand_id__in=brand_ids, is_active=True).count(),
        'total_out_of_stock': sum(1 for s in stocks if s.is_out_of_stock),
        'total_low_stock': sum(1 for s in stocks if s.is_low_stock),
        'store_config': store_config,
    }

    return render(request, 'management/stock_management.html', context)


@manager_required
@require_POST
def stock_add_product(request):
    """Add a product to stock tracking"""
    from apps.pos.models import StoreProductStock

    product_id = request.POST.get('product_id')
    daily_stock = request.POST.get('daily_stock', 0)
    low_stock_alert = request.POST.get('low_stock_alert', 5)

    if not product_id:
        messages.error(request, 'Please select a product')
        return redirect('management:stock_management')

    try:
        product = Product.objects.get(id=product_id)

        stock, created = StoreProductStock.objects.get_or_create(
            product_sku=product.sku,
            brand=product.brand,
            defaults={
                'product_id': product.id,
                'product_name': product.name,
                'daily_stock': Decimal(str(daily_stock)),
                'low_stock_alert': int(low_stock_alert),
                'last_reset_date': timezone.now().date(),
            }
        )

        if created:
            messages.success(request, f'"{product.name}" added to stock tracking with daily stock: {daily_stock}')
        else:
            # Reactivate if was deactivated
            if not stock.is_active:
                stock.is_active = True
                stock.daily_stock = Decimal(str(daily_stock))
                stock.sold_qty = Decimal('0')
                stock.low_stock_alert = int(low_stock_alert)
                stock.product_id = product.id
                stock.product_name = product.name
                stock.last_reset_date = timezone.now().date()
                stock.save()
                messages.success(request, f'"{product.name}" reactivated with daily stock: {daily_stock}')
            else:
                messages.warning(request, f'"{product.name}" is already being tracked')

    except Product.DoesNotExist:
        messages.error(request, 'Product not found')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')

    return redirect('management:stock_management')


@manager_required
@require_POST
def stock_update(request, stock_id):
    """Update stock record (daily_stock, low_stock_alert, sold_qty)"""
    from apps.pos.models import StoreProductStock

    try:
        stock = get_object_or_404(StoreProductStock, id=stock_id)

        update_fields = ['updated_at']

        daily_stock = request.POST.get('daily_stock')
        low_stock_alert = request.POST.get('low_stock_alert')
        sold_qty = request.POST.get('sold_qty')

        if daily_stock is not None:
            stock.daily_stock = Decimal(str(daily_stock))
            update_fields.append('daily_stock')
        if low_stock_alert is not None:
            stock.low_stock_alert = int(low_stock_alert)
            update_fields.append('low_stock_alert')
        if sold_qty is not None:
            stock.sold_qty = Decimal(str(sold_qty))
            update_fields.append('sold_qty')

        stock.save(update_fields=update_fields)
        messages.success(request, f'Stock updated for "{stock.product_name}"')

    except Exception as e:
        messages.error(request, f'Error: {str(e)}')

    return redirect('management:stock_management')


@manager_required
@require_POST
def stock_reset_daily(request, stock_id):
    """Reset sold_qty to 0 and optionally set new daily_stock"""
    from apps.pos.models import StoreProductStock

    try:
        stock = get_object_or_404(StoreProductStock, id=stock_id)
        new_stock = request.POST.get('daily_stock')

        if new_stock:
            stock.reset_daily(new_stock=Decimal(str(new_stock)))
        else:
            stock.reset_daily()

        messages.success(request, f'Daily stock reset for "{stock.product_name}"')

    except Exception as e:
        messages.error(request, f'Error: {str(e)}')

    return redirect('management:stock_management')


@manager_required
@require_POST
def stock_reset_all(request):
    """Reset ALL stock items - sold_qty to 0, keep daily_stock"""
    from apps.pos.models import StoreProductStock

    store_config = Store.get_current()
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True)
    brand_ids = list(store_brands.values_list('brand_id', flat=True))

    count = StoreProductStock.objects.filter(
        brand_id__in=brand_ids, is_active=True
    ).update(sold_qty=0, last_reset_date=timezone.now().date())

    messages.success(request, f'Daily reset complete: {count} products reset to full stock')
    return redirect('management:stock_management')


@manager_required
@require_POST
def stock_remove(request, stock_id):
    """Remove product from stock tracking (soft delete)"""
    from apps.pos.models import StoreProductStock

    try:
        stock = get_object_or_404(StoreProductStock, id=stock_id)
        stock.is_active = False
        stock.save(update_fields=['is_active', 'updated_at'])
        messages.success(request, f'"{stock.product_name}" removed from stock tracking')

    except Exception as e:
        messages.error(request, f'Error: {str(e)}')

    return redirect('management:stock_management')


@manager_required
@require_POST
def stock_sync_product_ids(request):
    """Re-match product_id after HO sync (run after daily sync)"""
    from apps.pos.models import StoreProductStock

    store_config = Store.get_current()
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True)
    brand_ids = list(store_brands.values_list('brand_id', flat=True))

    stocks = StoreProductStock.objects.filter(brand_id__in=brand_ids, is_active=True)
    updated = 0
    for stock in stocks:
        if stock.sync_product_id():
            updated += 1

    messages.success(request, f'Sync complete: {updated} product IDs updated, {stocks.count()} total checked')
    return redirect('management:stock_management')


# =============================================================================
# Payment Method Profiles
# =============================================================================

@manager_required
def payment_profile_list(request):
    """List all payment method profiles"""
    store_config, error_response = check_store_config(request, 'management/payment_profiles.html')
    if error_response:
        return error_response

    profiles = PaymentMethodProfile.objects.filter(
        company=store_config.company
    ).select_related('brand', 'store', 'media_group').prefetch_related('prompts').order_by('sort_order', 'name')

    # Use global brand filter - show brand-specific + company-wide
    context_brand_id = request.session.get('context_brand_id', '')
    if context_brand_id:
        profiles = profiles.filter(
            Q(brand_id=context_brand_id) | Q(brand__isnull=True)
        )

    context = {
        'store_config': store_config,
        'profiles': profiles,
        'context_brand_id': context_brand_id,
    }
    return render(request, 'management/payment_profiles.html', context)


def _clean_number_input(value, default=0):
    """Clean number input by removing dots/commas and converting to int"""
    if not value:
        return default
    try:
        # Remove dots and commas (thousand separators)
        cleaned = str(value).replace('.', '').replace(',', '').strip()
        return int(cleaned) if cleaned else default
    except (ValueError, AttributeError):
        return default


@manager_required
def payment_profile_create(request):
    """Create payment method profile with inline prompts"""
    store_config, error_response = check_store_config(request, 'management/payment_profile_form.html')
    if error_response:
        return error_response

    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')
    media_groups = MediaGroup.objects.filter(company=store_config.company, is_active=True)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        brand_id = request.POST.get('brand') or None
        store_id = request.POST.get('store') or None
        media_group_id = request.POST.get('media_group') or None
        media_id = request.POST.get('media_id', '0')
        color = request.POST.get('color', '#6b7280')
        sort_order = request.POST.get('sort_order', '0')
        smallest_denomination = request.POST.get('smallest_denomination', '0')
        allow_change = request.POST.get('allow_change') == 'on'
        open_cash_drawer = request.POST.get('open_cash_drawer') == 'on'
        legacy_method_id = request.POST.get('legacy_method_id', '')

        if not name:
            messages.error(request, 'Profile name is required')
        elif not code:
            messages.error(request, 'Profile code is required')
        else:
            # Check duplicate code within same scope
            dup_qs = PaymentMethodProfile.objects.filter(company=store_config.company, code=code)
            if brand_id:
                dup_qs = dup_qs.filter(brand_id=brand_id)
            else:
                dup_qs = dup_qs.filter(brand__isnull=True)
            if dup_qs.exists():
                scope = 'this brand' if brand_id else 'company-wide'
                messages.error(request, f'Profile with code "{code}" already exists for {scope}')
            else:
                try:
                    profile = PaymentMethodProfile.objects.create(
                        company=store_config.company,
                        brand_id=brand_id,
                        store_id=store_id if store_id else None,
                        media_group_id=media_group_id,
                        name=name,
                        code=code,
                        media_id=_clean_number_input(media_id, 0),
                        color=color,
                        sort_order=_clean_number_input(sort_order, 0),
                        smallest_denomination=_clean_number_input(smallest_denomination, 0),
                        allow_change=allow_change,
                        open_cash_drawer=open_cash_drawer,
                        legacy_method_id=legacy_method_id,
                    )
                    _save_prompts_from_post(request, profile)
                    messages.success(request, f'Payment profile "{name}" created successfully')
                    return redirect('management:payment_profile_list')
                except Exception as e:
                    messages.error(request, f'Error creating profile: {str(e)}')

    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'media_groups': media_groups,
        'is_edit': False,
        'legacy_method_choices': [
            ('cash', 'Cash'), ('card', 'Card'), ('qris', 'QRIS'),
            ('ewallet', 'E-Wallet'), ('transfer', 'Transfer'),
            ('voucher', 'Voucher'), ('debit', 'Debit'),
        ],
    }
    return render(request, 'management/payment_profile_form.html', context)


@manager_required
def payment_profile_edit(request, profile_id):
    """Edit payment method profile with inline prompts"""
    store_config, error_response = check_store_config(request, 'management/payment_profile_form.html')
    if error_response:
        return error_response

    profile = get_object_or_404(PaymentMethodProfile, id=profile_id, company=store_config.company)
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')
    media_groups = MediaGroup.objects.filter(company=store_config.company, is_active=True)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        brand_id = request.POST.get('brand') or None
        store_id = request.POST.get('store') or None
        media_group_id = request.POST.get('media_group') or None
        media_id = request.POST.get('media_id', '0')
        color = request.POST.get('color', '#6b7280')
        sort_order = request.POST.get('sort_order', '0')
        smallest_denomination = request.POST.get('smallest_denomination', '0')
        allow_change = request.POST.get('allow_change') == 'on'
        open_cash_drawer = request.POST.get('open_cash_drawer') == 'on'
        legacy_method_id = request.POST.get('legacy_method_id', '')

        if not name:
            messages.error(request, 'Profile name is required')
        elif not code:
            messages.error(request, 'Profile code is required')
        else:
            # Check duplicate code within same scope
            dup_qs = PaymentMethodProfile.objects.filter(company=store_config.company, code=code).exclude(id=profile.id)
            if brand_id:
                dup_qs = dup_qs.filter(brand_id=brand_id)
            else:
                dup_qs = dup_qs.filter(brand__isnull=True)
            if dup_qs.exists():
                scope = 'this brand' if brand_id else 'company-wide'
                messages.error(request, f'Profile with code "{code}" already exists for {scope}')
            else:
                try:
                    profile.brand_id = brand_id
                    profile.store_id = store_id if store_id else None
                    profile.media_group_id = media_group_id
                    profile.name = name
                    profile.code = code
                    profile.media_id = _clean_number_input(media_id, 0)
                    profile.color = color
                    profile.sort_order = _clean_number_input(sort_order, 0)
                    profile.smallest_denomination = _clean_number_input(smallest_denomination, 0)
                    profile.allow_change = allow_change
                    profile.open_cash_drawer = open_cash_drawer
                    profile.legacy_method_id = legacy_method_id
                    profile.save()
                    _save_prompts_from_post(request, profile)
                    messages.success(request, f'Payment profile "{name}" updated successfully')
                    return redirect('management:payment_profile_list')
                except Exception as e:
                    messages.error(request, f'Error updating profile: {str(e)}')

    import json
    prompts_list = list(profile.prompts.order_by('sort_order').values(
        'field_name', 'label', 'field_type', 'min_length', 'max_length',
        'placeholder', 'use_scanner', 'is_required',
    ))
    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'media_groups': media_groups,
        'profile': profile,
        'prompts': json.dumps(prompts_list),
        'is_edit': True,
        'legacy_method_choices': [
            ('cash', 'Cash'), ('card', 'Card'), ('qris', 'QRIS'),
            ('ewallet', 'E-Wallet'), ('transfer', 'Transfer'),
            ('voucher', 'Voucher'), ('debit', 'Debit'),
        ],
    }
    return render(request, 'management/payment_profile_form.html', context)


@manager_required
@require_POST
def payment_profile_delete(request, profile_id):
    """Delete payment method profile"""
    try:
        profile = get_object_or_404(PaymentMethodProfile, id=profile_id)
        name = profile.name
        profile.prompts.all().delete()
        profile.delete()
        messages.success(request, f'Payment profile "{name}" deleted successfully')
    except ProtectedError:
        messages.error(request, 'Cannot delete: this profile is used by existing payments')
    except Exception as e:
        messages.error(request, f'Error deleting profile: {str(e)}')
    return redirect('management:payment_profile_list')


@manager_required
@require_POST
def payment_profile_toggle(request, profile_id):
    """Toggle payment profile active status"""
    try:
        profile = get_object_or_404(PaymentMethodProfile, id=profile_id)
        profile.is_active = not profile.is_active
        profile.save(update_fields=['is_active'])
        status = 'activated' if profile.is_active else 'deactivated'
        return JsonResponse({'success': True, 'is_active': profile.is_active, 'message': f'Profile {status}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def _save_prompts_from_post(request, profile):
    """Parse prompt data from POST and save to profile"""
    import json
    profile.prompts.all().delete()

    prompts_json = request.POST.get('prompts_data', '[]')
    try:
        prompts = json.loads(prompts_json)
    except json.JSONDecodeError:
        prompts = []

    for i, p in enumerate(prompts):
        if not p.get('field_name'):
            continue
        DataEntryPrompt.objects.create(
            profile=profile,
            field_name=p.get('field_name', '').strip(),
            label=p.get('label', '').strip(),
            field_type=p.get('field_type', 'text'),
            min_length=_clean_number_input(p.get('min_length'), 0),
            max_length=_clean_number_input(p.get('max_length'), 99),
            placeholder=p.get('placeholder', '').strip(),
            use_scanner=p.get('use_scanner', False),
            is_required=p.get('is_required', False),
            sort_order=i,
        )


# ============================================================================
# MEDIA GROUPS
# ============================================================================

@manager_required
def media_group_list(request):
    """List all media groups"""
    store_config, error_response = check_store_config(request, 'management/media_groups.html')
    if error_response:
        return error_response

    groups = MediaGroup.objects.filter(company=store_config.company).order_by('sort_order', 'name')

    context = {
        'store_config': store_config,
        'groups': groups,
    }
    return render(request, 'management/media_groups.html', context)


@manager_required
def media_group_create(request):
    """Create new media group"""
    store_config, error_response = check_store_config(request, 'management/media_group_form.html')
    if error_response:
        return error_response

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        orafin_group = request.POST.get('orafin_group', '').strip()
        sort_order = request.POST.get('sort_order', '0')

        if not name:
            messages.error(request, 'Group name is required')
        elif not code:
            messages.error(request, 'Group code is required')
        elif MediaGroup.objects.filter(company=store_config.company, code=code).exists():
            messages.error(request, f'Media group with code "{code}" already exists')
        else:
            try:
                MediaGroup.objects.create(
                    company=store_config.company,
                    name=name,
                    code=code,
                    orafin_group=orafin_group,
                    sort_order=_clean_number_input(sort_order, 0),
                )
                messages.success(request, f'Media group "{name}" created successfully')
                return redirect('management:media_group_list')
            except Exception as e:
                messages.error(request, f'Error creating media group: {str(e)}')

    context = {
        'store_config': store_config,
        'is_edit': False,
    }
    return render(request, 'management/media_group_form.html', context)


@manager_required
def media_group_edit(request, group_id):
    """Edit media group"""
    store_config, error_response = check_store_config(request, 'management/media_group_form.html')
    if error_response:
        return error_response

    group = get_object_or_404(MediaGroup, id=group_id, company=store_config.company)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        orafin_group = request.POST.get('orafin_group', '').strip()
        sort_order = request.POST.get('sort_order', '0')

        if not name:
            messages.error(request, 'Group name is required')
        elif not code:
            messages.error(request, 'Group code is required')
        elif MediaGroup.objects.filter(company=store_config.company, code=code).exclude(id=group.id).exists():
            messages.error(request, f'Media group with code "{code}" already exists')
        else:
            try:
                group.name = name
                group.code = code
                group.orafin_group = orafin_group
                group.sort_order = _clean_number_input(sort_order, 0)
                group.save()
                messages.success(request, f'Media group "{name}" updated successfully')
                return redirect('management:media_group_list')
            except Exception as e:
                messages.error(request, f'Error updating media group: {str(e)}')

    context = {
        'store_config': store_config,
        'group': group,
        'is_edit': True,
    }
    return render(request, 'management/media_group_form.html', context)


@manager_required
@require_POST
def media_group_delete(request, group_id):
    """Delete media group"""
    try:
        group = get_object_or_404(MediaGroup, id=group_id)
        name = group.name
        group.delete()
        messages.success(request, f'Media group "{name}" deleted successfully')
    except ProtectedError:
        messages.error(request, 'Cannot delete: this group is used by payment profiles')
    except Exception as e:
        messages.error(request, f'Error deleting group: {str(e)}')
    return redirect('management:media_group_list')


@manager_required
@require_POST
def media_group_toggle(request, group_id):
    """Toggle media group active status"""
    try:
        group = get_object_or_404(MediaGroup, id=group_id)
        group.is_active = not group.is_active
        group.save(update_fields=['is_active'])
        status = 'activated' if group.is_active else 'deactivated'
        return JsonResponse({'success': True, 'is_active': group.is_active, 'message': f'Group {status}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =====================================================
# EFT Terminal Management
# =====================================================

@login_required
def eft_terminal_list(request):
    """List all EFT terminals"""
    store_config, error_response = check_store_config(request, 'management/eft_terminals.html')
    if error_response:
        return error_response

    terminals = EFTTerminal.objects.filter(company=store_config.company).order_by('sort_order', 'code')

    context = {
        'store_config': store_config,
        'terminals': terminals,
    }
    return render(request, 'management/eft_terminals.html', context)


@manager_required
def eft_terminal_create(request):
    """Create new EFT terminal"""
    store_config, error_response = check_store_config(request, 'management/eft_terminal_form.html')
    if error_response:
        return error_response

    if request.method == 'POST':
        code = request.POST.get('code', '').strip()
        name = request.POST.get('name', '').strip()
        sort_order = request.POST.get('sort_order', '0')

        if not code:
            messages.error(request, 'EFT code is required')
        elif not name:
            messages.error(request, 'Bank/terminal name is required')
        elif EFTTerminal.objects.filter(company=store_config.company, code=code).exists():
            messages.error(request, f'EFT terminal with code "{code}" already exists')
        else:
            try:
                EFTTerminal.objects.create(
                    company=store_config.company,
                    code=code,
                    name=name,
                    sort_order=_clean_number_input(sort_order, 0),
                )
                messages.success(request, f'EFT terminal "{code}: {name}" created successfully')
                return redirect('management:eft_terminal_list')
            except Exception as e:
                messages.error(request, f'Error creating EFT terminal: {str(e)}')

    context = {
        'store_config': store_config,
        'is_edit': False,
    }
    return render(request, 'management/eft_terminal_form.html', context)


@manager_required
def eft_terminal_edit(request, terminal_id):
    """Edit EFT terminal"""
    store_config, error_response = check_store_config(request, 'management/eft_terminal_form.html')
    if error_response:
        return error_response

    eft = get_object_or_404(EFTTerminal, id=terminal_id, company=store_config.company)

    if request.method == 'POST':
        code = request.POST.get('code', '').strip()
        name = request.POST.get('name', '').strip()
        sort_order = request.POST.get('sort_order', '0')

        if not code:
            messages.error(request, 'EFT code is required')
        elif not name:
            messages.error(request, 'Bank/terminal name is required')
        elif EFTTerminal.objects.filter(company=store_config.company, code=code).exclude(id=eft.id).exists():
            messages.error(request, f'EFT terminal with code "{code}" already exists')
        else:
            try:
                eft.code = code
                eft.name = name
                eft.sort_order = _clean_number_input(sort_order, 0)
                eft.save()
                messages.success(request, f'EFT terminal "{code}: {name}" updated successfully')
                return redirect('management:eft_terminal_list')
            except Exception as e:
                messages.error(request, f'Error updating EFT terminal: {str(e)}')

    context = {
        'store_config': store_config,
        'eft': eft,
        'is_edit': True,
    }
    return render(request, 'management/eft_terminal_form.html', context)


@manager_required
@require_POST
def eft_terminal_delete(request, terminal_id):
    """Delete EFT terminal"""
    try:
        eft = get_object_or_404(EFTTerminal, id=terminal_id)
        label = f"{eft.code}: {eft.name}"
        eft.delete()
        messages.success(request, f'EFT terminal "{label}" deleted successfully')
    except Exception as e:
        messages.error(request, f'Error deleting EFT terminal: {str(e)}')
    return redirect('management:eft_terminal_list')


@manager_required
@require_POST
def eft_terminal_toggle(request, terminal_id):
    """Toggle EFT terminal active status"""
    try:
        eft = get_object_or_404(EFTTerminal, id=terminal_id)
        eft.is_active = not eft.is_active
        eft.save(update_fields=['is_active'])
        status = 'activated' if eft.is_active else 'deactivated'
        return JsonResponse({'success': True, 'is_active': eft.is_active, 'message': f'EFT terminal {status}'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================================
# CHECKER TEMPLATES MANAGEMENT
# ============================================================================

@manager_required
def checker_template_list(request):
    """List all checker templates"""
    from apps.kitchen.models import CheckerTemplate

    store_config, error_response = check_store_config(request, 'management/checker_template_list.html')
    if error_response:
        return error_response

    templates = CheckerTemplate.objects.filter(
        company=store_config.company
    ).select_related('brand', 'store').order_by('-created_at')

    context_brand_id = request.session.get('context_brand_id')
    if context_brand_id:
        templates = templates.filter(
            Q(brand_id=context_brand_id) | Q(brand__isnull=True)
        )

    context = {
        'store_config': store_config,
        'templates': templates,
        'context_brand_id': context_brand_id,
    }

    return render(request, 'management/checker_template_list.html', context)


@manager_required
def checker_template_create(request):
    """Create new checker template"""
    from apps.kitchen.models import CheckerTemplate

    store_config, error_response = check_store_config(request, 'management/checker_template_form.html')
    if error_response:
        return error_response

    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')

    if request.method == 'POST':
        template_name = request.POST.get('template_name', '').strip()
        paper_width = request.POST.get('paper_width', '80')
        brand_id = request.POST.get('brand')
        store_id = request.POST.get('store')
        is_active = request.POST.get('is_active') == '1'

        # Header fields
        header_line_1 = request.POST.get('header_line_1', '').strip()
        header_line_2 = request.POST.get('header_line_2', '').strip()

        # Content display options
        show_bill_number = request.POST.get('show_bill_number') == '1'
        show_table_number = request.POST.get('show_table_number') == '1'
        show_date_time = request.POST.get('show_date_time') == '1'
        show_station_label = request.POST.get('show_station_label') == '1'
        show_item_notes = request.POST.get('show_item_notes') == '1'
        show_item_qty = request.POST.get('show_item_qty') == '1'
        show_checkbox = request.POST.get('show_checkbox') == '1'

        # Footer
        footer_line_1 = request.POST.get('footer_line_1', '').strip()
        footer_line_2 = request.POST.get('footer_line_2', '').strip()

        # Print settings
        auto_cut = request.POST.get('auto_cut') == '1'
        feed_lines = request.POST.get('feed_lines', '3')

        if not template_name:
            messages.error(request, 'Template name is required')
        else:
            try:
                template = CheckerTemplate(
                    company=store_config.company,
                    brand_id=brand_id if brand_id else None,
                    store_id=store_id if store_id else None,
                    template_name=template_name,
                    is_active=is_active,
                    paper_width=int(paper_width),
                    header_line_1=header_line_1,
                    header_line_2=header_line_2,
                    show_bill_number=show_bill_number,
                    show_table_number=show_table_number,
                    show_date_time=show_date_time,
                    show_station_label=show_station_label,
                    show_item_notes=show_item_notes,
                    show_item_qty=show_item_qty,
                    show_checkbox=show_checkbox,
                    footer_line_1=footer_line_1,
                    footer_line_2=footer_line_2,
                    auto_cut=auto_cut,
                    feed_lines=int(feed_lines),
                    created_by=request.user
                )
                template.save()
                messages.success(request, f'Checker template "{template_name}" created successfully')
                return redirect('management:checker_template_list')
            except Exception as e:
                messages.error(request, f'Error creating template: {str(e)}')

    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'is_edit': False,
    }

    return render(request, 'management/checker_template_form.html', context)


@manager_required
def checker_template_edit(request, template_id):
    """Edit checker template"""
    from apps.kitchen.models import CheckerTemplate

    store_config, error_response = check_store_config(request, 'management/checker_template_form.html')
    if error_response:
        return error_response

    template = get_object_or_404(CheckerTemplate, id=template_id, company=store_config.company)
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')

    if request.method == 'POST':
        template_name = request.POST.get('template_name', '').strip()
        paper_width = request.POST.get('paper_width', '80')
        brand_id = request.POST.get('brand')
        store_id = request.POST.get('store')
        is_active = request.POST.get('is_active') == '1'

        # Header fields
        header_line_1 = request.POST.get('header_line_1', '').strip()
        header_line_2 = request.POST.get('header_line_2', '').strip()

        # Content display options
        show_bill_number = request.POST.get('show_bill_number') == '1'
        show_table_number = request.POST.get('show_table_number') == '1'
        show_date_time = request.POST.get('show_date_time') == '1'
        show_station_label = request.POST.get('show_station_label') == '1'
        show_item_notes = request.POST.get('show_item_notes') == '1'
        show_item_qty = request.POST.get('show_item_qty') == '1'
        show_checkbox = request.POST.get('show_checkbox') == '1'

        # Footer
        footer_line_1 = request.POST.get('footer_line_1', '').strip()
        footer_line_2 = request.POST.get('footer_line_2', '').strip()

        # Print settings
        auto_cut = request.POST.get('auto_cut') == '1'
        feed_lines = request.POST.get('feed_lines', '3')

        if not template_name:
            messages.error(request, 'Template name is required')
        else:
            try:
                template.brand_id = brand_id if brand_id else None
                template.store_id = store_id if store_id else None
                template.template_name = template_name
                template.is_active = is_active
                template.paper_width = int(paper_width)
                template.header_line_1 = header_line_1
                template.header_line_2 = header_line_2
                template.show_bill_number = show_bill_number
                template.show_table_number = show_table_number
                template.show_date_time = show_date_time
                template.show_station_label = show_station_label
                template.show_item_notes = show_item_notes
                template.show_item_qty = show_item_qty
                template.show_checkbox = show_checkbox
                template.footer_line_1 = footer_line_1
                template.footer_line_2 = footer_line_2
                template.auto_cut = auto_cut
                template.feed_lines = int(feed_lines)
                template.updated_by = request.user
                template.save()
                messages.success(request, f'Checker template "{template_name}" updated successfully')
                return redirect('management:checker_template_list')
            except Exception as e:
                messages.error(request, f'Error updating template: {str(e)}')

    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'template': template,
        'is_edit': True,
    }

    return render(request, 'management/checker_template_form.html', context)


@manager_required
@require_POST
def checker_template_delete(request, template_id):
    """Delete checker template"""
    from apps.kitchen.models import CheckerTemplate

    store_config = Store.get_current()
    if not store_config:
        messages.error(request, 'Store configuration not found')
        return redirect('management:checker_template_list')

    try:
        template = get_object_or_404(CheckerTemplate, id=template_id, company=store_config.company)
        template_name = template.template_name
        template.delete()
        messages.success(request, f'Checker template "{template_name}" deleted successfully')
    except Exception as e:
        messages.error(request, f'Error deleting template: {str(e)}')

    return redirect('management:checker_template_list')


@manager_required
@require_POST
def checker_template_toggle(request, template_id):
    """Toggle checker template active status"""
    from apps.kitchen.models import CheckerTemplate

    store_config = Store.get_current()
    if not store_config:
        return JsonResponse({'success': False, 'error': 'Store not configured'}, status=400)

    try:
        template = get_object_or_404(CheckerTemplate, id=template_id, company=store_config.company)
        template.is_active = not template.is_active
        template.updated_by = request.user
        template.save()

        status = 'activated' if template.is_active else 'deactivated'
        messages.success(request, f'Template "{template.template_name}" {status}!')

        return JsonResponse({
            'success': True,
            'is_active': template.is_active,
            'message': f'Template {status}'
        })

    except Exception as e:
        logger.error(f"Error toggling checker template: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ==========================================
# Kitchen Ticket Template CRUD
# ==========================================

@manager_required
def kitchen_template_list(request):
    """List all kitchen ticket templates"""
    from apps.kitchen.models import KitchenTicketTemplate

    store_config, error_response = check_store_config(request, 'management/kitchen_template_list.html')
    if error_response:
        return error_response

    templates = KitchenTicketTemplate.objects.filter(
        company=store_config.company
    ).select_related('brand', 'store').order_by('-created_at')

    context_brand_id = request.session.get('context_brand_id')
    if context_brand_id:
        templates = templates.filter(
            Q(brand_id=context_brand_id) | Q(brand__isnull=True)
        )

    context = {
        'store_config': store_config,
        'templates': templates,
        'context_brand_id': context_brand_id,
    }

    return render(request, 'management/kitchen_template_list.html', context)


@manager_required
def kitchen_template_create(request):
    """Create new kitchen ticket template"""
    from apps.kitchen.models import KitchenTicketTemplate

    store_config, error_response = check_store_config(request, 'management/kitchen_template_form.html')
    if error_response:
        return error_response

    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')

    if request.method == 'POST':
        template_name = request.POST.get('template_name', '').strip()
        brand_id = request.POST.get('brand')
        store_id = request.POST.get('store')
        is_active = request.POST.get('is_active') == '1'

        header_line_1 = request.POST.get('header_line_1', '').strip()
        header_line_2 = request.POST.get('header_line_2', '').strip()

        show_bill_number = request.POST.get('show_bill_number') == '1'
        show_table_number = request.POST.get('show_table_number') == '1'
        show_customer_name = request.POST.get('show_customer_name') == '1'
        show_station_name = request.POST.get('show_station_name') == '1'
        show_date_time = request.POST.get('show_date_time') == '1'
        show_item_qty = request.POST.get('show_item_qty') == '1'
        show_item_notes = request.POST.get('show_item_notes') == '1'

        footer_line_1 = request.POST.get('footer_line_1', '').strip()
        footer_line_2 = request.POST.get('footer_line_2', '').strip()

        auto_cut = request.POST.get('auto_cut') == '1'
        feed_lines = request.POST.get('feed_lines', '3')

        if not template_name:
            messages.error(request, 'Template name is required')
        else:
            try:
                template = KitchenTicketTemplate(
                    company=store_config.company,
                    brand_id=brand_id if brand_id else None,
                    store_id=store_id if store_id else None,
                    template_name=template_name,
                    is_active=is_active,
                    header_line_1=header_line_1,
                    header_line_2=header_line_2,
                    show_bill_number=show_bill_number,
                    show_table_number=show_table_number,
                    show_customer_name=show_customer_name,
                    show_station_name=show_station_name,
                    show_date_time=show_date_time,
                    show_item_qty=show_item_qty,
                    show_item_notes=show_item_notes,
                    footer_line_1=footer_line_1,
                    footer_line_2=footer_line_2,
                    auto_cut=auto_cut,
                    feed_lines=int(feed_lines),
                    created_by=request.user
                )
                template.save()
                messages.success(request, f'Kitchen template "{template_name}" created successfully')
                return redirect('management:kitchen_template_list')
            except Exception as e:
                messages.error(request, f'Error creating template: {str(e)}')

    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'is_edit': False,
    }

    return render(request, 'management/kitchen_template_form.html', context)


@manager_required
def kitchen_template_edit(request, template_id):
    """Edit kitchen ticket template"""
    from apps.kitchen.models import KitchenTicketTemplate

    store_config, error_response = check_store_config(request, 'management/kitchen_template_form.html')
    if error_response:
        return error_response

    template = get_object_or_404(KitchenTicketTemplate, id=template_id, company=store_config.company)
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')

    if request.method == 'POST':
        template_name = request.POST.get('template_name', '').strip()
        brand_id = request.POST.get('brand')
        store_id = request.POST.get('store')
        is_active = request.POST.get('is_active') == '1'

        header_line_1 = request.POST.get('header_line_1', '').strip()
        header_line_2 = request.POST.get('header_line_2', '').strip()

        show_bill_number = request.POST.get('show_bill_number') == '1'
        show_table_number = request.POST.get('show_table_number') == '1'
        show_customer_name = request.POST.get('show_customer_name') == '1'
        show_station_name = request.POST.get('show_station_name') == '1'
        show_date_time = request.POST.get('show_date_time') == '1'
        show_item_qty = request.POST.get('show_item_qty') == '1'
        show_item_notes = request.POST.get('show_item_notes') == '1'

        footer_line_1 = request.POST.get('footer_line_1', '').strip()
        footer_line_2 = request.POST.get('footer_line_2', '').strip()

        auto_cut = request.POST.get('auto_cut') == '1'
        feed_lines = request.POST.get('feed_lines', '3')

        if not template_name:
            messages.error(request, 'Template name is required')
        else:
            try:
                template.brand_id = brand_id if brand_id else None
                template.store_id = store_id if store_id else None
                template.template_name = template_name
                template.is_active = is_active
                template.header_line_1 = header_line_1
                template.header_line_2 = header_line_2
                template.show_bill_number = show_bill_number
                template.show_table_number = show_table_number
                template.show_customer_name = show_customer_name
                template.show_station_name = show_station_name
                template.show_date_time = show_date_time
                template.show_item_qty = show_item_qty
                template.show_item_notes = show_item_notes
                template.footer_line_1 = footer_line_1
                template.footer_line_2 = footer_line_2
                template.auto_cut = auto_cut
                template.feed_lines = int(feed_lines)
                template.updated_by = request.user
                template.save()
                messages.success(request, f'Kitchen template "{template_name}" updated successfully')
                return redirect('management:kitchen_template_list')
            except Exception as e:
                messages.error(request, f'Error updating template: {str(e)}')

    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'template': template,
        'is_edit': True,
    }

    return render(request, 'management/kitchen_template_form.html', context)


@manager_required
@require_POST
def kitchen_template_delete(request, template_id):
    """Delete kitchen ticket template"""
    from apps.kitchen.models import KitchenTicketTemplate

    store_config = Store.get_current()
    if not store_config:
        messages.error(request, 'Store configuration not found')
        return redirect('management:kitchen_template_list')

    try:
        template = get_object_or_404(KitchenTicketTemplate, id=template_id, company=store_config.company)
        template_name = template.template_name
        template.delete()
        messages.success(request, f'Kitchen template "{template_name}" deleted successfully')
    except Exception as e:
        messages.error(request, f'Error deleting template: {str(e)}')

    return redirect('management:kitchen_template_list')


@manager_required
@require_POST
def kitchen_template_toggle(request, template_id):
    """Toggle kitchen ticket template active status"""
    from apps.kitchen.models import KitchenTicketTemplate

    store_config = Store.get_current()
    if not store_config:
        return JsonResponse({'success': False, 'error': 'Store not configured'}, status=400)

    try:
        template = get_object_or_404(KitchenTicketTemplate, id=template_id, company=store_config.company)
        template.is_active = not template.is_active
        template.updated_by = request.user
        template.save()

        status = 'activated' if template.is_active else 'deactivated'
        messages.success(request, f'Template "{template.template_name}" {status}!')

        return JsonResponse({
            'success': True,
            'is_active': template.is_active,
            'message': f'Template {status}'
        })

    except Exception as e:
        logger.error(f"Error toggling kitchen template: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================================================
# QRIS Audit Log
# ============================================================================

@manager_required
def qris_audit_log(request):
    """QRIS Audit Log — all QRIS payment events for bank analysis."""
    logs = QRISAuditLog.objects.select_related('bill', 'transaction', 'user').order_by('-created_at')

    # Filter by event type
    event_filter = request.GET.get('event', '')
    if event_filter:
        logs = logs.filter(event=event_filter)

    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        logs = logs.filter(status_after=status_filter)

    # Filter by date range
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)

    # Search by txn_ref or bill number
    search = request.GET.get('search', '')
    if search:
        logs = logs.filter(
            Q(txn_ref__icontains=search) |
            Q(bill__bill_number__icontains=search)
        )

    # Stats (before pagination)
    today = timezone.now().date()
    today_logs = QRISAuditLog.objects.filter(created_at__date=today)
    total_today = today_logs.count()
    paid_today = today_logs.filter(event='payment_confirmed').count()
    errors_today = today_logs.filter(event__in=['error', 'gateway_error', 'gateway_timeout']).count()
    avg_response = today_logs.filter(
        response_time_ms__isnull=False
    ).aggregate(avg=Avg('response_time_ms'))['avg']
    avg_wait = today_logs.filter(
        event='payment_confirmed',
        elapsed_since_create_s__isnull=False
    ).aggregate(avg=Avg('elapsed_since_create_s'))['avg']

    # QRIS transaction summary today
    txn_today = QRISTransaction.objects.filter(created_at__date=today)
    txn_created = txn_today.count()
    txn_paid = txn_today.filter(status='paid').count()
    txn_expired = txn_today.filter(status='expired').count()
    txn_cancelled = txn_today.filter(status='cancelled').count()
    txn_total_paid = txn_today.filter(status='paid').aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # Pagination
    paginator = Paginator(logs, 50)
    page = request.GET.get('page', 1)
    try:
        logs_page = paginator.page(page)
    except PageNotAnInteger:
        logs_page = paginator.page(1)
    except EmptyPage:
        logs_page = paginator.page(paginator.num_pages)

    context = {
        'logs': logs_page,
        'event_filter': event_filter,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search': search,
        # Stats today
        'total_today': total_today,
        'paid_today': paid_today,
        'errors_today': errors_today,
        'avg_response_ms': round(avg_response, 1) if avg_response else None,
        'avg_wait_s': round(avg_wait, 1) if avg_wait else None,
        # Transaction summary
        'txn_created': txn_created,
        'txn_paid': txn_paid,
        'txn_expired': txn_expired,
        'txn_cancelled': txn_cancelled,
        'txn_total_paid': txn_total_paid,
        # Event choices for filter dropdown
        'event_choices': QRISAuditLog.EVENT_CHOICES,
    }
    return render(request, 'management/qris_audit_log.html', context)


# ==========================================
# CUSTOMER DISPLAY PROMO MANAGEMENT
# ==========================================

@manager_required
def display_promo_list(request):
    """List all customer display promos"""
    from apps.core.models import CustomerDisplayPromo
    from datetime import date

    store_config, error_response = check_store_config(request, 'management/display_promo_list.html')
    if error_response:
        return error_response

    promos = CustomerDisplayPromo.objects.filter(
        company=store_config.company
    ).select_related('brand', 'store').order_by('order', '-created_at')

    # Apply brand context filter
    context_brand_id = request.session.get('context_brand_id')
    if context_brand_id:
        promos = promos.filter(
            Q(brand_id=context_brand_id) | Q(brand__isnull=True)
        )

    today = date.today()
    active_count = promos.filter(is_active=True).count()
    inactive_count = promos.filter(is_active=False).count()
    scheduled_count = promos.filter(
        is_active=True, start_date__gt=today
    ).count()

    context = {
        'store_config': store_config,
        'promos': promos,
        'active_count': active_count,
        'inactive_count': inactive_count,
        'scheduled_count': scheduled_count,
    }

    return render(request, 'management/display_promo_list.html', context)


@manager_required
def display_promo_create(request):
    """Create a new customer display promo"""
    from apps.core.models import CustomerDisplayPromo
    from datetime import date

    store_config, error_response = check_store_config(request, 'management/display_promo_form.html')
    if error_response:
        return error_response

    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        badge_text = request.POST.get('badge_text', 'HOT DEAL').strip()
        badge_color = request.POST.get('badge_color', 'red').strip()
        image_url = request.POST.get('image_url', '').strip()
        image_path = request.POST.get('image_path', '').strip()
        emoji_fallback = request.POST.get('emoji_fallback', '🍽️').strip()
        original_price = request.POST.get('original_price', '0')
        promo_price = request.POST.get('promo_price', '0')
        order = request.POST.get('order', '0')
        start_date = request.POST.get('start_date', '').strip() or None
        end_date = request.POST.get('end_date', '').strip() or None
        brand_id = request.POST.get('brand') or None
        store_id = request.POST.get('store') or None
        is_active = request.POST.get('is_active') == '1'

        if not title:
            messages.error(request, 'Promo title is required')
        elif not badge_text:
            messages.error(request, 'Badge text is required')
        else:
            try:
                promo = CustomerDisplayPromo(
                    company=store_config.company,
                    brand_id=brand_id,
                    store_id=store_id,
                    title=title,
                    description=description,
                    badge_text=badge_text,
                    badge_color=badge_color,
                    image_url=image_url,
                    image_path=image_path,
                    emoji_fallback=emoji_fallback or '🍽️',
                    original_price=int(original_price or 0),
                    promo_price=int(promo_price or 0),
                    order=int(order or 0),
                    start_date=start_date,
                    end_date=end_date,
                    is_active=is_active,
                )
                promo.save()
                messages.success(request, f'Promo "{title}" created successfully')
                return redirect('management:display_promo_list')
            except Exception as e:
                messages.error(request, f'Error creating promo: {str(e)}')

    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'is_edit': False,
    }

    return render(request, 'management/display_promo_form.html', context)


@manager_required
def display_promo_edit(request, promo_id):
    """Edit a customer display promo"""
    from apps.core.models import CustomerDisplayPromo

    store_config, error_response = check_store_config(request, 'management/display_promo_form.html')
    if error_response:
        return error_response

    promo = get_object_or_404(CustomerDisplayPromo, id=promo_id, company=store_config.company)
    store_brands = StoreBrand.objects.filter(store=store_config, is_active=True).select_related('brand')

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        badge_text = request.POST.get('badge_text', 'HOT DEAL').strip()
        badge_color = request.POST.get('badge_color', 'red').strip()
        image_url = request.POST.get('image_url', '').strip()
        image_path = request.POST.get('image_path', '').strip()
        emoji_fallback = request.POST.get('emoji_fallback', '🍽️').strip()
        original_price = request.POST.get('original_price', '0')
        promo_price = request.POST.get('promo_price', '0')
        order = request.POST.get('order', '0')
        start_date = request.POST.get('start_date', '').strip() or None
        end_date = request.POST.get('end_date', '').strip() or None
        brand_id = request.POST.get('brand') or None
        store_id = request.POST.get('store') or None
        is_active = request.POST.get('is_active') == '1'

        if not title:
            messages.error(request, 'Promo title is required')
        elif not badge_text:
            messages.error(request, 'Badge text is required')
        else:
            try:
                promo.title = title
                promo.description = description
                promo.badge_text = badge_text
                promo.badge_color = badge_color
                promo.image_url = image_url
                promo.image_path = image_path
                promo.emoji_fallback = emoji_fallback or '🍽️'
                promo.original_price = int(original_price or 0)
                promo.promo_price = int(promo_price or 0)
                promo.order = int(order or 0)
                promo.start_date = start_date
                promo.end_date = end_date
                promo.brand_id = brand_id
                promo.store_id = store_id
                promo.is_active = is_active
                promo.save()
                messages.success(request, f'Promo "{title}" updated successfully')
                return redirect('management:display_promo_list')
            except Exception as e:
                messages.error(request, f'Error updating promo: {str(e)}')

    context = {
        'store_config': store_config,
        'store_brands': store_brands,
        'promo': promo,
        'is_edit': True,
    }

    return render(request, 'management/display_promo_form.html', context)


@manager_required
@require_POST
def display_promo_delete(request, promo_id):
    """Delete a customer display promo"""
    from apps.core.models import CustomerDisplayPromo

    store_config = Store.get_current()
    if not store_config:
        messages.error(request, 'Store configuration not found')
        return redirect('management:display_promo_list')

    try:
        promo = get_object_or_404(CustomerDisplayPromo, id=promo_id, company=store_config.company)
        promo_title = promo.title
        promo.delete()
        messages.success(request, f'Promo "{promo_title}" deleted successfully')
    except Exception as e:
        messages.error(request, f'Error deleting promo: {str(e)}')

    return redirect('management:display_promo_list')


@manager_required
@require_POST
def display_promo_toggle(request, promo_id):
    """Toggle customer display promo active status"""
    from apps.core.models import CustomerDisplayPromo

    store_config = Store.get_current()
    if not store_config:
        return JsonResponse({'success': False, 'error': 'Store not configured'}, status=400)

    try:
        promo = get_object_or_404(CustomerDisplayPromo, id=promo_id, company=store_config.company)
        promo.is_active = not promo.is_active
        promo.save()

        status = 'activated' if promo.is_active else 'deactivated'
        messages.success(request, f'Promo "{promo.title}" {status}!')

        return JsonResponse({
            'success': True,
            'is_active': promo.is_active,
            'message': f'Promo {status}'
        })

    except Exception as e:
        logger.error(f"Error toggling promo: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@manager_required
@require_POST
def display_promo_upload_image(request):
    """Upload promo image to MinIO and return URL"""
    try:
        store_config = Store.get_current()
        if not store_config:
            return JsonResponse({'success': False, 'error': 'Store not configured'}, status=400)

        image = request.FILES.get('image')
        if not image:
            return JsonResponse({'success': False, 'error': 'No image file provided'}, status=400)

        # Validate file size (5MB max)
        if image.size > 5 * 1024 * 1024:
            return JsonResponse({'success': False, 'error': 'File size must be less than 5MB'}, status=400)

        # Validate file type
        allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/webp']
        if image.content_type not in allowed_types:
            return JsonResponse({'success': False, 'error': 'Only JPG, PNG, and WebP files are allowed'}, status=400)

        # Upload to MinIO
        from apps.core.minio_client import upload_to_minio
        import uuid

        file_ext = image.name.split('.')[-1].lower()
        unique_id = str(uuid.uuid4())[:8]
        filename = f"promo_{unique_id}.{file_ext}"

        bucket_name = 'customer-display'
        company_code = store_config.company.code
        object_path = f"{company_code}/promos/{filename}"

        image_url = upload_to_minio(
            bucket_name=bucket_name,
            object_name=object_path,
            file_data=image.read(),
            content_type=image.content_type
        )

        return JsonResponse({
            'success': True,
            'image_url': image_url,
            'image_path': object_path,
        })

    except Exception as e:
        logger.error(f"Error uploading promo image: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
